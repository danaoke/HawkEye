#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 18 22:03:35 2019

@author: jiahexu
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 24 13:43:48 2019

@author: jiahexu
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 23 22:04:27 2019

@author: jiahexu
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Aug 17 13:38:42 2019

@author: jiahexu
"""

# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
#import mcerp3

import tkinter as tk
import numpy as np
import copy
import xlrd
import matplotlib.ticker as mtick
from matplotlib import patches, lines, pyplot
import networkx as nx 
from scipy.stats import skew, kurtosis
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from scipy.stats.stats import pearsonr,spearmanr
from random import *
import pandas as pd
import scipy.stats as ss
from scipy.stats import triang
#from scipy.stats import poisson
import mcerp
from mcerp import PERT,Tri,Normal
import xlsxwriter
from scipy import stats
from scipy.stats import poisson

import seaborn as sns
from tkinter import *
import sys,os

import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
#from tkMessageBox import showinfo
# ALL STATIC VARIABLES ARE DEFINED HERE
TITLE_FONT = ("Times New Roman", 24)
HEADR_FONT = ("Times New Roman", 16)
MESSG_FONT = ("Arial", 10)
VERSION = "2.6.10"
############################################################################################################################################################
def bern_distro(param, itera):
    """ Returns a specified array of values according to Bernoulli distribution.

    :param param: The probability an event will occur (between 0 and 1).
    :param itera: The number of iterations.
    :return: An np.array of n=itera random values
    """
    return np.random.binomial(1, param, itera)


def uni_distro(low, high,itera):
    """ Returns a specified array of values according to Uniform distribution.

        :param param: The probability an event will occur (lambda).
        :param itera: The number of iterations.
        :return: An np.array of n=itera random values
        """
    
    return np.random.uniform(low,high,itera)

def pois_distro(param, itera):
    """ Returns a specified array of values according to Binomial distribution.

    :param param: The probability an event will occur (lambda).
    :param itera: The number of iterations.
    :return: An np.array of n=itera random values
    """
    return poisson.rvs(param, size=itera)
    #return np.random.poisson(param, itera)


def pert_distro(minimum, m, maximum, itera):
    """ Returns a specified array of values according to PERT distribution

    :param minimum: The minimum value parameter.
    :param m: The most likely value parameter.
    :param maximum: The maximum value parameter.
    :param itera: The number of iterations.
    :return: An np.array of n=itera random values.
    """
    mu = (minimum + 4 * m + maximum) / 6
    
   
    # special case if mu == mode
    if mu==m:
        
        alpha = 3
    
    else:
        alpha = (mu-minimum)*(2*m-minimum-maximum)/((m-mu)*(maximum-minimum))

   
    beta = alpha*(maximum-mu)/(mu-minimum)
    

    mcerp.npts = itera
    x = mcerp.np.random.beta(alpha, beta, itera)

    return minimum+x*(maximum-minimum)

def norm_distro(mean, sigma, itera):
    """ Returns a specified array of values according to normal distribution.

    :param mean: The mean value parameter.
    :param sigma: The standard deviation value parameter.
    :param itera: The number of random values to generate.
    :return: An np.array of n=itera random values.
    """
    mcerp.npts = itera
    return mcerp.np.random.normal(mean, sigma, itera)


def tri_distro(a, b, c, itera):
    """ Returns a specified array of values according to triangular distribution.

    :param a: The minimum value parameter.
    :param b: The most likely value parameter.
    :param c: The maximum value parameter.
    :param itera: The number of random values to generate.
    :return: An np.array of n=itera random values.
    """
    mcerp.npts = itera
    return mcerp.np.random.triangular(a, b, c, itera)
def no_distro(para,itera):
    x=np.array([])
    for i in range(0, itera):
      x=np.append(x,para)

    return x

def open_onclick(enter):
    """ Displays a tkinter file dialog box for the user.

    :param enter: A tkinter entry box to put the directory string into.
    """
    ftypes = [('Excel Files', '*.xlsx')]
    dlg = filedialog.askopenfilename(initialdir="/Users/Documents", title="Select File", filetypes=ftypes)
    enter.delete(0, 1000)
    enter.insert(0, dlg)
def make_table(data, string_1,unit='$',unitloc='l'):
    """ Displays the statistics of the generated data in a table format.
    :param data: The data.
    :param string_1: The title of the table.
    """
    window = tk.Toplevel()
    window.resizable(width=True, height=True)
    window.title("Table")
    main = tk.Label(window, text=string_1, font=('Calibri', 24, 'bold') )
    main.pack()
    style = ttk.Style()
    style.configure("mystyle.Treeview", font=('Calibri', 15),rowheight=40)  # Modify the font of the body
    style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
    tree = ttk.Treeview(window, selectmode="extended", show='tree',height="30", style="mystyle.Treeview")
    tree["columns"] = ("stat", "sdata", "perc", "pdata")
    tree.column("#0", stretch=tk.YES, minwidth=100, width=200)
    tree.column("stat",stretch=tk.YES, minwidth=100, width=200)
    tree.column("sdata", stretch=tk.YES, minwidth=100, width=200)
    tree.column("perc", stretch=tk.YES, minwidth=100, width=200)
    tree.column("pdata", stretch=tk.YES, minwidth=100, width=200)
    tree.heading("stat", text="Statistics")
    tree.heading("perc", text="Percentiles")
    s = float(skew(data))
    k = kurtosis(data)
    
    percentage = [5,10,25,50,75,90,95]
    percentiles = np.percentile(data,[5,10,25,50,75,90,95])
    
    functions = [['Minimum','min'],
                 ['Maximum','max'],
                 ['Mean','mean'],
                 ['Median','median'],
                 ['Std. Deviation','std'],
                 ['Skewness','skew'],
                 ['Kurtosis','kurtosis']]
    if unitloc == 'l':
        for i in range(5):
            tree.insert('', i, values=[functions[i][0], f'{unit}{int(getattr(np,functions[i][1])(data))}',
                                       f'{percentage[i]}th Percentile', f'{unit}{int(percentiles[i])}'])
        
        tree.insert('', 5, values=['Skewness',  str(round(s, 2)),
                               '90th Percentile', f'{unit}{int(percentiles[5])}'])
        tree.insert('', 6, values=['Kurtosis', str(round(k, 2)),
                                   '95th Percentile', f'{unit}{int(percentiles[6])}'])
        
    elif unitloc == 'r':
        for i in range(5):
            tree.insert('', i, values=[functions[i][0], f'{int(getattr(np,functions[i][1])(data))}{unit}',
                                       f'{percentage[i]}th Percentile', f'{int(percentiles[i])}{unit}'])
        
        tree.insert('', 5, values=['Skewness',  str(round(s, 2)),
                               '90th Percentile', f'{int(percentiles[5])}{unit}'])
        tree.insert('', 6, values=['Kurtosis', str(round(k, 2)),
                                   '95th Percentile', f'{int(percentiles[6])}{unit}'])
       
    '''    
    [, '$'+str(int(np.(data))),
                               $'+str(int(np.percentile(data, 5)))])
    tree.insert('', 1, values=[, '$'+str(int(np.max(data))),
                               '10th Percentile', '$'+str(int(np.percentile(data, 10)))])
    tree.insert('', 2, values=['Mean', '$'+str(int(np.mean(data))),
                               '25th Percentile', '$'+str(int(np.percentile(data, 25)))])
    tree.insert('', 3, values=['', '$'+str(int(np.median(data))),
                               '50th Percentile', '$'+str(int())])
    tree.insert('', 4, values=['Std. Deviation', '$'+str(int(np.std(data))),
                               '75th Percentile', '$'+str(int(np.percentile(data, 75)))])
    tree.insert('', 5, values=['Skewness',  str(round(s, 2)),
                               '90th Percentile', '$'+str(int(np.percentile(data, 90)))])
    tree.insert('', 6, values=['Kurtosis', str(round(k, 2)),
                               '95th Percentile', '$'+str(int(np.percentile(data, 95)))])
    '''
    tree.pack()    
def make_graph( data, string_1, i, j):

    """ Displays the primary data in graph from using matplotlib
    :param b: The number of bins to sort the data into.
    :param data: The data generated.
    :param string_1: The title of the graph.
    :param i: Boolean. Display CDF?
    :param j: Boolean. Display PDF?
    """
    if j > 0 and i == 0:
        pyplot.clf()
        host = pyplot.subplot(111)
        host.set_xlabel(string_1, fontsize=12)
        host.set_ylabel("Relative Frequency, %")
        host.hist(data, bins=50, alpha=0.6, color='blue')
        host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    elif i > 0 and j == 0:
        pyplot.clf()
        host = pyplot.subplot(111)
        host.set_xlabel(string_1, fontsize=12)
        host.set_ylabel("Cumulative Probability", fontsize=12)
        #host.hist(data, bins=1000, density=True, cumulative=True, histtype='step', color='blue')
        ser = pd.Series(data)
        ax = ser.hist(bins=1000, density=True, cumulative=True, histtype='step', color='blue',grid=False)
        # how to delete the vertical line for cdf graph
        poly = ax.findobj(pyplot.Polygon)[0]
        vertices = poly.get_path().vertices
        # Keep everything above y == 0. You can define this mask however
        # you need, if you want to be more careful in your selection.
        keep = vertices[:, 1] > 0
        # Construct new polygon from these "good" vertices
        new_poly = pyplot.Polygon(vertices[keep], closed=False, fill=False, edgecolor=poly.get_edgecolor(),
                                  linewidth=poly.get_linewidth())
        poly.set_visible(False)
        ax.add_artist(new_poly)
    else:
        #pyplot.style.use('classic')
        pyplot.clf()
        host = pyplot.subplot(111)
        par1 = host.twinx()
        host.set_xlabel(string_1, fontsize=12)
        host.set_ylabel("Relative Frequency, %", fontsize=12)
        par1.set_ylabel("Cumulative Probability", fontsize=12)
        host.hist(data, bins=50, alpha=0.6, label="PDF", color='blue')
        #par1.hist(data, bins=1000, density=True, cumulative=True, histtype='step', color='blue')
        ser = pd.Series(data)
        bx = ser.hist(bins=1000, density=True, cumulative=True, histtype='step', color='blue',grid=False)
        # how to delete the vertical line for cdf graph
        poly = bx.findobj(pyplot.Polygon)[0]
        vertices = poly.get_path().vertices
        # Keep everything above y == 0. You can define this mask however
        keep = vertices[:, 1] > 0
        # Construct new polygon from these "good" vertices
        new_poly = pyplot.Polygon(vertices[keep], closed=False, fill=False, edgecolor=poly.get_edgecolor(),
                                  linewidth=poly.get_linewidth())
        poly.set_visible(False)
        bx.add_artist(new_poly)
        blue_line = lines.Line2D([], [], color="blue", label="CDF", linewidth=2.0)
        blue_patch = patches.Patch(alpha=0.6, color="blue", label="PDF")
        pyplot.legend(handles=[blue_line, blue_patch], loc='upper left')
        host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    fmt = '${x:,.0f}'
    tick = mtick.StrMethodFormatter(fmt)
    host.xaxis.set_major_formatter(tick)
    pyplot.draw()
    pyplot.show()
def make_risk_graph(data, temp, string_1, i, j):
        
    #temp = data + data2
    if i > 0 and j == 0:
        pyplot.clf()
        host = pyplot.subplot(111)
        par1 = host.twinx()
        host.set_xlabel(string_1, fontsize=12)
        host.set_ylabel("Cumulative Probability", fontsize=12)
        host.hist(temp, 1000, density=True, cumulative=True, histtype='step', color='red', linewidth=2)
        par1.hist(data, 1000, density=True, cumulative=True, histtype='step', color='blue', linewidth=2)
        blue_line = lines.Line2D([], [], color="blue", label="Without Risk", linewidth=2)
        risk_line = lines.Line2D([], [], color="red", label="With Risk", linewidth=2)
        pyplot.legend(handles=[risk_line, blue_line], loc='upper left')
        fmt = '${x:,.0f}'
        tick = mtick.StrMethodFormatter(fmt)
        host.xaxis.set_major_formatter(tick)

        def fix_hist_step_vertical_line_at_end(ax):
            axpolygons = [poly for poly in ax.get_children() if isinstance(poly, patches.Polygon)]
            for poly in axpolygons:
                poly.set_xy(poly.get_xy()[:-1])
        fix_hist_step_vertical_line_at_end(host)
        fix_hist_step_vertical_line_at_end(par1)
    if j > 0 and i == 0:
        pyplot.style.use('classic')
        pyplot.clf()
        host = pyplot.subplot(111)
        host.set_xlabel(string_1, fontsize=12)
        host.set_ylabel("Relative Frequency, %", fontsize=12)
        host.hist(data, bins=50, alpha=0.6, label="Without Risk", color='blue')
        host.hist(temp, bins=50, alpha=0.6, label="With Risk", color='red')
        risk_patch = patches.Patch(alpha=0.6, color="red", label="With Risk")
        blue_patch = patches.Patch(alpha=0.6, color="blue", label="Without Risk")
        pyplot.legend(handles=[risk_patch, blue_patch])
        host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
        fmt = '${x:,.0f}'
        tick = mtick.StrMethodFormatter(fmt)
        host.xaxis.set_major_formatter(tick)
    pyplot.draw()
    pyplot.show()





def read_NPV_excel(director):
    book=xlrd.open_workbook(director)
    sheet=book.sheet_by_name("Sheet1")
    df=int(sheet.cell_value(1,0))
    YearNo = []
    Capital_Cost_Dis = []
    Capital_Cost_Max = []
    Capital_Cost_Most = []
    Capital_Cost_Min = []
    Capital_Cost_sigma = []
    Capital_Cost_Event = []
    Annual_Rev_Dis = []
    Annual_Rev_Max = []
    Annual_Rev_Min = []
    Annual_Rev_Most = []
    Annual_Rev_sigma = []
    Annual_Rev_Event = []
    Annual_Cost_Dis = []
    Annual_Cost_Max = []
    Annual_Cost_Min = []
    Annual_Cost_Most = []
    Annual_Cost_sigma = []
    Annual_Cost_Event = []
    Discount_Rate_Most = []

    for i in range(0,df):
        YearNo.append(sheet.cell_value(i+2,0))
        Capital_Cost_Dis.append(sheet.cell_value(i+2,6))
        Capital_Cost_Max.append(sheet.cell_value(i+2,3))
        Capital_Cost_Most.append(sheet.cell_value(i+2,2))
        Capital_Cost_Min.append(sheet.cell_value(i+2,1))
        Capital_Cost_sigma.append(sheet.cell_value(i+2,4))
        Capital_Cost_Event.append(sheet.cell_value(i+2,5))
        Annual_Rev_Dis.append(sheet.cell_value(i+2,12))
        Annual_Rev_Max.append(sheet.cell_value(i+2,9))
        Annual_Rev_Min.append(sheet.cell_value(i+2,7))
        Annual_Rev_Most.append(sheet.cell_value(i+2,8))
        Annual_Rev_sigma.append(sheet.cell_value(i+2,10))
        Annual_Rev_Event.append(sheet.cell_value(i+2,11))
        Annual_Cost_Dis.append(sheet.cell_value(i+2,18))
        Annual_Cost_Max.append(sheet.cell_value(i+2,15))
        Annual_Cost_Min.append(sheet.cell_value(i+2,13))
        Annual_Cost_Most.append(sheet.cell_value(i+2,14))
        Annual_Cost_sigma.append(sheet.cell_value(i+2,16))
        Annual_Cost_Event.append(sheet.cell_value(i+2,17))
        Discount_Rate_Most.append(sheet.cell_value(i+2,19))
    Capital = Capital_Cost_Dis, Capital_Cost_Max, Capital_Cost_Most, Capital_Cost_Min, Capital_Cost_sigma ,Capital_Cost_Event
    Annual_Rev = Annual_Rev_Dis, Annual_Rev_Max, Annual_Rev_Min, Annual_Rev_Most, Annual_Rev_sigma, Annual_Rev_Event
    Annual_Cost= Annual_Cost_Dis, Annual_Cost_Max, Annual_Cost_Min, Annual_Cost_Most, Annual_Cost_sigma, Annual_Cost_Event
    return  YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most

def read_cost_excel(director):
    book=xlrd.open_workbook(director)
    sheet=book.sheet_by_index(0)
    if sheet.cell_value(0,0)=='':
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:df=int(sheet.cell_value(0,0))
    if sheet.cell_value(df+3,0)=='':
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:dt1=int(sheet.cell_value(df+3,0))
        
    if sheet.cell_value(df+3+dt1+1,0)=='':
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:dt2=int(sheet.cell_value(df+3+dt1+1,0))
    
    task=[]
    simul=[]
    distro=[]
    most=[]
    minn=[]
    maxn=[]
    std=[]
    event=[]
    
    risk_bin = []  # an array w/risk bernoulli probability
    risk_dis = []  # an array w/risk distribution
    risk_tsk = []  # an array w/risk task name
    risk_sim = []  # an array w/risk distribution info
    risk_most = []
    risk_min = []
    risk_max = []
    risk_std = []
    risk_event = []

    for i in range(0,df):
        if sheet.cell_value(i+1,1)=='':
            messagebox.showerror("Error", "Missing Task Name or Input Read Number out of Data Range.")
            return
        else:
            task.append(sheet.cell_value(i+1,1))
        if sheet.cell_value(i+1,3)=='Poisson' or 'Pert' or 'Triangular' or 'Normal' or '':
            
            distro.append(sheet.cell_value(i+1,3))
        else:
            messagebox.showerror("Error", "Duration Missing Spealling/ Formating/ Input.")
            return
        most.append(sheet.cell_value(i+1,2))

        minn.append(sheet.cell_value(i+1,4))

        maxn.append(sheet.cell_value(i+1,5))

        std.append(sheet.cell_value(i+1,6))

        event.append(sheet.cell_value(i+1,7))


        if distro[i] == 'Poisson':
            if minn[i]=='' or most[i]=='' or maxn[i]=='' or event[i]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return
            else:
                simul.append(event[i])

        elif distro[i] == 'Pert':
            if minn[i]=='' or most[i]=='' or maxn[i]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return
            else:
                simul.append([minn[i], most[i], maxn[i]])

        elif distro[i] == 'Triangular':
            if minn[i]=='' or most[i]=='' or maxn[i]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return
            else:
                simul.append([minn[i], most[i], maxn[i]])

        elif distro[i] == 'Normal':
            if  std[i]=='' or most[i]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return
            else:
                simul.append([most[i], std[i]])

        else:
            if most[i]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return
            else:
                simul.append(most[i])


    for j in range(df+3,df+3+dt1):
        if sheet.cell_value(j+1,1)=='':
            messagebox.showerror("Error", "Missing Risk Task Name or Input Read Number out of Data Range.")
            return
        else:
            risk_tsk.append(sheet.cell_value(j+1,1))
        if sheet.cell_value(j+1,3)=='Poisson' or 'Pert' or 'Triangular' or 'Normal' or '':
            risk_dis.append(sheet.cell_value(j+1,3))
        else:
            messagebox.showerror("Error", "Risk Duration Missing Spealling/ Formating/ Input.")
            return

        risk_most.append(sheet.cell_value(j+1,2))

        risk_min.append(sheet.cell_value(j+1,4))

        risk_max.append(sheet.cell_value(j+1,5))

        risk_std.append(sheet.cell_value(j+1,6))

        risk_event.append(sheet.cell_value(j+1,8))

        risk_bin.append(sheet.cell_value(j+1,7))
    
    for j in range(df+3+dt1+1,df+3+dt1+1+dt2):
        if sheet.cell_value(j+1,1)=='':
            messagebox.showerror("Error", "Missing Risk Task Name or Input Read Number out of Data Range.")
            return
        else:
        
            risk_tsk.append(sheet.cell_value(j+1,1))
        if sheet.cell_value(i+1,3)=='Poisson' or 'Pert' or 'Triangular' or 'Normal' or '':
            risk_dis.append(sheet.cell_value(j+1,3))
        else:
            messagebox.showerror("Error", "Risk Duration Missing Spealling/ Formating/ Input.")
            return
        risk_most.append(sheet.cell_value(j+1,2))

        risk_min.append(sheet.cell_value(j+1,4))

        risk_max.append(sheet.cell_value(j+1,5))

        risk_std.append(sheet.cell_value(j+1,6))

        risk_event.append(sheet.cell_value(j+1,8))

        risk_bin.append(sheet.cell_value(j+1,7))
    
    for j in range(0,len(risk_dis)):

        if risk_dis[j] == 'Poisson':
            if risk_min[j]=='' or risk_most[j]=='' or risk_max[j]=='' or risk_event[j]=='':
               messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
               return 
            else:
                risk_sim.append([risk_min[j], risk_most[j], risk_max[j],risk_event[j]])

        elif risk_dis[j] == 'Pert':
            if risk_min[j]=='' or risk_most[j]=='' or risk_max[j]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return 
            else:
                risk_sim.append([risk_min[j], risk_most[j], risk_max[j]])

        elif risk_dis[j] == 'Triangular':
            if risk_min[j]=='' or risk_most[j]=='' or risk_max[j]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return 
            else:

                risk_sim.append([risk_min[j], risk_most[j], risk_max[j]])

        elif risk_dis[j] == 'Normal':
            if risk_most[j]=='' or risk_std[j]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return 
            else:

                risk_sim.append([risk_most[j], risk_std[j]])

        else:
            if risk_most[j]=='':
                messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                return 
            else:risk_sim.append(risk_most[j])
            
    return simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most,minn,maxn,risk_min,risk_max

def upload_file(director):
    book=xlrd.open_workbook(director)
    sheet=book.sheet_by_index(0)
    if sheet.cell_value(0,0)=='':
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:
        t=int(sheet.cell_value(0,0))
    if sheet.cell_value(t+3,0)=='' :
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:
        rt=int(sheet.cell_value(t+3,0))
    if sheet.cell_value(t+3+rt+1,0)=='':
        messagebox.showerror("Error", "Missing Range Number Input.")
        return
    else:   
        ro=int(sheet.cell_value(t+3+rt+1,0))
    
    name=[]
        
    for i in range(0,t):
        if sheet.cell_value(i+1,1)=='':
            messagebox.showerror("Error", "Missing Task Name or Input Read Number out of Data Range.")
            return
        else:
            
            name.append(sheet.cell_value(i+1,1)) #task name
    task=[]
    for i in range(0,t):
        if sheet.cell_value(i+1,2)=='':
            messagebox.showerror("Error", "Duration Missing Spealling/ Formating/ Input.")
            return
        else:
            
            task.append(sheet.cell_value(i+1,2))  #task duration
    distr=[]
    for i in range(0,t):
        if sheet.cell_value(i+1,3)=='' or 'Poisson' or 'Pert' or 'Triangular' or 'Normal':
            distr.append(sheet.cell_value(i+1,3))  #task distribution    
        else:
            messagebox.showerror("Error", "Distribution Missing Spealling/ Formating/ Input.")
            return
           
    mim=[]
    for i in range(0,t):
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,4)=='':
                 messagebox.showerror("Error", "Minimum Missing Spealling/ Formating/ Input.")
                 return
            else:
                 mim.append(sheet.cell_value(i+1,4)) #min
        else:
            mim.append(sheet.cell_value(i+1,4)) #min
    maxm=[]
    for i in range(0,t):
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,5)=='':
                 messagebox.showerror("Error", "Maximum Missing Spealling/ Formating/ Input.")
                 return
            else:
                 maxm.append(sheet.cell_value(i+1,5))
        else:
            maxm.append(sheet.cell_value(i+1,5))
   
    predss_check = []
    for i in range(0, t):
        predss_check.append(sheet.cell_value(i + 1, 6)) #predssor
            
    predss = []
    predss.append([0])
    
    for i in range(0, t):
        if isinstance(predss_check[i], str):
            a = predss_check[i].split(" ")
            b=[]
            if len(a)>=3:
                messagebox.showerror("Error", "Predss Input out of range.")
                return 
            for j in range(len(a)):
                
                if a[j].isdigit()== True:
                    b.append(int(a[j]))
                else:
                    messagebox.showerror("Error", "Predss Missing Spealling/ Formating/ Input.")
                    return 
                
            
            predss.append(b)
        else:
            
            predss.append([int(predss_check[i])]) #predssor
            
        
    
    name_rt=[]
    mod_rt=[]
    distr_rt=[]
    task_rt=[]
    mim_rt=[]
    maxm_rt=[]
    event_rt=[]
    for i in range(t+3,t+3+rt):
        if sheet.cell_value(i+1,1)=='':
             messagebox.showerror("Error", "Threat Name Missing or Out of range.")
             return 
        else:
            name_rt.append(sheet.cell_value(i+1,1))
        if sheet.cell_value(i+1,2)=='':
             messagebox.showerror("Error", "Threat Duration Missing Spealling/ Formating/ Input.")
             return 
        else:
            mod_rt.append(sheet.cell_value(i+1,2))
        
        if sheet.cell_value(i+1,3)=='' or 'Poisson' or 'Pert' or 'Triangular' or 'Normal':
             distr_rt.append(sheet.cell_value(i+1,3))
             
        else:
            messagebox.showerror("Error", "Threat Distribution Missing Spealling/ Formating/ Input.")
            return 
        
        if sheet.cell_value(i+1,6)=='' or isinstance(sheet.cell_value(i+1,6),str)==True :
             messagebox.showerror("Error", "Threat Predssor Missing Spealling/ Formating/ Input.")
             return 
        else:
            
            task_rt.append(sheet.cell_value(i+1,6))
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,4)=='':
                
                 messagebox.showerror("Error", "Threat Minimum Missing Spealling/ Formating/ Input.")
                 return
            else:
                mim_rt.append(sheet.cell_value(i+1,4))    
        else:
            mim_rt.append(sheet.cell_value(i+1,4))
            
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,5)=='':
                 messagebox.showerror("Error", "Threat Maximum Missing Spealling/ Formating/ Input.")
                 return 
            else:
                maxm_rt.append(sheet.cell_value(i+1,5))
        else:
            maxm_rt.append(sheet.cell_value(i+1,5))
        if sheet.cell_value(i+1,3)=='Poisson' and sheet.cell_value(i+1,8)=='':
             messagebox.showerror("Error", "Threat Events Missing Spealling/ Formating/ Input.")
             return 
        else:
            event_rt.append(sheet.cell_value(i+1,8))
        
    name_ro=[]
    mod_ro=[]
    distr_ro=[]
    task_ro=[]
    mim_ro=[]
    maxm_ro=[]
    event_ro=[]
    for i in range(t+3+rt+1,t+3+rt+1+ro):
        if sheet.cell_value(i+1,1)=='':
             messagebox.showerror("Error", "Opportunity Name Missing or Out of range.")
             return 
        else:
            name_ro.append(sheet.cell_value(i+1,1))
        
        if sheet.cell_value(i+1,2)=='':
             messagebox.showerror("Error", "Opportunity Duration Missing Spealling/ Formating/ Input.")
             return 
        else:
            mod_ro.append(sheet.cell_value(i+1,2))
        if sheet.cell_value(i+1,3)=='' or 'Poisson' or 'Pert' or 'Triangular' or 'Normal':
        
            distr_ro.append(sheet.cell_value(i+1,3))
        else:
            messagebox.showerror("Error", "Opportunity Distribution Missing Spealling/ Formating/ Input.")
            return 
        if sheet.cell_value(i+1,6)=='' or isinstance(sheet.cell_value(i+1,6),str)==True :
             messagebox.showerror("Error", "Opportunity Predssor Missing Spealling/ Formating/ Input.")
             return 
        else:
            task_ro.append(sheet.cell_value(i+1,6))
        
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,4)=='':
                 messagebox.showerror("Error", "Opportunity Minimum Missing Spealling/ Formating/ Input.")
                 return 
            else:
                 mim_ro.append(sheet.cell_value(i+1,4))
        else:
            mim_ro.append(sheet.cell_value(i+1,4))
        
        if sheet.cell_value(i+1,3)!='' and sheet.cell_value(i+1,3)!='Normal':
            if sheet.cell_value(i+1,5)=='':
                 messagebox.showerror("Error", "Opportunity Maximum Missing Spealling/ Formating/ Input.")
                 return 
            else:
                maxm_ro.append(sheet.cell_value(i+1,5))
        else:
            maxm_ro.append(sheet.cell_value(i+1,5))
        
        if sheet.cell_value(i+1,3)=='Poisson' and sheet.cell_value(i+1,8)=='':
             messagebox.showerror("Error", "Opportunity Events Missing Spealling/ Formating/ Input.")
             return 
        else:
            event_ro.append(sheet.cell_value(i+1,8))
        
    return name,task,distr,predss,t,mim,maxm,name_rt,mod_rt,distr_rt,task_rt,rt,mim_rt,maxm_rt,event_rt,name_ro,mod_ro,distr_ro,task_ro,ro,mim_ro,maxm_ro,event_ro
        





#simul, distro, task,risk_sim,risk_dis,risk_tsk,risk_bin=read_cost_excel('sample_projectcost - Copy-1.xlsx')
############################################################################################################################################################
#INTERFACE FOR HAWKEYE

class App(tk.Tk):
    """ This is the base of the program, its structure is what allows for multiple 'pages'.
    Note: Every 'page' manipulated is defined as its own unique class below this one.

    """
    def __init__(self, *args, **kwargs):
        """ Constructor function.

        :param args: Inherited from tk.Tk
        :param kwargs: Inherited from tk.Tk
        """
        tk.Tk.__init__(self, *args, **kwargs)

        # add a window icon (*must* be .ico)
        # tk.Tk.iconbitmap(self, default="hawkeye.ico")
        if getattr(sys, 'frozen', False):
            application_path = sys._MEIPASS
        elif __file__:
            application_path = os.path.dirname(__file__)
            
        self.iconbitmap(default=os.path.join(application_path,'hawkeye.ico'))
        
        # add a window title
        tk.Tk.wm_title(self, "HawkEye "+VERSION)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        # all new page classes *must* be added to this loop
        for F in (StartPage, ChoosePage,Mainpage,OneDimension,OneDimension1,NetPresentValue,NPVSimul,ProjectCost,PCRiskSimul,ProSchedule,ProScheduleOptions,
                  Optimization,Optimization_Simulation,PortOOptions,Decisiontree,DecisionSolution):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

    # this function "raises" the given frame to the top of the stack
    def show_frame(self, cont):
        """ 'Raises' the given frame to the top of the stack.

        :param cont: The frame being manipulated.
        :return:
        """
        frame = self.frames[cont]
        frame.tkraise()


# ALL PAGE CLASSES SHOULD BE DEFINED BELOW HERE
# a note on page formatting:
# in order to maintain a good UI/UX, all organization is handled by
# "interior" frames that organize the widgets appropriately, i.e.
# page frame <== interior frame <== widget


class StartPage(tk.Frame):
    """ This is the first page of the program that the user sees.

    """
    def __init__(self, parent, controller):
        """ Class constructor.

        :param parent: The parent frame of this page.
        :param controller: The controller runs between all pages and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)

        title_frame = tk.Frame(self)
        title_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)

        label_1 = tk.Label(title_frame, text="HawkEYE "+VERSION, font=TITLE_FONT)
        label_1.pack(padx=15, pady=15)
        label_2 = tk.Label(center_frame, text="User Agreement", font=HEADR_FONT)
        label_2.pack()
        message_1 = tk.Message(center_frame, text="This is an alpha program. As such we will not be held liable for"
                               "any actions, taken on the part of the user, that are based upon the data. "
                               "USE AT YOUR OWN RISK.", justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        button_1 = ttk.Button(button_frame, text="Accept", command=lambda: controller.show_frame(ChoosePage))
        button_1.pack(side="right", padx=5)
        button_2 = ttk.Button(button_frame, text="Decline", command=controller.destroy)
        button_2.pack(side="right", padx=5)

class ChoosePage(tk.Frame):
    """ This page allows the user to choose the appropriate model that they would like to use.

    """
    def __init__(self, parent, controller):
        """ Class constructor.

        :param parent: The parent frame of the page.
        :param controller: The controller runs between all pages and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)

        # this chunk resets the global variables if you return to the choose page
        
        

        top_frame = tk.Frame(self)
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)

        label_1 = tk.Label(center_frame, text="Available Simulation Models", font=HEADR_FONT)
        label_1.pack()
        message_1 = tk.Message(center_frame, text="Choose the simulation you'd like to run from the list below.\n",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()

        button_one = ttk.Button(center_frame, text="Sensitivity Analysis", width=30,
                                command=lambda: controller.show_frame(Mainpage))
        button_one.pack()
        button_npv = ttk.Button(center_frame, text="Net Present Value", width=30,
                                command=lambda: controller.show_frame(NetPresentValue))
        button_npv.pack()
        button_pcost = ttk.Button(center_frame, text="Project Cost", width=30,
                                  command=lambda: controller.show_frame(ProjectCost))
        button_pcost.pack()
        button_sched = ttk.Button(center_frame, text="Project Schedule", width=30,
                                  command=lambda: controller.show_frame(ProSchedule))
        button_sched.pack()
        button_port = ttk.Button(center_frame, text='Portfolio Optimization', width=30,
                                 command=lambda: controller.show_frame(Optimization))
        button_port.pack()
        
        button_dece = ttk.Button(center_frame, text='Decision Tree', width=30,
                                 command=lambda: controller.show_frame(Decisiontree))
        button_dece.pack()
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: controller.show_frame(StartPage))
        button_1.pack(side="right", padx=5)

class Mainpage(tk.Frame):

    """ This page allows the user to upload an excel file for one dimensional analysis.



    """

    def __init__(self, parent, controller):

        """

        Class constructor.

        :param parent: The parent frame of the page.

        :param controller: The controller runs between all pages and allows them to call the show_page function.

        """

        tk.Frame.__init__(self, parent)
        



        def get_string():

            """ Retrieves the directory string from the message box widget then displays ODimDefine.



            """


            controller.show_frame(OneDimension)



        def get_string2():

            """ Retrieves the directory string from the message box widget then displays ODimDefine.



            """


            controller.show_frame(OneDimension1)



        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", expand=True, padx=10, pady=5, anchor='s')

        center_frame2 = tk.Frame(self)

        center_frame2.pack(side="top", expand=True, padx=10, pady=5, anchor="n")

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Sensitivity Analysis | Type Select", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the model type to analyze.\n",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()

        #entry_1 = ttk.Entry(center_frame, width=50)

        #entry_1.pack(side="left")

        #button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))

        #button_2.pack(side="left", pady=5)

        button_3 = ttk.Button(center_frame2, text="NPV", command=lambda: get_string())

        button_3.pack(side="right", pady=5)

        button_4 = ttk.Button(center_frame2, text="Project Cost", command=lambda: get_string2())

        button_4.pack(side="right", pady=5)

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: controller.show_frame(ChoosePage))

        button_1.pack(side="right", padx=5)

class OneDimension(tk.Frame):

    """ This page allows the user to upload an excel file for one dimensional analysis.



    """

    def __init__(self, parent, controller):

        """

        Class constructor.

        :param parent: The parent frame of the page.

        :param controller: The controller runs between all pages and allows them to call the show_page function.

        """

        tk.Frame.__init__(self, parent)
        global directory
        directory =""



        def get_string():
            """ Retrieves the directory string from the entry widget.

            """
            global directory
            directory = entry_1.get()
           
            if directory == "":
                messagebox.showerror("Error", "An error has occurred in your File.")
            else:
                 global  YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
                 YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most =read_NPV_excel(directory)
                 make_tornado(entry_change.get())
                 
        def clear_val():
            
            global directory,YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
            
            if directory!="":
                YearNo=[]
                Capital=[]
                Annual_Rev=[]
                Annual_Cost=[]
                Discount_Rate_Most=[]
                del directory,YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
            
            controller.show_frame(Mainpage)
            entry_1.delete(0, 'end')
    
       
        def sens_npv(YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most,change):
        
            #global title,yahoo_npv
        
            rate = Discount_Rate_Most[1]
            Capital_Cost = []
            Capital_Cost.append(Capital[2][0] * (1 - float(change)))
            Capital_Cost.append(Capital[2][0])
            Capital_Cost.append(Capital[2][0] * (1 + float(change)))
            print("Cpaital_Cost_list:" + str(Capital_Cost))
        
            Annual_Rev_min = []
            for i in range(0, len(Annual_Rev[3])):
                if Annual_Rev[3][i]=='':
                    Annual_Rev[3][i]=0
                if Annual_Cost[3][i]=='':
                    Annual_Cost[3][i]=0
            Annual_Rev_min.append(np.array(Annual_Rev[3]) * (1 - float(change)))
            Annual_Rev_mean = Annual_Rev[3]
            Annual_Rev_max = []
            Annual_Rev_max.append(np.array(Annual_Rev[3]) * (1 + float(change)))
            print("Annual_Rev_min:" + str(Annual_Rev_min))
            print("Annual_Rev_mean:" + str(Annual_Rev_mean))
            print("Annual_Rev_max:" + str(Annual_Rev_max))
        
            Annual_Cos_min = []
            Annual_Cos_min.append(np.array(Annual_Cost[3]) * (1 - float(change)))
            Annual_Cos_mean = Annual_Cost[3]
            Annual_Cos_max = []
            Annual_Cos_max.append(np.array(Annual_Cost[3]) * (1 + float(change)))
            print("Annual_Cos_min:" + str(Annual_Cos_min))
            print("Annual_Cos_mean:" + str(Annual_Cos_mean))
            print("Annual_Cos_max:" + str(Annual_Cos_max))
        
            discf = []
        
            for i in range(1, len(YearNo)):
                discf.append(1 / ((1 + rate) ** i))
            print("discf:" + str(discf))
            ############################################################################
            Capital_npv = []
            Net_capital = np.array(Annual_Rev[3]) - np.array(Annual_Cost[3])
            with_discf_capital = Net_capital[1:] * discf
            sum_with_capital = np.sum(with_discf_capital)
            for i in range(0, len(Capital_Cost)):
                Capital_npv.append(sum_with_capital - Capital_Cost[i])
            print("Capital_npv:" + str(Capital_npv))
            ############################################################################
            Annual_Rev_min_npv = []
            Annual_Rev_mean_copy = np.copy(Annual_Rev_mean)
            
            for i in range(1, len(Annual_Rev_mean)):
                Annual_Rev_mean[i] = Annual_Rev_min[0][i]
        
               
                Net_rev_min = np.array(Annual_Rev_mean[1:]) - np.array(Annual_Cos_mean[1:])
                
        
                with_discf_rev = Net_rev_min * discf
               
                sum_with_rev_min = np.sum(with_discf_rev)
                print(sum_with_rev_min)
                Annual_Rev_min_npv.append(sum_with_rev_min - Capital[2][0])
               
                Annual_Rev_mean[i] = Annual_Rev_mean_copy[i]
                
        
            print("Annual_Rev_min_npv:" + str(Annual_Rev_min_npv))
        
            Annual_Rev_mean_npv = []
            Net_rev_mean = np.array(Annual_Rev[3]) - np.array(Annual_Cost[3])
            with_discf_rev_mean = np.sum(Net_rev_mean[1:] * discf)
            Annual_Rev_mean_npv.append(with_discf_rev_mean - Capital[2][0])
            print("Annual_Rev_mean_npv:" + str(Annual_Rev_mean_npv))
        
            Annual_Rev_max_npv = []
            Annual_Rev_mean_copy = np.copy(Annual_Rev_mean)
            for i in range(1, len(Annual_Rev_mean)):
                Annual_Rev_mean[i] = Annual_Rev_max[0][i]
        
                # print (Annual_Rev_mean)
                Net_rev_max = np.array(Annual_Rev_mean[1:]) - np.array(Annual_Cos_mean[1:])
                # print (Net_rev)
        
                with_discf_rev = Net_rev_max * discf
                # print (with_discf_rev)
                sum_with_rev_max = np.sum(with_discf_rev)
        
                Annual_Rev_max_npv.append(sum_with_rev_max - Capital[2][0])
                # print (Annual_Rev_min_npv)
                Annual_Rev_mean[i] = Annual_Rev_mean_copy[i]
                # print (Annual_Rev_mean)
        
            print("Annual_Rev_max_npv:" + str(Annual_Rev_max_npv))
            ############################################################################
            Annual_Cos_min_npv = []
            Annual_Cos_mean_copy = np.copy(Annual_Cos_mean)
        
            for i in range(1, len(Annual_Cos_mean)):
                Annual_Cos_mean[i] = Annual_Cos_min[0][i]
        
               
                Net_cos_min = np.array(Annual_Rev_mean[1:]) - np.array(Annual_Cos_mean[1:])
                
        
                with_discf_cos = Net_cos_min * discf
                
                sum_with_cos_min = np.sum(with_discf_cos)
                
                Annual_Cos_min_npv.append(sum_with_cos_min - Capital[2][0])
               
                Annual_Cos_mean[i] = Annual_Cos_mean_copy[i]
                
        
            print("Annual_Cos_min_npv:" + str(Annual_Cos_min_npv))
            
            Annual_Cos_max_npv = []
            Annual_Cos_mean_copy = np.copy(Annual_Cos_mean)
            for i in range(1, len(Annual_Cos_mean)):
                Annual_Cos_mean[i] = Annual_Cos_max[0][i]
        
               
                Net_cos_max = np.array(Annual_Rev_mean[1:]) - np.array(Annual_Cos_mean[1:])
               
        
                sum_with_cos_max = np.sum(Net_cos_max * discf)
        
                Annual_Cos_max_npv.append(sum_with_cos_max - Capital[2][0])
                
                Annual_Cos_mean[i] = Annual_Cos_mean_copy[i]
               
        
            print("Annual_Cos_max_npv:" + str(Annual_Cos_max_npv))
            ############################################################################
            discount_rate_range = [rate * (1 - float(change)), rate * rate * (1 + float(change))]
        
            discount_rate_min_npv = []
            discf_min = []
        
            for i in range(1, len(YearNo)):
                discf_min.append(1 / ((1 + discount_rate_range[0]) ** i))
            print("discf_min:" + str(discf_min))
            Net_dis_min = np.array(Annual_Rev[3]) - np.array(Annual_Cost[3])
            with_discf_min = Net_dis_min[1:] * discf_min
            sum_with_discf_min = np.sum(with_discf_min)
            discount_rate_min_npv.append(sum_with_discf_min - Capital[2][0])
            print("discount_rate_min_npv:" + str(discount_rate_min_npv))
        
            discount_rate_max_npv = []
            discf_max = []
        
            for i in range(1, len(YearNo)):
                discf_max.append(1 / ((1 + discount_rate_range[1]) ** i))
            print("discf_max:" + str(discf_max))
            Net_dis_max = np.array(Annual_Rev[3]) - np.array(Annual_Cost[3])
            with_discf_max = Net_dis_max[1:] * discf_max
            sum_with_discf_max = np.sum(with_discf_max)
            discount_rate_max_npv.append(sum_with_discf_max - Capital[2][0])
            print("discount_rate_max_npv:" + str(discount_rate_max_npv))
        
            titles = ["CapCost"]
            for i in range(1, len(YearNo)):
                titles.append("Rev" + str(YearNo[i]))
        
            for i in range(1, len(YearNo)):
                titles.append("Cost" + str(YearNo[i]))
        
            titles.append("DisRate")
        
            print("titles:" + str(titles))
        
            NPV_values = []  # yahoo npv
            NPV_values.append([Capital_npv[2], Capital_npv[0]])
            for i in range(0, len(Annual_Rev_min_npv)):
                NPV_values.append([Annual_Rev_min_npv[i], Annual_Rev_max_npv[i]])
        
            for i in range(0, len(Annual_Cos_min_npv)):
                NPV_values.append([Annual_Cos_min_npv[i], Annual_Cos_max_npv[i]])
        
            NPV_values.append([discount_rate_min_npv[0], discount_rate_max_npv[0]])
        
            print("NPV_values:" + str(NPV_values))
        
            yahoo_npv = []
            yahoo_npv.append(Capital_npv[2])
            yahoo_npv.append(Capital_npv[0])
            for i in range(0, len(Annual_Rev_min_npv)):
                yahoo_npv.append(Annual_Rev_min_npv[i])
            for i in range(0, len(Annual_Rev_max_npv)):
                yahoo_npv.append(Annual_Rev_max_npv[i])
        
            for i in range(0, len(Annual_Cos_min_npv)):
                yahoo_npv.append(Annual_Cos_min_npv[i])
            for i in range(0, len(Annual_Cos_max_npv)):
                yahoo_npv.append(Annual_Cos_max_npv[i])
        
            yahoo_npv.append(discount_rate_min_npv[0])
            yahoo_npv.append(discount_rate_max_npv[0])
        
            yahoo = []
            for i in range(0, len(titles)):
                yahoo.append([titles[i], NPV_values[i]])
            print("Yahoo:" + str(yahoo))
            avg_npv = np.average(yahoo_npv)
            print("avg_npv:" + str(avg_npv))
        
            base = avg_npv
        
            values = []
            for i in range(0, len(yahoo)):
                y = yahoo[i][1][1] - yahoo[i][1][0]
        
                values.append(y)
            print("values:" + str(values))
            lows_list = []
            for z in range(0, len(values)):
                lows_list.append(base - values[z] / 2)
            print("lows_list:" + str(lows_list))
        
            lows_values = []
            for x in range(0, len(lows_list)):
                y = (titles[x], lows_list[x], values[x])
                lows_values.append(y)
            print("lows_values:" + str(values))
        
            dtype = [('name', 'S10'), ('lows', float), ('value', int)]
        
            a = np.array(lows_values, dtype=dtype)  # create a structured array
            b = np.sort(a, order='lows')
            print("b:" + str(b))
        
            variables_sorted = []
            lows_sorted = []
            values_sorted = []
            for i in range(0, len(b)):
                v = b[i][0]
                variables_sorted.append(v)
                l = b[i][1]
                lows_sorted.append(l)
                va = b[i][2]
                values_sorted.append(va)
            print(variables_sorted)
            print(lows_sorted)
            print(values_sorted)
        
            # The actual drawing part
        
            # The y position for each variable
            ys = range(len(values_sorted))[::-1]  # top to bottom
        
            # Plot the bars, one by one
            for y, low, value in zip(ys, lows_sorted, values_sorted):
                pyplot.broken_barh(
                    [(low,value)], (y - 0.4, 0.8),
        
                    facecolors=['lightgreen'],
        
                    edgecolors=['white'],
        
                    linewidth=1, 
                )
        
                x = base + 1
                
                pyplot.text(x ,y,'$'+str(value),  va='center', ha='center')
               
        
                # Draw a vertical line down the middle
            
            pyplot.axvline(base, color='black')
        
            # Position the x-axis on the top, hide all the other spines (=axis lines)
            axes = pyplot.gca()  # (gca = get current axes)
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
        
            fmt = '${x:,.0f}'
        
            tick = mtick.StrMethodFormatter(fmt)
        
            axes.xaxis.set_major_formatter(tick)
        
            # Make the y-axis display the variables
            pyplot.yticks(ys, variables_sorted)
        
            # Set the portion of the x- and y-axes to show
            pyplot.xlim(base - 5, base + 5)
            pyplot.ylim(-1, len(variables_sorted))
            pyplot.suptitle("NPV,Millions")
        
            pyplot.xlabel('The Average is $' + str(round(base, 2)), fontsize=12)
            pyplot.show()


        def make_tornado(change):
            
            sens_npv(YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most,change)






        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", expand=True, padx=10, pady=5, anchor='s')

        center_frame2 = tk.Frame(self)

        center_frame2.pack(side="top", expand=True, padx=10, pady=5, anchor="n")

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Sensitivity Analysis | File Upload", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the MS Excel file with the data to analyze: (.xlsx)\n",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()

        entry_1 = ttk.Entry(center_frame, width=50)

        entry_1.pack(side="left")
        entry_frame = tk.Frame(center_frame)

        entry_frame.pack()

        entry_change = ttk.Entry(entry_frame, width=8)

        entry_change.grid(row=2, column=0)

        entry_change.insert(0, "0.2")

        button_3 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1) )

        button_3.pack(side="left", pady=5)
        button_4 = ttk.Button(center_frame, text="Submit", command=lambda:get_string() )

        button_4.pack(side="left", padx=5, pady=5)

       
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())

        button_1.pack(side="right", padx=5)


class OneDimension1(tk.Frame):

    """ This page allows the user to upload an excel file for one dimensional analysis.



    """

    def __init__(self, parent, controller):

        """

        Class constructor.

        :param parent: The parent frame of the page.

        :param controller: The controller runs between all pages and allows them to call the show_page function.

        """

        tk.Frame.__init__(self, parent)
        var_1 = tk.IntVar()
        global if_risk
        global directory
        directory =""

        def get_string():
            """ Retrieves the directory string from the entry widget.

            """
            global directory
            directory = entry_1.get()
           
            if directory == "":
                messagebox.showerror("Error", "An error has occurred in your File.")
            else:
                 global task,most,risk_tsk,risk_most,minn,maxn,risk_min,risk_max
                 simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most,minn,maxn,risk_min,risk_max =read_cost_excel(directory)
                 for i in range(0,len(minn)):
                     if minn[i]=='':
                         minn[i]=0
                     if maxn[i]=='':
                         maxn[i]=0
                     
                 make_tornado(var_1.get())
                 
        def clear_val():
            
            global directory,task,most,risk_tsk,risk_most,minn,maxn,risk_min,risk_max
            
            if directory!="":
                
                
                task=[]
                most=[]
                
                risk_tsk=[]
                minn=[]
                maxn=[]
                risk_min=[]
                risk_max=[]
                risk_most=[]
                del directory,task,most,risk_tsk,risk_most,minn,maxn,risk_min,risk_max
            
            controller.show_frame(Mainpage)
            entry_1.delete(0, 'end')
       

        def cost_tornado():
            variables=task
            print ("cost_variables:"+str(variables))
            base = np.average(most)
            values=np.array(maxn)-np.array(minn)
            print ("cost_values:"+str(values))
        
        
        
            lows_list = []
            for z in range(0, len(values)):
                lows_list.append(base - values[z] / 2)
            print("lows_list:" + str(lows_list))
        
            lows_values = []
            for x in range(0, len(lows_list)):
                y = (variables[x], lows_list[x], values[x])
                lows_values.append(y)
            print("lows_values:" + str(lows_values))
            # print (lows_values.sort(lows_values[]))
        
            dtype = [('name', 'S10'), ('lows', float), ('value', int)]
        
            a = np.array(lows_values, dtype=dtype)  # create a structured array
            b = np.sort(a, order='lows')
            print("b:" + str(b))
        
            variables_sorted = []
            lows_sorted = []
            values_sorted = []
            for i in range(0, len(b)):
                v = b[i][0]
                variables_sorted.append(v)
                l = b[i][1]
                lows_sorted.append(l)
                va = b[i][2]
                values_sorted.append(va)
            print(variables_sorted)
            print(lows_sorted)
            print(values_sorted)
        
        
        
            # print (np.sort(values))
            # The actual drawing part
        
            # The y position for each variable
            # ys = range(len(values))[::-1]  # top to bottom
            ys = range(len(variables_sorted))[::-1]
        
            # Plot the bars, one by one
            for y, low, value in zip(ys, lows_sorted, values_sorted):
                # The width of the 'low' and 'high' pieces
                #low_width = base - low
                #high_width = low + value - base
        
                # Each bar is a "broken" horizontal bar chart
                pyplot.broken_barh(
                    [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=[ 'lightgreen'],
        
                    edgecolors=[ 'white'],
                    linewidth=1,
                )
        
                # Display the value as text. It should be positioned in the center of
                # the 'high' bar, except if there isn't any room there, then it should be
                # next to bar instead.
                x = base + value / 2
                if x <= base + 50:
                    x = base + value/2 + 50
                pyplot.text(x, y, '$'+str(value), va='center', ha='center')
        
            # Draw a vertical line down the middle
            pyplot.axvline(base, color='black')
        
            # Position the x-axis on the top, hide all the other spines (=axis lines)
            axes = pyplot.gca()  # (gca = get current axes)
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '${x:.0f}'
        
            tick = mtick.StrMethodFormatter(fmt)
        
            axes.xaxis.set_major_formatter(tick)
        
            # Make the y-axis display the variables
            # plt.yticks(ys, variables)
            pyplot.yticks(ys, variables_sorted)
        
            pyplot.xlim(base - 9000, base + 9000)
        
            pyplot.ylim(-1, len(variables_sorted))
        
            pyplot.title("Project Cost")
            base_money = '${:,.2f}'.format(np.round(base, 2))
        
            pyplot.xlabel('Average =' + str(base_money), fontsize=12)
        
            pyplot.show()
        def cost_tornado_risk():
            variables = task+risk_tsk
            print("cost_risk_variables:" + str(variables))
            base = np.average(risk_most)+np.average(most)
        
            a=abs(np.array(maxn)-np.array(minn))
            print (type(a))
            b=abs(np.array(risk_max)-np.array(risk_min))
            print ("b:"+str(b))
            print (type(b))
            values=np.concatenate((a, b), axis=None)
            print("cost_risk_values:" + str(values))
        
            lows_list = []
            for z in range(0, len(values)):
                lows_list.append(base - values[z] / 2)
            print("lows_list:" + str(lows_list))
        
            lows_values = []
            for x in range(0, len(lows_list)):
                y = (variables[x], lows_list[x], values[x])
                lows_values.append(y)
            print("lows_values:" + str(lows_values))
            # print (lows_values.sort(lows_values[]))
        
            dtype = [('name', 'S10'), ('lows', float), ('value', int)]
        
            a = np.array(lows_values, dtype=dtype)  # create a structured array
            b = np.sort(a, order='lows')
            print("b:" + str(b))
        
            variables_sorted = []
            lows_sorted = []
            values_sorted = []
            for i in range(0, len(b)):
                v = b[i][0]
                variables_sorted.append(v)
                l = b[i][1]
                lows_sorted.append(l)
                va = b[i][2]
                values_sorted.append(va)
            print(variables_sorted)
            print(lows_sorted)
            print(values_sorted)
        
            # print (np.sort(values))
            # The actual drawing part
        
            # The y position for each variable
            # ys = range(len(values))[::-1]  # top to bottom
            ys = range(len(variables_sorted))[::-1]
        
            # Plot the bars, one by one
            for y, low, value in zip(ys, lows_sorted, values_sorted):
                # The width of the 'low' and 'high' pieces
                #low_width = base - low
                #high_width = low + value - base
        
                # Each bar is a "broken" horizontal bar chart
                pyplot.broken_barh(
                    [(low, value)],
                    (y - 0.4, 0.8),
                    facecolors=[ 'lightgreen'],
        
                    edgecolors=['white'],
                    linewidth=1,
                )
        
                # Display the value as text. It should be positioned in the center of
                # the 'high' bar, except if there isn't any room there, then it should be
                # next to bar instead.
                x = base + value / 2
                if x <= base + 50:
                    x = base + value/2 + 50
                pyplot.text(x, y, '$'+str(value), va='center', ha='center')
        
            # Draw a vertical line down the middle
            pyplot.axvline(base, color='black')
        
            # Position the x-axis on the top, hide all the other spines (=axis lines)
            axes = pyplot.gca()  # (gca = get current axes)
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '${x:.0f}'
        
            tick = mtick.StrMethodFormatter(fmt)
        
            axes.xaxis.set_major_formatter(tick)
        
            # Make the y-axis display the variables
            # plt.yticks(ys, variables)
            pyplot.yticks(ys, variables_sorted)
        
            pyplot.xlim(base - 9000, base + 9000)
        
            pyplot.ylim(-1, len(variables_sorted))
        
            pyplot.title("Project Cost With Risk")
            base_money = '${:,.2f}'.format(np.round(base, 2))
        
            pyplot.xlabel('Average =' + str(base_money), fontsize=12)
        
            pyplot.show()

        def make_tornado(if_risk):
            if_risk = var_1.get()
           

            if if_risk > 0:
                # simul_pcost(int(entry_simul.get()))

                #simul_risk(10000)
                #simul_pcost(10000)
                

                cost_tornado_risk()



            else:
                #simul_pcost(10000)
                

                cost_tornado()



        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", expand=True, padx=10, pady=5, anchor='s')

        center_frame2 = tk.Frame(self)

        center_frame2.pack(side="top", expand=True, padx=10, pady=5, anchor="n")

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Sensitivity Analysis | File Upload", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the MS Excel file with the data to analyze: (.xlsx)\n",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()
        check_1 = ttk.Checkbutton(center_frame, text="Add Risk", variable=var_1)

        check_1.pack(side="left", padx=5, pady=5)

        check_1.state(['!alternate'])
        entry_1 = ttk.Entry(center_frame, width=50)

        entry_1.pack(side="left")

        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))

        button_2.pack(side="left", pady=5)
        button_3 = ttk.Button(center_frame, text="Submit", command=lambda: get_string())

        button_3.pack(side="left", padx=5, pady=5)

        #button_3 = ttk.Button(center_frame2, text="Run as NPV", command=lambda: get_string())

        #button_3.pack(side="right", pady=5)

        #button_4 = ttk.Button(center_frame2, text="Run as Project Cost", command=lambda: get_string2())

        #button_4.pack(side="right", pady=5)

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())

        button_1.pack(side="right", padx=5)

 
        
#########################################################################################################################################
class NetPresentValue(tk.Frame):

    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)
        global directory
        directory =""

        def get_string():
            """ Retrieves the directory string from the entry widget.

            """
            global directory
            directory = entry_1.get()
           
            if directory == "":
                messagebox.showerror("Error", "An error has occurred in your File.")
            else:
                 global  YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
                 YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most =read_NPV_excel(directory)
                 controller.show_frame(NPVSimul)
        def clear_val():
            
            global directory,YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
            
            if directory!="":
                YearNo=[]
                Capital=[]
                Annual_Rev=[]
                Annual_Cost=[]
                Discount_Rate_Most=[]
                del directory,YearNo, Capital, Annual_Rev, Annual_Cost, Discount_Rate_Most
            
            controller.show_frame(ChoosePage)
            entry_1.delete(0, 'end')
    
        
        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Net Present Value | File Upload", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the MS Excel file with your NPV data: (.xlsx)\n",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()

        entry_1 = ttk.Entry(center_frame, width=50)

        entry_1.pack(side="left")

        #button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1)NPV_excel(enter))

        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))

        button_2.pack(side="left", pady=5)

        button_3 = ttk.Button(center_frame, text="Submit", command=lambda: get_string())

        button_3.pack(side="left", padx=5, pady=5)

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())

        button_1.pack(side="right", padx=5)


class NPVSimul(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

    
        def check_capital_distro(year_no,):
            """
            Displays the distribution graph for the given input / task.
            :param task_no: The index number of the task to display.
            :return: None.
            """
            i = year_no

            if i > len(YearNo) - 1 or i < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Capital[0][i] =='':
                messagebox.showerror("Hey!",YearNo[i]+" has no distribution. If this is a mistake check "
                                       "your excel file for typos!")
           
            if Capital[0][i] == "Normal":
                
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Capital Cost, Millions")

                b = Capital[2][i]
                c = Capital[4][i]

                x = Normal(b, c)
                x.plot()
                x.plot(hist=True)
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Capital Cost ' + str(i)+"\n"+"Normal Distribution")
                pyplot.draw()
                pyplot.show()

            elif Capital[0][i] == "Pert":
                pyplot.clf()
                host = pyplot.subplot(111)
                a = Capital[3][i]
                b = Capital[2][i]
                c = Capital[1][i]

                x = PERT(a, b, c)
                x.plot()
                x.plot(hist=True)
                # data = pert_distro(Capital_Cost_Min[i], Capital_Cost_Most[i], Capital_Cost_Max[i], 10000)

                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Capital Cost, Millions")
                # host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Capital Cost ' + str(i)+"\n"+"Pert Distribution")
                pyplot.draw()
                pyplot.show()

            elif Capital[0][i] == "Triangular":
                #data = np.random.triangular(Capital_Cost_Min[i], Capital_Cost_Most[i], Capital_Cost_Max[i], 10000)
                #data = tri_distro(Capital_Cost_Min[i], Capital_Cost_Most[i], Capital_Cost_Max[i],10000)

                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Capital Cost, Millions")
                a = Capital[3][i]
                b = Capital[2][i]
                c = Capital[1][i]

                x = Tri(a, b, c)
                x.plot()
                x.plot(hist=True)
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Capital Cost ' + str(i)+"\n"+"Triangular Distribution")
                pyplot.draw()
                pyplot.show()

        # this function displays the task statistics picked by the user
        def task_capital_stats(year_no):
            i = year_no
           
            title = 'Capital Cost ' + str(i)+"(Million)"
            if i > len(YearNo) - 1 or i < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Capital[0][i] =='':
                messagebox.showerror("Hey!",YearNo[i]+" has no distribution. If this is a mistake check "
                                       "your excel file for typos!")
           
            elif Capital[0][i] == "Normal":
                #data = norm_distro(, 10000)
                data = np.random.normal(Capital[2][i], Capital[4][i], 10000)
                make_table(data, title)
            elif Capital[0][i] == "Pert":
                data = pert_distro(Capital[3][i], Capital[2][i], Capital[1][i], 10000)

                make_table(data, title)
            elif Capital[0][i] == "Triangular":
                #data = tri_distro(Capital_Cost_Min[i], Capital_Cost_Most[i], Capital_Cost_Max[i],10000)
                data = np.random.triangular(Capital[3][i], Capital[2][i], Capital[1][i], 10000)
                make_table(data, title)

        def check_anrev_distro(yr):
            """
            Displays the distribution graph for the given input / task.
            :param task_no: The index number of the task to display.
            :return: None.
            """
            j = yr

            if j > len(YearNo) - 1 or j < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Annual_Rev[0][j] =='':
                messagebox.askokcancel("Hey!", YearNo[j] + " has no distribution. If this is a mistake check "
                                                           "your excel file for typos!")
            if Annual_Rev[0][j] == "Normal":
             
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Revenue, Millions")

                b = Annual_Rev[3][j]
                c = Annual_Rev[4][j]

                x = Normal(b, c)
                x.plot()
                x.plot(hist=True)
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Revenue ' + str(j)+"\n"+"Normal Distribution")
                pyplot.draw()
                pyplot.show()
            elif Annual_Rev[0][j] == "Pert":
                pyplot.clf()
                host = pyplot.subplot(111)
                a = Annual_Rev[2][j]
                b = Annual_Rev[3][j]
                c = Annual_Rev[1][j]

                x = PERT(a, b, c)
                x.plot()
                x.plot(hist=True)

                # data = pert_distro(Annual_Rev_Min[i], Annual_Rev_Most[i], Annual_Rev_Max[i], 10000)

                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Revenue, Millions")
                # host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Revenue ' + str(j)+"\n"+"Pert Distribution")
                pyplot.draw()
                pyplot.show()

            elif Annual_Rev[0][j] == "Triangular":
                #data = np.random.triangular(Annual_Rev_Min[j], Annual_Rev_Most[j], Annual_Rev_Max[j], 10000)
                # data = tri_distro(Annual_Rev_Min[i], Annual_Rev_Most[i], Annual_Rev_Max[i], 10000)
                pyplot.clf()
                host = pyplot.subplot(111)
                a = Annual_Rev[2][j]
                b = Annual_Rev[3][j]
                c = Annual_Rev[1][j]

                x = Tri(a, b, c)
                x.plot()
                x.plot(hist=True)

                # data = pert_distro(Annual_Rev_Min[i], Annual_Rev_Most[i], Annual_Rev_Max[i], 10000)

                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Revenue, Millions")
                # host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Revenue ' + str(j)+"\n"+"Triangular Distribution")
                pyplot.draw()
                pyplot.show()



        # this function displays the task statistics picked by the user
        def task_anrev_stats(yr):
            j = yr

            title = 'Annual Revenue ' + str(j )+"(Million)"

            if j > len(YearNo) - 1 or j < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Annual_Rev[0][j] =='':
                messagebox.askokcancel("Hey!", YearNo[j] + " has no distribution. If this is a mistake check "
                                                         "your excel file for typos!")
            elif Annual_Rev[0][j] == "Normal":
                #data = norm_distro(Annual_Rev_Most[i], Annual_Rev_sigma[i], 10000)
                data = np.random.normal(Annual_Rev[3][j], Annual_Rev[4][j], 10000)
                make_table(data, title)
            elif Annual_Rev[0][j] == "Pert":
                data = pert_distro(Annual_Rev[2][j], Annual_Rev[3][j], Annual_Rev[1][j], 10000)
                make_table(data, title)
            elif Annual_Rev[0][j] == "Triangular":
                #data = tri_distro(Annual_Rev_Min[i], Annual_Rev_Most[i], Annual_Rev_Max[i], 10000)
                data = np.random.triangular(Annual_Rev[2][j], Annual_Rev[3][j], Annual_Rev[1][j], 10000)
                make_table(data, title)


        def check_ancost_distro(year):
            """
            Displays the distribution graph for the given input / task.
            :param task_no: The index number of the task to display.
            :return: None.
            """
            z = year

            if z > len(YearNo) - 1 or z < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Annual_Cost[0][z] =='':
                messagebox.askokcancel("Hey!", YearNo[z] + " has no distribution. If this is a mistake check "
                                                         "your excel file for typos!")


            elif Annual_Cost[0][z] == "Normal":
                #data = np.random.normal(Annual_Cost_Most[z], Annual_Cost_sigma[z], 10000)
                #data = norm_distro(Annual_Cost_Most[i], Annual_Cost_sigma[i], 10000)

                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Cost, Millions")

                b = Annual_Cost[3][z]
                c = Annual_Cost[4][z]

                x = Normal( b, c)
                x.plot()
                x.plot(hist=True)
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Cost ' + str(z)+"\n"+"Normal Distribution" )
                pyplot.draw()
                pyplot.show()

            elif Annual_Cost[0][z] == "Pert":
                pyplot.clf()
                host = pyplot.subplot(111)
                a = Annual_Cost[2][z]
                b = Annual_Cost[3][z]
                c = Annual_Cost[1][z]

                x = PERT(a, b, c)
                x.plot()
                x.plot(hist=True)

                #data = pert_distro(Annual_Cost_Min[i], Annual_Cost_Most[i], Annual_Cost_Max[i], 10000)

                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Cost, Millions")
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Cost' + str(z)+"\n"+"Pert Distribution" )
                pyplot.draw()
                pyplot.show()

            elif Annual_Cost[0][z] == "Triangular":
                #data = np.random.triangular(Annual_Cost_Min[z], Annual_Cost_Most[z], Annual_Cost_Max[z], 10000)
                #data = tri_distro(Annual_Cost_Min[i], Annual_Cost_Most[i], Annual_Cost_Max[i], 10000)

                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Relative Frequency")
                host.set_xlabel("Annual Cost, Millions")
                a = Annual_Cost[2][z]
                b = Annual_Cost[3][z]
                c = Annual_Cost[1][z]

                x = Tri(a, b, c)
                x.plot()
                x.plot(hist=True)
                #host.hist(data, bins=50)
                fmt = '${x:,.2f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)

                pyplot.title('Annual Cost ' + str(z )+"\n"+"Triangular Distribution" )
                pyplot.draw()
                pyplot.show()

        # this function displays the task statistics picked by the user
        def task_ancost_stats(year):
            z = year

            title = 'Annual Cost ' + str(z)+"(Million)"

            if z > len(YearNo) - 1 or z < 0:
                messagebox.askokcancel("Hey!", "This task lies outside of the range of specified.")
            elif Annual_Cost[0][z]=='':
                messagebox.askokcancel("Hey!", YearNo[z] + " has no distribution. If this is a mistake check "
                                                        "your excel file for typos!")
            elif Annual_Cost[0][z] == "Normal":
                #data = norm_distro(Annual_Cost_Most[i], Annual_Cost_sigma[i], 10000)
                data = np.random.normal(Annual_Cost[3][z], Annual_Cost[4][z], 10000)
                make_table(data, title)
            elif Annual_Cost[0][z] == "Pert":
                data = pert_distro(Annual_Cost[2][z], Annual_Cost[3][z], Annual_Cost[1][z], 10000)
                make_table(data, title)
            elif Annual_Cost[0][z] == "Triangular":
                #data = tri_distro(Annual_Cost_Min[i], Annual_Cost_Most[i], Annual_Cost_Max[i], 10000)
                data = np.random.triangular(Annual_Cost[2][z], Annual_Cost[3][z], Annual_Cost[1][z], 10000)
                make_table(data, title)
        
        def simul_npv(f):
                  
                   # global CapitalCost, CapitalCost1, AnnualRev, AnnualRev1, AnnualCost, AnnualCost1, NPV, DiscountRate
                        
            CapitalCost = []
            AnnualRev = []
            AnnualCost = []
            AnnualNetCF = []
            CapitalCost1=[]
            PV_NetCF=[]
            for i in range(0, f): 
                
                CapitalCost.append([])
                AnnualRev.append([])
                AnnualCost.append([])
                AnnualNetCF.append([])
                
                for j in range(0, len(Capital[0])):
                    x = 0
                    
                    CapitalCost[i].append(x)
                    AnnualRev[i].append(x)
                    AnnualCost[i].append(x)
                    AnnualNetCF[i].append(x)
            DiscountFactor = []
            for i in range(0, len(YearNo)):
                if Discount_Rate_Most[i]=='':
                    Discount_Rate_Most[i]=0
                d = 1 / ((1 + Discount_Rate_Most[i]) ** YearNo[i])
                DiscountFactor.append(d)
                
            for i in range(0, len(Capital[0])):
                if Capital[0][i] == "Pert":
                    a = pert_distro(Capital[3][i], Capital[2][i], Capital[1][i], f)
                    x += a
                elif Capital[0][i] == "Triangular":
                    a = tri_distro(Capital[3][i], Capital[2][i], Capital[1][i], f)
                    x += a
                elif Capital[0][i] == "Normal":
                    a = norm_distro(Capital[2][i], Capital[4][i], f)
                    x += a
                elif Capital[0][i] == "Poisson":
                    a = pois_distro(Capital[5][i], f)
                    x += a
                else:
                    if Capital[2][i]=='':
                        a= no_distro(0, f)
                        x += a
                    else:
                        a = no_distro(Capital[2][i], f)
                        x += a
                for j in range(0,f):
                    CapitalCost[j][i]=a[j]
            for j in range(0,f):
                 CapitalCost1.append(x[j])
            CapitalCost=np.array(CapitalCost).T.tolist()        
            
            c1=0
            c2=0
            for i in range(0, len(Annual_Rev[0])):
                if Annual_Rev[0][i] == "Pert":
                    a = pert_distro(Annual_Rev[2][i], Annual_Rev[3][i], Annual_Rev[1][i], f)
                    a1=a*DiscountFactor[i]
                    c1 +=a1
                    #AnnualRev.append(a)
                elif Annual_Rev[0][i] == "Triangular":
                    a = tri_distro(Annual_Rev[2][i], Annual_Rev[3][i], Annual_Rev[1][i], f)
                    a1=a*DiscountFactor[i]
                    c1+=a1
                    #AnnualRev.append(a)
                elif Annual_Rev[0][i] == "Normal":
                    a = norm_distro(Annual_Rev[3][i], Annual_Rev[4][i], f)
                    a1=a*DiscountFactor[i]
                    c1+=a1
                    #AnnualRev.append(a)
                elif Annual_Rev[0][i] == "Poisson":
                    a = pois_distro(Annual_Rev[5][i], f)
                    a1=a*DiscountFactor[i]
                    c1+=a1
                    #AnnualRev.append(a)
                else:
                    if Annual_Rev[3][i]=='':
                        a = no_distro(0, f)
                        a1=a*DiscountFactor[i]
                        c1+=a1
                    else:
                        
                        a = no_distro(Annual_Rev[3][i], f)
                        a1=a*DiscountFactor[i]
                        c1+=a1
                
                if Annual_Cost[0][i] == "Pert":
                    b = pert_distro(Annual_Cost[2][i], Annual_Cost[3][i], Annual_Cost[1][i], f)
                    b2=b*DiscountFactor[i]
                    c2+=b2
                    #AnnualCost.append(a)
                elif Annual_Cost[0][i] == "Triangular":
                    b = tri_distro(Annual_Cost[2][i], Annual_Cost[3][i], Annual_Cost[1][i], f)
                    b2=b*DiscountFactor[i]
                    c2+=b2
                    #AnnualCost.append(a)
                elif Annual_Cost[0][i] == "Normal":
                    b = norm_distro(Annual_Cost[3][i], Annual_Cost[4][i], f)
                    b2=b*DiscountFactor[i]
                    c2+=b2
                    #AnnualCost.append(a)
                elif Annual_Cost[0][i] == "Poisson":
                    b = pois_distro(Annual_Cost[5][i], f)
                    b2=b*DiscountFactor[i]
                    c2+=b2
                    #AnnualCost.append(a)
                else:
                    if Annual_Cost[3][i]=="":
                        b = no_distro(0, f)
                        b2=b*DiscountFactor[i]
                        c2+=b2
                    else:
                        b = no_distro(Annual_Cost[3][i], f)
                        b2=b*DiscountFactor[i]
                        c2+=b2
                    #AnnualCost.append(a)
               
                for j in range(0,f):
                    AnnualRev[j][i]=a[j]
                    AnnualCost[j][i]=b[j]
                    AnnualNetCF[j][i]=a[j]-b[j]
            for j in range(0,f):
                 PV_NetCF.append(c1[j]-c2[j])
            
            AnnualRev=np.array(AnnualRev).T.tolist()    
            AnnualCost=np.array(AnnualCost).T.tolist()    
            
        
            Npv =(np.array(PV_NetCF) - np.array(CapitalCost1)).tolist()
            #Ave=np.average(Npv)
            title=['Capital_Cost'] 
            for i in range(0,len(AnnualRev)):
                title.append("Annualrev"+str(YearNo[i]))
            for i in range(0,len(AnnualCost)):
                title.append("AnnualCost"+str(YearNo[i]))
           
            return Npv,CapitalCost1,AnnualRev,AnnualCost,title            
          
        
        def one_dim_tornado_npv(Npv,CapitalCost1,AnnualRev,AnnualCost,YearNo,title):
             
            aaa=[CapitalCost1]+AnnualRev+AnnualCost
           
            bbb=title
            avg_npv=np.average(Npv)
            low = []
            high = []
            value = []
            name=[]
            x=Npv
            def sort_second(val):
                return val[1]
            
            for i in range(0, len(aaa)):
                
                s=aaa[i]
                avs=np.mean(s)
                #avx=np.mean(x)
                     
                if round(np.average(s))==0:
                   i+1            
                else:
                    a=round(0.05*len(s))
                    l=[]
                    h=[]
                    for j in range(0,a):
                        
                        minpos = s.index(min(s)) 
                        s[minpos]= avs
                        u=x[minpos]
                        #x[minpos]= avx
                        maxpos = s.index(max(s))
                        s[maxpos]=avs
                        v=x[maxpos]
                        #x[maxpos]=avx
                        l.append(u)
                        h.append(v)
        
                    v=np.average(l)
                    u=np.average(h)
                    if u>v:
                        low.append(round(v,2))
                        high.append(round(u,2))
                    elif u<=v:
                        low.append(round(u,2))
                        high.append(round(v,2))
                        
                    name.append(bbb[i]) 
            mains = []
        
            for i in range(0, len(low)):
                v = round(float(high[i]) - float(low[i]),3)
                value.append(v)
                x = (name[i], value[i], low[i], high[i])
                mains.append(x)
        
            mains.sort(key=sort_second, reverse=True)
            values = []
            variables = []
            lows = []
            highs=[]
            for i in range(0, len(mains)):
                x = (mains[i][0])
                variables.append(x)
            
            for i in range(0, len(mains)):
                y = mains[i][1]
                values.append(y)
            
            for i in range(0, len(mains)):
                z = mains[i][2]
                lows.append(z)
            for i in range(0, len(mains)):
                h = mains[i][3]
                highs.append(h)
           
            base = int(avg_npv)
            return variables,base, values, lows, highs
           
        def run_simul(f):
            global  Npv,CapitalCost1,AnnualRev,AnnualCost,title
            Npv,CapitalCost1,AnnualRev,AnnualCost,title=simul_npv(int(entry_simul.get()))
   
        def plot():
              
            variables,base, values, lows, highs=one_dim_tornado_npv(Npv,CapitalCost1,AnnualRev,AnnualCost,YearNo,title)
            
            ys = range(len(values))[::-1]
    
    
            for y, low, high,value in zip(ys, lows,highs, values):
                
                
               
                pyplot.broken_barh(
                   [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=['lightgreen'],
                    edgecolors=['white'],
                    linewidth=1,
                )
                x = base + value / 4
                if x <= base + 1:
                   x = base + value/4 + 1
                pyplot.text(x, y, '$'+str(round(value)), va='center', ha='center')
            
            
            pyplot.axvline(base, color='black') 
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '${x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            axes.xaxis.set_major_formatter(tick)
            pyplot.yticks(ys, variables)
            pyplot.xlim(base - 3, base + 4)
            pyplot.ylim(-1, len(variables))
            pyplot.title("Project Cost, Millions")
            base_money = '${:,.2f}'.format(round(base))
            pyplot.xlabel('Average =' + str(base_money), fontsize=12)
            pyplot.show()
    
        def write_to_excel_npv( ):
        
            messagebox.showwarning("Hey", "1. The iteration data are stored in new Excel file titled 'NPV Output Data' inside "
        
                                         "the source folder. \n \n2. Before you run the next simulation, please make sure the "
        
                                         "'NPV Output Data' Excel file is closed, so that the new iteration data "
        
                                         "can override the old iteration data.")
        
            wb = xlsxwriter.Workbook('NPV Output Data.xlsx')
        
            ws = wb.add_worksheet()
        
            A1 = 'NPV'
           
            titled = np.append(A1, title)
        
            item = titled
            yahoo_npv=[Npv]+[CapitalCost1]+AnnualRev+AnnualCost
        
            column = 0
        
            for i in item:
        
                ws.write(0, column, i)
        
                column += 1
        
        
            col = 0
        
            for i in yahoo_npv:
        
                r = 1
        
                for j in i:
        
                    # print('data: %f, r:%d'%(j,r))
        
                    ws.write(r, col, j)
        
        
        
                    r += 1
        
                col += 1
        
        
        
            wb.close()
       
        # check to see if any data exists before displaying stats
        def run_stats():
            if np.any(Npv):
                make_table(Npv, "NPV Statistics")

        # checks to see if data exists before displaying graphs
        def run_graph(title, cdf, pdf):
            if np.any(Npv):
                make_graph(Npv, title, cdf, pdf)

        def make_tornado():

            plot()
            
        def write_excel_npv():
           
            write_to_excel_npv()
        def clear_val():
            global  Npv,CapitalCost1,AnnualRev,AnnualCost,title
            
            
            Npv=[]
            CapitalCost1=[]
            AnnualRev=[]
            AnnualCost=[]
            title=[]
            del Npv,CapitalCost1,AnnualRev,AnnualCost,title
            
            controller.show_frame(NetPresentValue)
            
       


        top_frame = tk.Frame(self)
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)
        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)
        label_1 = tk.Label(center_frame, text="Net Present Value | Simulation Settings", font=HEADR_FONT)
        label_1.pack()
        message_1 = tk.Message(center_frame, text="Check the distributions of your individual tasks to see that "

                               "they are correct.\nThen input 'Number of Iterations' and select 'Run Simulation'.",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        message_2 = tk.Message(center_frame, text="  "

                               "  ",
                               justify="center", width=500, font=MESSG_FONT)
        message_2.pack()

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())
        button_1.pack(side="right", padx=5)

        entry_frame = tk.Frame(center_frame)
        entry_frame.pack()

        entry_year = ttk.Entry(entry_frame, width=8)
        entry_year.grid(row=0, column=0)
        entry_year.insert(0, "0")
        entry_year_rev = ttk.Entry(entry_frame, width=8)
        entry_year_rev.grid(row=1, column=0)
        entry_year_rev.insert(0, "1")
        entry_year_cost = ttk.Entry(entry_frame, width=8)
        entry_year_cost.grid(row=2, column=0)
        entry_year_cost.insert(0, "1")
        button_2 = ttk.Button(entry_frame, text="Capital Cost Distribution",command=lambda: check_capital_distro(int(entry_year.get())))
        button_2.grid(row=0, column=1)
        button_4 = ttk.Button(entry_frame, text="Capital Cost Statistics",command=lambda: task_capital_stats(int(entry_year.get())))
        button_4.grid(row=0, column=2)
        button_5 = ttk.Button(entry_frame, text="Annual Revenue Distribution",command=lambda: check_anrev_distro(int(entry_year_rev.get())))
        button_5.grid(row=1, column=1)
        button_6 = ttk.Button(entry_frame, text="Annual Revenue Statistics",command=lambda: task_anrev_stats(int(entry_year_rev.get())))
        button_6.grid(row=1, column=2)
        button_7 = ttk.Button(entry_frame, text="Annual Cost Distribution",command=lambda: check_ancost_distro(int(entry_year_cost.get())))
        button_7.grid(row=2, column=1)
        button_8 = ttk.Button(entry_frame, text="Annual Cost Statistics",command=lambda: task_ancost_stats(int(entry_year_cost.get())))
        button_8.grid(row=2, column=2)

        entry_simul = ttk.Entry(entry_frame, width=8)
        entry_simul.grid(row=3, column=0)
        entry_simul.insert(0, "1000")

        label_sims = tk.Label(entry_frame, text="Number of Iterations")
        label_sims.grid(row=3, column=1, columnspan=2)

        separate = ttk.Separator(entry_frame, orient="horizontal")
        separate.grid(row=6, column=0, columnspan=3, sticky="ew", pady=5)
        label_3 = tk.Label(entry_frame, text=" ")
        label_3.grid(row=7, column=1, columnspan=2)
        
        label_2 = tk.Label(entry_frame, text="Output Display:")
        label_2.grid(row=8, column=0, columnspan=3)
        label_5 = tk.Label(entry_frame, text=" "
                              " ")
        label_5.grid(row=4, column=1, columnspan=2)
        button_3 = ttk.Button(entry_frame, text="Run Simulation", command=lambda: run_simul(int(entry_simul.get())))
        button_3.grid(row=5, column=0, columnspan=3, sticky="ew")
        b_cp = ttk.Button(entry_frame, text="CDF & PDF Plot", command=lambda: run_graph("NPV", 1, 1))
        b_cp.grid(row=9, column=0,sticky="ew" )
        #b_cdf = ttk.Button(entry_frame, text="Plot CDF", command=lambda: run_graph("NPV", 1, 0))
        #b_cdf.grid(row=10, column=0)
        #b_pdf = ttk.Button(entry_frame, text="Plot PDF", command=lambda: run_graph("NPV", 0, 1))
        #b_pdf.grid(row=10, column=1)
        b_stat = ttk.Button(entry_frame, text="Statistics", command=lambda: run_stats())
        b_stat.grid(row=9, column=1,sticky="ew")
        b_tornado = ttk.Button(entry_frame, text="Tornado Diagram",
                               command=lambda: make_tornado())
        b_tornado.grid(row=9, column=2,sticky="ew")
        separate1 = ttk.Separator(entry_frame, orient="horizontal")
        separate1.grid(row=11, column=0, columnspan=3, sticky="ew", pady=5)
        label_4 = tk.Label(entry_frame, text=" "
                           " ")
        label_4.grid(row=12, column=1, columnspan=2)
        b_write_excel_npv = ttk.Button(entry_frame, text="Output Data",
                                        command=lambda: write_excel_npv())

        b_write_excel_npv.grid(row=13, column=0, columnspan=3, sticky="ew")




####################################################################################################################################
class ProjectCost(tk.Frame):
    """
    This page allows the user to select the excel file that contains the Project Cost data they would like to import.
    """
    def __init__(self, parent, controller):
        """
        Class constructor.
        :param parent: The parent frame.
        :param controller: The controller runs through every page and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)
        global directory
        directory =""
        #var_1 = tk.IntVar()
        
    


        def get_string():
            """ Retrieves the directory string from the entry widget.

            """
            global directory
            directory = entry_1.get()
           
            if directory == "":
                messagebox.showerror("Error", "An error has occurred in your File.")
            else:
                 global simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most
                 simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most,minn,maxn,risk_min,risk_max =read_cost_excel(directory)
                 controller.show_frame(PCRiskSimul)
                 
        def clear_val():
            
            global directory,simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most
            
            if directory!="":
                simul=[]
                distro=[]
                task=[]
                most=[]
                risk_sim=[]
                risk_dis=[]
                risk_tsk=[]
                risk_bin=[]
                risk_most=[]
                del directory,simul,distro,task,most,risk_sim,risk_dis,risk_tsk,risk_bin,risk_most
                controller.show_frame(ChoosePage)
                entry_1.delete(0, 'end')
            
            else:
                controller.show_frame(ChoosePage)
            
           
       
      
        
        top_frame = tk.Frame(self)
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)
        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)

        label_1 = tk.Label(center_frame, text="Project Cost Model | File Upload", font=HEADR_FONT)
        label_1.pack()
        message_1 = tk.Message(center_frame, text="Select the MS Excel file with your project cost data (.xlsx)\n",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        
        #check_1 = ttk.Checkbutton(center_frame, text="Add Risk", variable=var_1)
        #check_1.pack(side="left", padx=5, pady=5)
        #check_1.state(['!alternate'])
        
        entry_1 = ttk.Entry(center_frame, width=50)
        entry_1.pack(side="left")
        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))
        button_2.pack(side="left", pady=5)
        button_3 = ttk.Button(center_frame, text="Submit", command=lambda: get_string())
        button_3.pack(side="left", padx=5, pady=5)
        button_1 = ttk.Button(bottom_frame, text="Back", command= lambda: clear_val())
        button_1.pack(side="right", padx=5)


########################################################################################################################
class PCRiskSimul(tk.Frame):

    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)

        
        def simule(itera):
            global new_data,cost_simul_data
            global risk_dat,cost_simul_data_risk,ccc
            
            new_data,cost_simul_data=simul_pcost(itera,simul,distro)
            risk_dat,cost_simul_data_risk=simul_risk(itera,risk_sim,risk_dis,risk_bin)
            ccc=[x + y for x, y in zip(new_data,risk_dat)]
           
        def clear():
            
            global new_data,cost_simul_data
            global risk_dat,cost_simul_data_risk,ccc
            new_data=[]
            cost_simul_data=[]
            risk_dat=[]
            cost_simul_data_risk=[]
            ccc=[]
            
            del new_data,cost_simul_data
            del risk_dat,cost_simul_data_risk,ccc
            controller.show_frame(ProjectCost)
           
                
            
        

        def simul_pcost(itera,simul,distro):
           
            x = 0
           
            new_data=[]
            cost_simul_data=[]
            for i in range(0, itera): 
                
                cost_simul_data.append([])
                for j in range(0, len(distro)):
                    x = 0
                    
                    cost_simul_data[i].append(x)
                
            for i in range(0, len(distro)): 
                
                argu = distro[i]
                if argu == "Triangular":  # distribution
                    a=tri_distro(simul[i][0], simul[i][1], simul[i][2], itera)
                    x += a
                elif argu == "Pert":
                    a=pert_distro(simul[i][0], simul[i][1], simul[i][2], itera)
                    x += a
                elif argu == "Normal":
                    a=norm_distro(simul[i][0], simul[i][1], itera)
                    x += a
                elif argu == "Uniform":
                    a=uni_distro(simul[i][0], simul[i][1], itera)
                    x += a
                else:
                    a = no_distro(simul[i], itera)
                    x += a
                for j in range(0,itera):
            
                    cost_simul_data[j][i] = a[j] 
            
            for j in range(0,itera):
                new_data.append(x[j])
                    
            cost_simul_data=np.array(cost_simul_data).T.tolist()
            return  new_data,cost_simul_data
            
          

        def simul_risk(itera,risk_sim,risk_dis,risk_bin):
            z=0
            risk_dat=[]
            cost_simul_data_risk=[]
            for i in range(0, itera): 
                
                cost_simul_data_risk.append([])
                for j in range(0, len(risk_dis)):
                    x = 0
                    
                    cost_simul_data_risk[i].append(x)
            
            for i in range(0, len(risk_dis)): 
                b = bern_distro(risk_bin[i], itera)
                argu = risk_dis[i]
                if argu == "Triangular":  # distribution
                    x = tri_distro(risk_sim[i][0], risk_sim[i][1], risk_sim[i][2], itera)
                    a= x*b
                    z += a
                elif argu == "Pert":
                    x = pert_distro(risk_sim[i][0], risk_sim[i][1], risk_sim[i][2], itera)
                    a = x * b
                    z += a
                elif argu =="Poisson":
                    x +=pert_distro(risk_sim[i][0], risk_sim[i][1], risk_sim[i][2],itera)
                    a = x * np.random.poisson(risk_sim[i][3],itera)
                    z += a
                else:
                    x = no_distro(risk_sim[i],itera)
                    a = x * b
                    z += a
                for j in range(0,itera):
            
                    cost_simul_data_risk[j][i] = a[j] 
            
            for j in range(0,itera):
                 
                 risk_dat.append(z[j])
            cost_simul_data_risk=np.array(cost_simul_data_risk).T.tolist()

            return risk_dat,cost_simul_data_risk

       
         
    
    
        def check_distro(task_no,simul, distro, task):
            
            i = task_no
            if i > len(distro)-1 or i < 0:
    
                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")
            elif distro[i]=='':
                messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + " has no distribution.")
            elif distro[i] == "Normal":
                pyplot.clf()
                host = pyplot.subplot(111)
                mu = simul[i][0]
                sigma = simul[i][1]
                x = np.linspace(mu - 5 * sigma, mu + 5 * sigma, 100)
                pyplot.plot(x, ss.norm.pdf(x, mu, sigma))
                pyplot.fill(x, ss.norm.pdf(x, mu, sigma))
                host.set_ylabel("Probability Density")
                host.set_xlabel("Task Cost")
                fmt = '${x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
                pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Normal Distribution")
                pyplot.show()
            
            elif distro[i] == "Pert":
                pyplot.clf()
                host = pyplot.subplot(111)
                a = simul[i][0]
                b = simul[i][1]
                c = simul[i][2]
                x = PERT(a, b, c)
                x.plot(hist=True)
                host.set_ylabel("Probability Density")
                host.set_xlabel("Task Cost")
                fmt = '${x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
                pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Pert Distribution")
                pyplot.show()
    
            elif distro[i] == "Triangular":
                
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Probability Density")
                host.set_xlabel("Task Cost")
                c = (simul[i][1] - simul[i][0]) / (simul[i][2]-simul[i][0])
                mean, var, skew, kurt = triang.stats(c, moments='mvsk')
                x = np.linspace(simul[i][0], simul[i][2], 1000)
                host.plot(x, triang.pdf(x, c, loc=simul[i][0], scale=simul[i][2] - simul[i][0]), 'b', lw=2)
                pyplot.fill(x, triang.pdf(x, c, loc=simul[i][0], scale=simul[i][2] - simul[i][0]))
                fmt = '${x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
                pyplot.title( 'Task '+str(i+1)+': '+task[i]+"\n"+"Triangular Distribution")
                pyplot.show()
                
            elif distro[i]=="Uniform":
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_ylabel("Probability Density")
                host.set_xlabel("Task Cost")
                a=simul[i][0]
                b=simul[i][1]
                mean, var, skew, kurt = ss.uniform.stats(moments='mvsk')
                x = np.linspace(a, b, 100)
                host.plot(x, ss.uniform.pdf(x, loc=a, scale=b-1), 'b', lw=2, alpha=0.6, label='uniform pdf')
                fmt = '${x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
                pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Uniform Distribution")
                pyplot.show()
    
    
       
        def task_stats(task_no,simul, distro, task):

            i = task_no
            
            if i > len(distro)-1or i < 0:

                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")
            else:
                
               
                mcerp.npts = int(entry_simul.get())
    
                title = 'Task ' + str(i+1) + ': ' + task[i]
    
                if distro[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + " has no distribution.")
    
                elif distro[i] == "Normal":
    
                    data = np.random.normal(simul[i][0], simul[i][1], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Pert":
    
                    a = simul[i][0]
    
                    b = simul[i][1]
    
                    c = simul[i][2]
    
                    alpha = ((4 * b) + c - (5 * a)) / (c - a)
    
                    beta = ((5 * c) - a - (4 * b)) / (c - a)
    
                    x = np.random.beta(alpha, beta, int(entry_simul.get()))
    
                    data = a + (x * (c - a))
    
                    make_table(data, title)
    
                elif distro[i] == "Triangular":
    
                    data = np.random.triangular(simul[i][0], simul[i][1], simul[i][2], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Uniform":
    
                    data = np.random.uniform(simul[i][0],simul[i][1], int(entry_simul.get()))
    
                    make_table(data, title)

        def tornado(new_data,cost_simul_data,task):
            avg_npv=np.average(new_data)
            low = []
            high = []
            value = []
            name=[]
            def sort_second(val):
                return val[1]
            
            for i in range(0, len(cost_simul_data)):
                x = new_data
                s=cost_simul_data[i]
                avs=np.mean(s)
                avx=np.mean(x)
                     
                if round(np.average(s))==0:
                   i+1              
                else:
                    a=round(0.05*len(s))
                    l=[]
                    h=[]
                    for j in range(0,a):
                        
                        minpos = s.index(min(s)) 
                        s[minpos]= avs
                        u=x[minpos]
                        x[minpos]= avx
                        maxpos = s.index(max(s))
                        s[maxpos]=avs
                        v=x[maxpos]
                        x[maxpos]=avx
                        l.append(u)
                        h.append(v)
        
                    v=np.average(l)
                    u=np.average(h)
                    if u>v:
                        low.append(int(v))
                        high.append(int(u))
                    elif u<=v:
                        low.append(int(u))
                        high.append(int(v))
                        
                    name.append(task[i])     
            mains = []
        
            for i in range(0, len(low)):
                v = float(high[i]) - float(low[i])
                value.append(v)
                x = (name[i], value[i], low[i], high[i])
                mains.append(x)
        
            mains.sort(key=sort_second, reverse=True)
            values = []
            variables = []
            lows = []
            highs=[]
            
            for i in range(0, len(mains)):
                x = (mains[i][0])
                variables.append(x)
            
            for i in range(0, len(mains)):
                y = mains[i][1]
                values.append(y)
            
            for i in range(0, len(mains)):
                z = mains[i][2]
                lows.append(z)
            for i in range(0, len(mains)):
                h = mains[i][3]
                highs.append(h)
       
            base = int(avg_npv)
            return variables,base, values, lows, highs
        
        def tornado_risk(ccc,cost_simul_data,cost_simul_data_risk,task,risk_tsk):
            
            
            aaa=cost_simul_data+cost_simul_data_risk
            bbb=task+risk_tsk
            avg_npv=np.average(ccc)
            low = []
            high = []
            value = []
            name=[]
            def sort_second(val):
                return val[1]
            
            for i in range(0, len(aaa)):
                x = ccc
                s=aaa[i]
                avs=np.mean(s)
                avx=np.mean(x)
                     
                if round(np.average(s))==0:
                   i+1            
                else:
                    a=round(0.05*len(s))
                    l=[]
                    h=[]
                    for j in range(0,a):
                        
                        minpos = s.index(min(s)) 
                        s[minpos]= avs
                        u=x[minpos]
                        x[minpos]= avx
                        maxpos = s.index(max(s))
                        s[maxpos]=avs
                        v=x[maxpos]
                        x[maxpos]=avx
                        l.append(u)
                        h.append(v)
        
                    v=np.average(l)
                    u=np.average(h)
                    if u>v:
                        low.append(int(v))
                        high.append(int(u))
                    elif u<=v:
                        low.append(int(u))
                        high.append(int(v))
                        
                    name.append(bbb[i]) 
            mains = []
        
            for i in range(0, len(low)):
                v = float(high[i]) - float(low[i])
                value.append(v)
                x = (name[i], value[i], low[i], high[i])
                mains.append(x)
        
            mains.sort(key=sort_second, reverse=True)
            values = []
            variables = []
            lows = []
            highs=[]
            for i in range(0, len(mains)):
                x = (mains[i][0])
                variables.append(x)
            
            for i in range(0, len(mains)):
                y = mains[i][1]
                values.append(y)
            
            for i in range(0, len(mains)):
                z = mains[i][2]
                lows.append(z)
            for i in range(0, len(mains)):
                h = mains[i][3]
                highs.append(h)
       
            base = int(avg_npv)
            return variables,base, values, lows, highs
       
        def plot():
            
            variables,base, values, lows, highs=tornado(new_data,cost_simul_data,task)
            ys = range(len(values))[::-1]
    
    
            for y, low, high,value in zip(ys, lows,highs, values):
                
               
                pyplot.broken_barh(
                   [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=['lightgreen'],
                    edgecolors=['white'],
                    linewidth=1,
                )
                x = base + value / 4
                if x <= base + 3000:
                   x = base + value/4 + 3000
                pyplot.text(x, y, '$'+str(round(value)), va='center', ha='center')
            
            
                #pyplot.text(high+50,y-0.2,'$'+str(round(high)),  color='lightgreen', fontweight='bold')
                #pyplot.text(low-200,y-0.2, '$'+str(round(low)),  va='center',va='center')
            
            pyplot.axvline(base, color='black') 
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '${x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            axes.xaxis.set_major_formatter(tick)
            pyplot.yticks(ys, variables)
            pyplot.xlim(base - 12000, base + 12000)
            pyplot.ylim(-1, len(variables))
            pyplot.title("Project Cost")
            base_money = '${:,.0f}'.format(round(base))
            pyplot.xlabel('Average =' + str(base_money), fontsize=12)
            pyplot.show()

            
        def plot_risk():
            variables,base, values, lows, highs=tornado_risk(ccc,cost_simul_data,cost_simul_data_risk,task,risk_tsk)
            ys = range(len(values))[::-1]
            for y, low, high,value in zip(ys, lows,highs, values):
                
               
                pyplot.broken_barh(
                   [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=['lightgreen'],
                    edgecolors=['white'],
                    linewidth=1,
                )
                x = base + value/4
                if x <= base + 3000:
                   x = base + value/4 + 3000
                pyplot.text(x, y, '$'+str(round(value)), va='center', ha='center')
                
            
                #pyplot.text(high+100,y-0.2,'$'+str(round(high)),  color='lightgreen', fontweight='bold')
                
            pyplot.axvline(base, color='black')
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '${x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            axes.xaxis.set_major_formatter(tick)
            pyplot.yticks(ys, variables)
            pyplot.xlim(base - 12000, base + 12000)
            pyplot.ylim(-1, len(variables))
            pyplot.suptitle("Project Cost with Risk")
            base_money = '${:,.0f}'.format(base)
            pyplot.xlabel('Average ='+str(base_money), fontsize=12)
            pyplot.show()

        
        def check_distro_risk (risk_no,risk_sim,risk_dis,risk_tsk,risk_bin):

            """

            Displays the distribution graph for the given input / task.

            :param task_no: The index number of the task in risk register to display.

            :return: None.

            """

            i = risk_no



            if i > len(risk_dis)-1or i < 0:

                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")

            elif risk_dis[i]=='':
                messagebox.showerror("Hey!", 'Risk '+str(i+1)+': '+risk_tsk[i]+" has no distribution.")
             
                

            elif risk_dis[i] == "Pert":
                if risk_sim[i][0]=='' or risk_sim[i][1]=='' or risk_sim[i][2]=='' :
                    messagebox.showerror("Hey!", 'Risk '+str(i+1)+': '+risk_tsk[i]+" \n has missing input data, please go to excel file and input the missing data.")

                else:
                    
                    if risk_sim[i][0]>0:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle('Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Pert Distribution & Risk Impact Cost",
    
                                     fontsize=12)
    
    
    
                        host = pyplot.subplot(121)
    
                        a = risk_sim[i][0]
    
                        b = risk_sim[i][1]
    
                        c = risk_sim[i][2]
    
    
    
                        x = PERT(a, b, c)
    
                        x.plot()
    
                        x.plot(hist=True)
    
    
    
                        host.set_ylabel("Probability Density")
    
                        host.set_xlabel("Pert Distribution")
    
    
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host.xaxis.set_major_formatter(tick)
    
    
    
                        #pyplot.title('Risk ' + str(i) + ': ' + risk_tsk[i]+"\n"+"Pert Distribution")
    
                        host1 = pyplot.subplot(122)
    
                        data_pert = pert_distro(a, b, c, 10000)
    
                        s = np.random.binomial(1, risk_bin[i], 10000)
    
                        data = s *  data_pert
    
                        host1.set_ylabel("Relative Frequency, %")
    
                        host1.set_xlabel("Risk Impact Cost")
    
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
    
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host1.xaxis.set_major_formatter(tick)
    
                        pyplot.show()
    
                    else:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle('Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Pert Distribution & Risk Impact Cost",
    
                                     fontsize=12)
    
                        host = pyplot.subplot(121)
    
                        a = risk_sim[i][2]
    
                        b = risk_sim[i][1]
    
                        c = risk_sim[i][0]
    
    
    
                        #x = PERT(a, b, c)
                        x = pert_distro(c,b, a, 10000)
                        sns.kdeplot(x, data2=None, shade=True, vertical=False, kernel='gau', bw='scott',
                                        gridsize=100, cut=3, clip=None, legend=True, cumulative=False, shade_lowest=True, cbar=False, cbar_ax=None, cbar_kws=None, ax=None,color="b")
                        
    
                        
    
    
    
                        host.set_ylabel("Probability Density")
    
                        host.set_xlabel("Pert Distribution")
                        
                        host.set_ylabel("Probability Density")
                        host.set_xlabel("Task Duration, days")
                        
    
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host.xaxis.set_major_formatter(tick)
    
                        host1 = pyplot.subplot(122)
    
                        #data_pert = pert_distro(c,b, a, 10000)
    
                        s = np.random.binomial(1, risk_bin[i], 10000)
    
                        data = s * x
    
                        host1.set_ylabel("Relative Frequency, %")
    
                        host1.set_xlabel("Risk Impact Cost")
    
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
    
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host1.xaxis.set_major_formatter(tick)
    
                        pyplot.show()

            elif risk_dis[i] == "Triangular":
                if risk_sim[i][0]=='' or risk_sim[i][1]=='' or risk_sim[i][2]=='' :
                    messagebox.showerror("Hey!", 'Risk '+str(i+1)+': '+risk_tsk[i]+" \n has missing input data, please go to excel file and input the missing data.")
                else:
                    

                    if risk_sim[i][0]>0:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle('Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Triangular Distribution & Risk Impact Cost",
    
                                     fontsize=12)
    
                        host = pyplot.subplot(121)
    
                        host.set_ylabel("Probability Density")
    
                        host.set_xlabel("Triangular Distribution")
    
                        c = (risk_sim[i][1] - risk_sim[i][0]) / (risk_sim[i][2] - risk_sim[i][0])
    
                        mean, var, skew, kurt = triang.stats(c, moments='mvsk')
    
                        x = np.linspace(risk_sim[i][0], risk_sim[i][2], 1000)
    
                        host.plot(x, triang.pdf(x, c, loc=risk_sim[i][0], scale=risk_sim[i][2] - risk_sim[i][0]), 'b', lw=2)
    
                        pyplot.fill(x, triang.pdf(x, c, loc=risk_sim[i][0], scale=risk_sim[i][2] - risk_sim[i][0]))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host.xaxis.set_major_formatter(tick)
    
                        host1 = pyplot.subplot(122)
    
                        data_trian = np.random.triangular(risk_sim[i][0], risk_sim[i][1], risk_sim[i][2], 10000)
    
                        s = np.random.binomial(1,risk_bin[i], 10000)
    
                        data = s * data_trian
    
                        host1.set_ylabel("Relative Frequency, %")
    
                        host1.set_xlabel("Risk Impact Cost")
    
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
    
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host1.xaxis.set_major_formatter(tick)
    
                        pyplot.show()
    
                    else:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle(
    
                            'Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Triangular Distribution & Risk Impact Cost",
    
                            fontsize=12)
    
                        host = pyplot.subplot(121)
    
                        host.set_ylabel("Probability Density")
    
                        host.set_xlabel("Triangular Distribution")
    
                        c = (risk_sim[i][1] - risk_sim[i][2]) / (risk_sim[i][0] - risk_sim[i][2])
    
                        mean, var, skew, kurt = triang.stats(c, moments='mvsk')
    
                        x = np.linspace(risk_sim[i][2], risk_sim[i][0], 1000)
    
                        host.plot(x, triang.pdf(x, c, loc=risk_sim[i][2], scale=risk_sim[i][0] - risk_sim[i][2]), 'b', lw=2)
    
                        pyplot.fill(x, triang.pdf(x, c, loc=risk_sim[i][2], scale=risk_sim[i][0] - risk_sim[i][2]))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host.xaxis.set_major_formatter(tick)
    
                        host1 = pyplot.subplot(122)
    
                        data_trian = np.random.triangular(risk_sim[i][2], risk_sim[i][1], risk_sim[i][0], 10000)
    
                        s = np.random.binomial(1, risk_bin[i], 10000)
    
                        data = s * data_trian
    
                        host1.set_ylabel("Relative Frequency, %")
    
                        host1.set_xlabel("Risk Impact Cost")
    
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
    
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host1.xaxis.set_major_formatter(tick)
    
                        pyplot.show()

            elif risk_dis[i] == "Poisson":
                if risk_sim[i][3]=='' :
                    messagebox.showerror("Hey!", 'Risk '+str(i+1)+': '+risk_tsk[i]+" \n has missing input data, please go to excel file and input the missing data.")

                else:
                    
                    if risk_sim[i][3]>0:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle('Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Poisson Distribution & Risk Impact Cost", fontsize=12)
    
                        host = pyplot.subplot(121)
    
                        host.set_ylabel("Probability Density, %")
    
                        host.set_xlabel("Poisson Distribution")
    
                        m=risk_sim[i][3]
    
                        data_poisson = poisson.rvs(mu=m, size=10000)
                        sns.distplot(data_poisson,
                                     bins=30,
                                     kde=False,
                                     color='skyblue',
                                     hist_kws={"linewidth": 15, 'alpha': 1})
                        host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data_poisson), symbol=''))
    
                        host1 = pyplot.subplot(122)
    
                        data_trian = np.random.triangular(risk_sim[i][0],risk_sim[i][1], risk_sim[i][2], 10000)
    
                        s = np.random.poisson(m, 10000)
    
                        data = s * data_trian
    
                        host1.set_ylabel("Relative Frequency, %")
    
                        host1.set_xlabel("Risk Impact Cost")
    
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
    
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
    
                        fmt = '${x:,.0f}'
    
                        tick = mtick.StrMethodFormatter(fmt)
    
                        host1.xaxis.set_major_formatter(tick)
    
                        pyplot.show()
    
                    else:
    
                        fig = pyplot.figure()
    
                        pyplot.clf()
    
                        fig.suptitle(
    
                            'Risk ' + str(i+1) + ': ' + risk_tsk[i] + "\n" + "Poisson Distribution & Risk Impact Cost",
    
                            fontsize=12)
    
                        host = pyplot.subplot(121)
                        host.set_ylabel("Probability Density, %")
                        host.set_xlabel("Poisson Distribution")
                        m = risk_sim[i][3]
                        data_poisson = poisson.rvs(mu=m, size=10000)
                        sns.distplot(data_poisson,
                                     bins=30,
                                     kde=False,
                                     color='skyblue',
                                     hist_kws={"linewidth": 15, 'alpha': 1})
                        host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data_poisson), symbol=''))
                        host1 = pyplot.subplot(122)
                        data_trian = np.random.triangular(risk_sim[i][2], risk_sim[i][1], risk_sim[i][0], 10000)
                        s = np.random.poisson(m, 10000)
                        data = s * data_trian
                        host1.set_ylabel("Relative Frequency, %")
                        host1.set_xlabel("Risk Impact Cost")
                        host1.hist(data, bins=50, alpha=0.6, color='blue')
                        host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
                        fmt = '${x:,.0f}'
                        tick = mtick.StrMethodFormatter(fmt)
                        host1.xaxis.set_major_formatter(tick)
                        pyplot.show()
        
        def task_stats_risk (task_no,risk_sim,risk_dis,risk_tsk,risk_bin):

            i = task_no
            if i > len(risk_dis)-1 or i < 0:
                messagebox.showwarning("Hey!", "This risk lies outside of the range of specified.")
            else:    
                mcerp.npts = int(entry_simul.get())
                title = 'Risk ' + str(i+1) + ': ' + risk_tsk[i]
                if risk_dis[i] == "":
                    messagebox.showerror("Hey!",'Risk '+str(i)+': '+risk_tsk[i]+ " has no distribution.")
                elif risk_dis[i] == "Pert":
                    a = risk_sim[i][0]
                    b = risk_sim[i][1]
                    c = risk_sim[i][2]
                    data_pert=pert_distro(a,b,c,int(entry_simul.get()))
                    data=data_pert*risk_bin[i]
                    make_table(data, title)
                elif risk_dis[i] == "Triangular":   
                    data_tri=tri_distro(risk_sim[i][0], risk_sim[i][1], risk_sim[i][2], int(entry_simul.get()))
                    data=data_tri*risk_bin[i]
                    make_table(data, title)
                elif risk_dis[i]=="Poisson":
                    data_poi=pois_distro(risk_sim[i][3], int(entry_simul.get()))
                    data_tri=tri_distro(risk_sim[i][0],risk_sim[i][1],risk_sim[i][2],int(entry_simul.get()))
                    data=data_tri*data_poi
                    make_table(data, title)
                    
        def write_to_excel(tsk, data, data_set):
            messagebox.showwarning("Hey","1. The iteration data are stored in new Excel files titled 'Project Cost Without Risk Output Data' and"
                                   " 'Project Cost With Risk Output Data' inside the source folder. \n \n2. "
                                   "Before you run the next simulation, please make sure Excel files are closed,"
                                   " so that the new iteration data can override the old iteration data.")

   
            wb = xlsxwriter.Workbook('Project Cost Without Risk Output Data.xlsx')
            ws = wb.add_worksheet()
            A1 = 'Output Data'
            title = np.append(A1, tsk)
        
            item = title
            column = 0
            for i in item:
                ws.write(0, column, i)
                column += 1
            row = 1
        
            for i in data:
                ws.write(row, 0, i)
                row += 1
        
            col = 1
            
            for i in data_set:
                
                r = 1
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    r += 1
                col += 1
        
            wb.close()
           
        def write_to_excel_risk(tsk, tsk_risk, total_data, data_set, data_set_risk):

            wb = xlsxwriter.Workbook('Project Cost With Risk Output Data.xlsx')
        
            ws = wb.add_worksheet()
        
            A1 = 'Total output'
        
            title = np.append(A1, tsk + tsk_risk)
            item = title
        
            column = 0
        
            for i in item:
        
                ws.write(0, column, i)
        
                column += 1
            row = 1
        
            #total_data = data + data_risk
        
            for i in total_data:
        
                ws.write(row, 0, i)
        
                row += 1
            col = 1
        
            for i in data_set + data_set_risk:
        
                r = 1
        
                for j in i:
        
                    # print('data: %f, r:%d'%(j,r))
        
                    ws.write(r, col, j)
        
        
        
                    r += 1
        
                col += 1
            wb.close()
        def savetoflie():
             write_to_excel(task, new_data, cost_simul_data)
        
             write_to_excel_risk(task, risk_tsk, ccc, cost_simul_data, cost_simul_data_risk)
    
        top_frame = tk.Frame(self)
    
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
    
        center_frame = tk.Frame(self)
    
        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)
    
        bottom_frame = tk.Frame(self)
    
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)
    
        label_1 = tk.Label(center_frame, text="Project Cost with and without Risk | Simulation Settings", font=HEADR_FONT)
    
        label_1.pack()
    
        message_1 = tk.Message(center_frame, text="Check the distributions of your individual tasks to see that "
    
                               "they are correct.\nThen input 'Number of Iterations' and click on 'Run Simulation'.",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_1.pack()
        message_2 = tk.Message(center_frame, text="",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_2.pack()
    
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda:clear())
    
        button_1.pack(side="right", padx=5)
    
    
    
        entry_frame = tk.Frame(center_frame)
    
        entry_frame.pack()
    
        button_2 = ttk.Button(entry_frame, text="Task Distribution",
    
                              command=lambda: check_distro(int(entry_task.get())-1,simul, distro, task))
    
        button_2.grid(row=0, column=1)
    
        button_4 = ttk.Button(entry_frame, text="Task Statistics",command=lambda:task_stats(int(entry_task.get())-1,simul, distro, task))
    
        button_4.grid(row=0, column=2)
    
        button_5 = ttk.Button(entry_frame, text="Risk Distribution",command=lambda:check_distro_risk(int(entry_risk.get())-1,risk_sim,risk_dis,risk_tsk,risk_bin))
    
        button_5.grid(row=1, column=1)
    
        button_6 = ttk.Button(entry_frame, text="Risk Statistics",command=lambda:task_stats_risk(int(entry_risk.get())-1,risk_sim,risk_dis,risk_tsk,risk_bin))
    
        button_6.grid(row=1, column=2)
    
        entry_simul = ttk.Entry(entry_frame, width=8)
    
        entry_simul.grid(row=2, column=0)
    
        entry_simul.insert(0, "100")
    
    
        entry_task = ttk.Entry(entry_frame, width=8)
    
        entry_task.grid(row=0, column=0)
    
        entry_risk = ttk.Entry(entry_frame, width=8)
    
        entry_risk.grid(row=1, column=0)
    
        entry_task.insert(0,"1")
    
        entry_risk.insert(0, "1")
    
        label_sims = tk.Label(entry_frame, text="Number of Iterations")
    
        label_sims.grid(row=2, column=1, columnspan=2)
    
        separate = ttk.Separator(entry_frame, orient="horizontal")
    
        separate.grid(row=4, column=0, columnspan=3, sticky="ew", pady=5)
        message_4 = tk.Message(entry_frame, text="                    "
                               "                                                  ",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_4.grid()
    
        label_2 = tk.Label(entry_frame, text="Output Display Without Risk:")
    
        label_2.grid(row=6, column=0, columnspan=3)
    
        button_3 = ttk.Button(entry_frame, text="Run Simulation", command=lambda:simule(int(entry_simul.get())))
    
        button_3.grid(row=3, column=0, columnspan=3, sticky="ew")
    
        b_pc = ttk.Button(entry_frame, text="PDF & CDF Plot",command=lambda: make_graph(new_data, "Project Cost without Risk",1,1))
    
        b_pc.grid(row=7, column=0)
    
        b_st = ttk.Button(entry_frame, text="Statistics",command=lambda: make_table(new_data, "Project Cost without Risk Statistics") )
    
        b_st.grid(row=7, column=1)
        
        b_tornado1 = ttk.Button(entry_frame, text="Tornado Diagram",command=lambda: plot())

        b_tornado1.grid(row=7, column=2)

        separate2 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate2.grid(row=8, column=0, columnspan=3, sticky="ew", pady=5)
        message_5 = tk.Message(entry_frame, text="                    "
                               "                                                  ",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_5.grid()
    
        label_3 = tk.Label(entry_frame, text="Output Display With Risk:")
    
        label_3.grid(row=10, column=0, columnspan=3)
        
        b_pc2 = ttk.Button(entry_frame, text="PDF & CDF Plot",command=lambda: make_graph(ccc, "Project Cost With Risk",1,1))
    
        b_pc2.grid(row=11, column=0)
    
        b_st2 = ttk.Button(entry_frame, text="Statistics",command=lambda: make_table(ccc, "Project Cost With Risk Statistics") )
    
        b_st2.grid(row=11, column=1)
        
        b_tornado = ttk.Button(entry_frame, text="Tornado Diagram",command=lambda: plot_risk())

        b_tornado.grid(row=11, column=2)
        separate4 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate4.grid(row=12, column=0, columnspan=3, sticky="ew", pady=5)
        message_6 = tk.Message(entry_frame, text="                    "
                               "                                                  ",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_6.grid()
    
        label_5 = tk.Label(entry_frame, text="Output Display Comparing with VS.without Risks :")
    
        label_5.grid(row=13, column=0, columnspan=3)
        
        b_pc3 = ttk.Button(entry_frame, text="PDF Plots",command=lambda: make_risk_graph(new_data,ccc, "Project Cost",0,1))
    
        b_pc3.grid(row=14, column=1)
        b_pp3 = ttk.Button(entry_frame, text="CDF Plots",command=lambda: make_risk_graph(new_data,ccc, "Project Cost",1,0))
    
        b_pp3.grid(row=14, column=2)
        separate3 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate3.grid(row=15, column=0, columnspan=3, sticky="ew", pady=5)
        
        message_3 = tk.Message(entry_frame, text="                    "
                               "                                                  ",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_3.grid()
        
    
        label_4 = tk.Label(entry_frame, text="    ")
    
        label_4.grid(row=16, column=0, columnspan=3)
       
    
    
        b_write_excel_risk = ttk.Button(entry_frame, text="Output Data", command=lambda: savetoflie())
    
        b_write_excel_risk.grid(row=17, column=0, columnspan=3, sticky="ew")

################################################################################################################################    
class ProSchedule(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        global directory
                 
        def get_string():
            global directory
            directory = entry_1.get()
            if directory == "":
                messagebox.askretrycancel("Error", "An error has occurred in your File.")
            else:
                global name,task,distr,predss,t,mim,maxm,name_rt,mod_rt,distr_rt,task_rt,rt,mim_rt,maxm_rt,event_rt,name_ro,mod_ro,distr_ro,task_ro,ro,mim_ro,maxm_ro,event_ro
                name,task,distr,predss,t,mim,maxm,name_rt,mod_rt,distr_rt,task_rt,rt,mim_rt,maxm_rt,event_rt,name_ro,mod_ro,distr_ro,task_ro,ro,mim_ro,maxm_ro,event_ro =upload_file(directory)               
                controller.show_frame(ProScheduleOptions)
        def clear_val():
            
            global directory,name,task,distr,predss,t,mim,maxm,name_rt,mod_rt,distr_rt,task_rt,rt,mim_rt,maxm_rt,event_rt,name_ro,mod_ro,distr_ro,task_ro,ro,mim_ro,maxm_ro,event_ro
            
            if directory!="":
                name=[]
                task=[]
                distr=[]
                predss=[]
                t=0
                mim=[]
                maxm=[]
                name_rt=[]
                mod_rt=[]
                distr_rt=[]
                task_rt=[]
                rt=0
                mim_rt=[]
                maxm_rt=[]
                event_rt=[]
                name_ro=[]
                mod_ro=[]
                distr_ro=[]
                task_ro=[]
                ro=0
                mim_ro=[]
                maxm_ro=[]
                event_ro=[]
                
                del directory,name,task,distr,predss,t,mim,maxm,name_rt,mod_rt,distr_rt,task_rt,rt,mim_rt,maxm_rt,event_rt,name_ro,mod_ro,distr_ro,task_ro,ro,mim_ro,maxm_ro,event_ro
            controller.show_frame(ChoosePage)
            entry_1.delete(0, 'end')


        top_frame = tk.Frame(self)
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)
        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)

        label_1 = tk.Label(center_frame, text="Project Schedule | File Upload", font=HEADR_FONT)
        label_1.pack()
        message_1 = tk.Message(center_frame, text="Select the MS Excel file with your project schedule data (.xlsx)\n",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        entry_1 = ttk.Entry(center_frame, width=50)
        entry_1.pack(side="left")
        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))
        button_2.pack(side="left", pady=5)
        button_3 = ttk.Button(center_frame, text="Open", command=lambda: get_string())
        button_3.pack(side="left", padx=5, pady=5)
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())
        button_1.pack(side="right", padx=5)

class ProScheduleOptions(tk.Frame):
    def __init__(self, parent, controller):
        
        tk.Frame.__init__(self, parent)
        '''
        def gerneral_cdf(xk,pk,itera):
            
            custm = stats.rv_discrete(name='custm', values=(xk, pk))
            g=custm.rvs(size=itera)
            return g
        '''
       

        def gerneral_cdf(xk,pk,itera):
            pk=np.array(pk)*itera
            custm=[]
            for i in range(len(pk)):
                bu=np.ones(int(pk[i]))*xk[i]
                #print(bu)
                custm=custm+bu.tolist()
            shuffle(custm)
            return custm
        
        def effect(prob, relate, prob2,relate2,itera):
            b1=np.zeros((len(prob),itera))
            b2=np.zeros((len(prob2),itera))
            
            for i in range(0,len(prob)):
                if prob[i]=="N":
                    b1[i]=b1[i]+np.ones(itera)
                    
                elif prob[i]!="N":
                    if int(relate[i])==0:
                        b1[i]=b1[i]+gerneral_cdf(np.arange(2),[1-prob[i],prob[i]],itera)
                    else:
                        z=int(relate[i])
                        pk=[prob[i],prob2[z-1],1-prob[i]-prob2[z-1]]
                        cm=gerneral_cdf(np.arange(3),pk,itera)
                        for j in range(len(cm)):
                            if cm[j] ==0:
                                b1[i][j]=1
                            elif cm[j] ==1:
                                b2[z-1][j]=1
            b1=b1.tolist()
            b2=b2.tolist()
            return b1,b2         
                    
         
        
        def upload_riskfile(director,itera):
            book=xlrd.open_workbook(director)
            sheet=book.sheet_by_index(0)
            t=int(sheet.cell_value(0,0))
            rt=int(sheet.cell_value(t+3,0))
            ro=int(sheet.cell_value(t+3+rt+1,0))
            
            prob_rt=[]
            relate_rt=[]
            affect_rt=[]
           
            
            for i in range(t+3,t+3+rt):
                prob_rt.append(sheet.cell_value(i+1,7))
                relate_rt.append(sheet.cell_value(i+1,9))
                affect_rt.append(sheet.cell_value(i+1,10))
            
            prob_ro=[]
            relate_ro=[]
            affect_ro=[]
           
            for i in range(t+3+rt+1,t+3+rt+1+ro):
                prob_ro.append(sheet.cell_value(i+1,7))
                relate_ro.append(sheet.cell_value(i+1,9))
                affect_ro.append(sheet.cell_value(i+1,10))
                
            b1,b2=effect(prob_rt,relate_rt,prob_ro,relate_ro,itera)
            
            return b1,b2 ,affect_rt,affect_ro   
            
            
        def simulation(task,distr, t, mim, maxm,itera):
           
            no_risk_duration=[]
            for i in range(0, itera):
                
                no_risk_duration.append([])
                
                
                for j in range(0, t):
                    x = 0
                    
                    no_risk_duration[i].append(x)
            
                    
            for i in range(0,t):
           
                if distr[i] == "Triangular" :
                    a=tri_distro(mim[i], task[i], maxm[i], itera)
                elif distr[i] == "Pert":
                    a=pert_distro(mim[i], task[i], maxm[i], itera)
                    
                elif distr[i] == "Normal":
                    a=norm_distro(mim[i], task[i], maxm[i], itera)
                elif distr[i] == "Uniform":
                    a=uni_distro(mim[i], maxm[i], itera)
                
                
                else:
                    a = no_distro(task[i], itera)
               # print(a) 
            
                for j in range(0,itera):
                    
                    no_risk_duration[j][i] = a[j] 
            return no_risk_duration 
          
        def risk_schedul_simulation(task,distr, t, mim, maxm,event,b,itera):
           
            risk_duration=[]
           
            for i in range(0, itera):
                
                risk_duration.append([])
                
                
                for j in range(0, t):
                    x = 0
                    
                    risk_duration[i].append(x)
                    
            
                    
            for i in range(0,t):
                
                m=0
           
                if distr[i] == "Triangular" :
                    a=tri_distro(mim[i], task[i], maxm[i], itera)*b[i]
                elif distr[i] == "Pert":
                    a=pert_distro(mim[i], task[i], maxm[i], itera)*b[i]
                    
                elif distr[i] == "Normal":
                    a=norm_distro(mim[i], task[i], maxm[i], itera)*b[i]
                elif distr[i] == "Uniform":
                    a=uni_distro(mim[i], maxm[i], itera)*b[i]
                elif distr[i] == "Poisson" :
                    for j in range(0,itera):
                        
                        if int(b[i][j])<2:
                            
                            x =pert_distro(mim[i], task[i], maxm[i],itera)
                            
                            a = x * np.random.poisson(event[i],itera)*b[i][j]
                        else:
                            for m in range(0,b[i][j]):
                               x +=pert_distro(mim[i], task[i], maxm[i],itera) 
                               m+=1 
                   
                
                else:
                    if task[i]==0:
                        a=no_distro(1, itera)*b[i]
                    else:
                        a = no_distro(task[i], itera)*b[i]
               
                
                for j in range(0,itera):
                    
                    risk_duration[j][i] = a[j] 
                    
                        
                        
            return risk_duration
        
        def write_to_excel2(tsk, data, data_set):
            messagebox.showwarning("Hey","1. The iteration data are stored in new Excel files titled 'Schedule Output Data' and"
                                   " 'Risk Schedule Output Data' inside the source folder. \n \n2. "
                                   "Before you run the next simulation, please make sure Excel files are closed,"
                                   "so that the new iteration data can override the old iteration data.")

   
            wb = xlsxwriter.Workbook('Schedule Output Data.xlsx')
            ws = wb.add_worksheet()
            A1 = 'Output Data'
            title = np.append(A1, tsk)
        
            item = title
            column = 0
            for i in item:
                ws.write(0, column, i)
                column += 1
            row = 1
        
            for i in data:
                ws.write(row, 0, i)
                row += 1
        
            r = 1
            
            for i in data_set:
                i.remove(i[0])
                col = 1
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    col += 1
                r += 1
        
            wb.close()
        def risk_result(durart,duraro):
            wb = xlsxwriter.Workbook('Risk  Output Data.xlsx')
            ws = wb.add_worksheet()
            title=['Application is rejected', 'Trial 1 of prototype testing is a total failure.','Supply chain disruptions are caused by a hurricane(s).','Application is approved quickly.',
                   'Trial 1 of prototype testing is a total success.']
            item = title
            data_set=np.transpose(np.transpose(np.array(durart)).tolist()+np.transpose(np.array(duraro)).tolist()).tolist()
            column = 0
            for i in item:
                ws.write(0, column, i)
                column += 1
            
        
            
        
            r = 1
            
            for i in data_set:
                
                col=0
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    col += 1
                r += 1
        
            wb.close()
        def risk_write_to_excel(tsk, data, data_set):
            wb = xlsxwriter.Workbook('Risk Schedule Output Data.xlsx')
            ws = wb.add_worksheet()
            A1 = 'Output Data'
            title = np.append(A1, tsk)
        
            item = title
            column = 0
            for i in item:
                ws.write(0, column, i)
                column += 1
            row = 1
        
            for i in data:
                ws.write(row, 0, i)
                row += 1
        
            r = 1
            
            for i in data_set:
                i.remove(i[0])
                col = 1
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    col += 1
                r += 1
        
            wb.close()
        
        
        def cpm(task,depen,t):
            A=[]
            for i in range(t+1):
                A.append(-1)
        
            i=2
            A[0]=0
            A[1]=task[1]
            while (i<t+1):
                #print("in loop",i)
                for j in range(2,t+1):
                    if A[j]!=-1:
                        continue
        
                    else:
                        d = depen[j]       #d also becomes an array cause depend was 2D array
                        #print(d)
                        f=0
                        A1=[]
                        for k in d:
                            if A[k] == -1:
                                f=1
                                break
                            else:
                                A1.append(A[k])
                        #print(A)
                        if f==0:
        
                            A[j] = max(A1) + task[j]
                            i=i+1
            #print("out of loop")
            total= max(A)
        
            A.pop(0)
            return total
        def cpmno(task, depen, t):   # CPM but no printing just for evaluating
            A=[]
            for i in range(t+1):
                A.append(-1)
        
            i=2
            A[0]=0
            A[1]=task[1]
            while (i<t+1):
                #print("in loop",i)
                for j in range(2,t+1):
                    if A[j]!=-1:
                        continue
        
                    else:
                        d = depen[j]       # d also becomes an array cause depend was 2D array
                        # print(d)
                        f=0
                        A1=[]
                        for k in d:
                            if A[k] == -1:
                                f=1
                                break
                            else:
                                A1.append(A[k])
                        #print(A)
                        if f==0:
        
                            A[j] = max(A1) + task[j]
                            i=i+1
            #print("out of loop")
            return max(A)
        def impact_on(t, risk_t,impact,duration,riskduration, affect_rt,itera):
            
            a=0
            rduration=[]
            for i in range(0, itera):
                
                rduration.append([])
                
                
                for j in range(0, t):
                    x = 0
                    
                    rduration[i].append(x)
            
            for i in range(0,t):
                for j in range(0,itera):
                    
                    a=duration[j][i]
                    rduration[j][i]=a
                    if (i+1) in impact:
                        for k in range(0,risk_t):
                            
                            if round(impact[k])==i+1:
                                b=riskduration[j][k]
                                
                                
                                if affect_rt[k]== "ADD":
                                    rduration[j][i]=a+b
                                elif affect_rt[k]== "REPLACE":
                                    if b!=0:
                                        
                                        rduration[j][i]=b
                                elif affect_rt[k]== "MINUS":
                                    rduration[j][i]=a-b
                                
                               
                    
                        
                    
            return rduration
        
        def impact_on2(t, risk_t,impact,duration,riskduration, affect_ro,itera):
            a=0
            rduration=[]
            for i in range(0, itera):
                
                rduration.append([])
                
                
                for j in range(0, t):
                    x = 0
                    
                    rduration[i].append(x)
            
            for i in range(0,t):
                for j in range(0,itera):
                    #print(impact)
                    a=duration[j][i]
                    rduration[j][i]=a
                    if (i+1) in impact:
                        for k in range(0,risk_t):
                           # print(round(impact[k]))
                            if round(impact[k])==i+1:
                                b=riskduration[j][k]
                                if affect_ro[k]== "ADD":
                                    rduration[j][i]=a+b
                                elif affect_ro[k]== "REPLACE":
                                    if b==1:
                                        
                                        rduration[j][i]=0
                                elif affect_ro[k]== "MINUS":
                                    rduration[j][i]=a-b
                                
                                
                        
                    
            return rduration
        
        
        def noriskduration(no_risk_duration,risk_duration,predss,t,itera):
            
            
        
            crit_index=[]  # stores index of critical tasks,  0 allowed here
            risk_crit_index=[]
            
               
            for i in range(0,t+1):
                crit_index.append(0)
                risk_crit_index.append(0)
            
            no_risk_cpm_array = []
            no_risk_task=[]
            no_risk_ciripath=[]
            risk_cpm_array = []
            risk_task=[]
            risk_ciripath=[]
         
            
            for i in range(0, t+1):  
                no_risk_ciripath.append([])
                risk_ciripath.append([])
                risk_task.append([])
                no_risk_task.append([])
         
            for i in range(0, itera):
                b1 = [0]
                b2 = [0]
               
                for k1 in risk_duration[i]:
                    b1.append(k1)
                risk_duration[i]=b1 
                
                for k2 in no_risk_duration[i]:
                    b2.append(k2)
                no_risk_duration[i]=b2
                c = cpm(b1, predss, t)
                
                d = cpm(b2, predss, t)
                #print(d)
                risk_cpm_array.append(c)
                no_risk_cpm_array.append(d) #CPM values for The simulations are
                
                for k in range(1,t+1):
                    no_risk_duration[i][k]=no_risk_duration[i][k]+1
                    totalcheck = cpmno(no_risk_duration[i],predss,t)
                    no_risk_duration[i][k]=no_risk_duration[i][k]-1
                    if (totalcheck > d):
                        no_risk_ciripath[k].append(d)   # cirtical path 
                        no_risk_task[k].append(no_risk_duration[i][k])
                        crit_index[k]=crit_index[k]+1
                    
                    risk_duration[i][k]=risk_duration[i][k]+1
                    risk_totalcheck = cpmno(risk_duration[i],predss,t)
                    risk_duration[i][k]=risk_duration[i][k]-1
                    if (risk_totalcheck > c):
                        risk_ciripath[k].append(c)   # cirtical path 
                        risk_task[k].append(risk_duration[i][k])
                        risk_crit_index[k]=risk_crit_index[k]+1
                        
            return crit_index, no_risk_cpm_array, no_risk_ciripath,risk_crit_index, risk_cpm_array, risk_ciripath,no_risk_task,risk_task
        
        

        def get_value(directory):
            global no_risk_duration,dura_rt,dura_ro,b1,b2,d1,d2,affect_rt,affect_ro
            global crit_index, no_risk_cpm_array, no_risk_ciripath,risk_crit_index, risk_cpm_array, risk_ciripath,no_risk_task,risk_task
            
        
           
            
            no_risk_duration=simulation(task,distr, t, mim, maxm, int(entry_simul.get()))
            
            b1,b2,affect_rt,affect_ro =upload_riskfile(directory,int(entry_simul.get()))
            dura_rt=risk_schedul_simulation(mod_rt,distr_rt, len(mod_rt), mim_rt, maxm_rt,event_rt,b1, int(entry_simul.get()))
            dura_ro=risk_schedul_simulation(mod_ro,distr_ro, len(mod_ro), mim_ro, maxm_ro,event_ro,b2, int(entry_simul.get())) 
            
            
            d1=impact_on(t, rt,task_rt,no_risk_duration,dura_rt,affect_rt, int(entry_simul.get()))
            d2=impact_on2(t, ro,task_ro,d1,dura_ro,affect_ro, int(entry_simul.get()))
            crit_index, no_risk_cpm_array, no_risk_ciripath,risk_crit_index, risk_cpm_array, risk_ciripath,no_risk_task,risk_task=noriskduration(no_risk_duration,d2,predss,t, int(entry_simul.get()))
            messagebox.showwarning("Hey!", "Simulation is complete. You may browse output now.")
        
       
       
        
        def make_risk_graph2(data, temp, string_1, i, j):
        
            #temp = data + data2
            if i > 0 and j == 0:
                pyplot.clf()
                host = pyplot.subplot(111)
                par1 = host.twinx()
                host.set_xlabel(string_1, fontsize=12)
                host.set_ylabel("Cumulative Probability", fontsize=12)
                host.hist(temp, 10000, density=True, cumulative=True, histtype='step', color='red', linewidth=2)
                par1.hist(data, 10000, density=True, cumulative=True, histtype='step', color='blue', linewidth=2)
                blue_line = lines.Line2D([], [], color="blue", label="Without Risk", linewidth=2)
                risk_line = lines.Line2D([], [], color="red", label="With Risk", linewidth=2)
                pyplot.legend(handles=[risk_line, blue_line], loc='upper left')
                fmt = '{x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
        
                def fix_hist_step_vertical_line_at_end(ax):
                    axpolygons = [poly for poly in ax.get_children() if isinstance(poly, patches.Polygon)]
                    for poly in axpolygons:
                        poly.set_xy(poly.get_xy()[:-1])
                fix_hist_step_vertical_line_at_end(host)
                fix_hist_step_vertical_line_at_end(par1)
            if j > 0 and i == 0:
                pyplot.style.use('classic')
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_xlabel(string_1, fontsize=12)
                host.set_ylabel("Relative Frequency, %", fontsize=12)
                host.hist(temp, bins=50, alpha=0.6, label="With Risk", color='red')
                host.hist(data, bins=50, alpha=0.6, label="Without Risk", color='blue')
                risk_patch = patches.Patch(alpha=0.6, color="red", label="With Risk")
                blue_patch = patches.Patch(alpha=0.6, color="blue", label="Without Risk")
                pyplot.legend(handles=[risk_patch, blue_patch])
                host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
                fmt = '{x:,.0f}'
                tick = mtick.StrMethodFormatter(fmt)
                host.xaxis.set_major_formatter(tick)
            pyplot.draw()
            pyplot.show()

        def make_graph2( data, string_1, i, j):
        
            """ Displays the primary data in graph from using matplotlib
            :param b: The number of bins to sort the data into.
            :param data: The data generated.
            :param string_1: The title of the graph.
            :param i: Boolean. Display CDF?
            :param j: Boolean. Display PDF?
            """
            if j > 0 and i == 0:
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_xlabel(string_1, fontsize=12)
                host.set_ylabel("Relative Frequency, %")
                host.hist(data, bins=50, alpha=0.6, color='blue')
                host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
            elif i > 0 and j == 0:
                pyplot.clf()
                host = pyplot.subplot(111)
                host.set_xlabel(string_1, fontsize=12)
                host.set_ylabel("Cumulative Probability", fontsize=12)
                #host.hist(data, bins=1000, density=True, cumulative=True, histtype='step', color='blue')
                ser = pd.Series(data)
                ax = ser.hist(bins=1000, density=True, cumulative=True, histtype='step', color='blue',grid=False)
                # how to delete the vertical line for cdf graph
                poly = ax.findobj(pyplot.Polygon)[0]
                vertices = poly.get_path().vertices
                # Keep everything above y == 0. You can define this mask however
                # you need, if you want to be more careful in your selection.
                keep = vertices[:, 1] > 0
                # Construct new polygon from these "good" vertices
                new_poly = pyplot.Polygon(vertices[keep], closed=False, fill=False, edgecolor=poly.get_edgecolor(),
                                          linewidth=poly.get_linewidth())
                poly.set_visible(False)
                ax.add_artist(new_poly)
            else:
                #pyplot.style.use('classic')
                pyplot.clf()
                host = pyplot.subplot(111)
                par1 = host.twinx()
                host.set_xlabel(string_1, fontsize=12)
                host.set_ylabel("Relative Frequency, %", fontsize=12)
                par1.set_ylabel("Cumulative Probability", fontsize=12)
                host.hist(data, bins=50, alpha=0.6, label="PDF", color='blue')
                #par1.hist(data, bins=1000, density=True, cumulative=True, histtype='step', color='blue')
                ser = pd.Series(data)
                bx = ser.hist(bins=1000, density=True, cumulative=True, histtype='step', color='blue',grid=False)
                # how to delete the vertical line for cdf graph
                poly = bx.findobj(pyplot.Polygon)[0]
                vertices = poly.get_path().vertices
                # Keep everything above y == 0. You can define this mask however
                keep = vertices[:, 1] > 0
                # Construct new polygon from these "good" vertices
                new_poly = pyplot.Polygon(vertices[keep], closed=False, fill=False, edgecolor=poly.get_edgecolor(),
                                          linewidth=poly.get_linewidth())
                poly.set_visible(False)
                bx.add_artist(new_poly)
                blue_line = lines.Line2D([], [], color="blue", label="CDF", linewidth=2.0)
                blue_patch = patches.Patch(alpha=0.6, color="blue", label="PDF")
                pyplot.legend(handles=[blue_line, blue_patch], loc='upper left')
                host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
            fmt = '{x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            host.xaxis.set_major_formatter(tick)
            pyplot.draw()
            pyplot.show()
            
        def run_stats1():
            """
            Checks to see if any data exists before displaying any statistics.
            :return: None.
            """
            
            make_table(no_risk_cpm_array, "Project Duration (Without Risks Events) Statistics ",' days','r')
        def run_stats2():
            """
            Checks to see if any data exists before displaying any statistics.
            :return: None.
            """
            
            make_table(risk_cpm_array, "Project Schedule (With Risks Events) Statistics ",' days','r')
        
        def tornado(no_risk_ciripath,no_risk_task,n,no_risk_cpm_array):
            avg_npv=np.average(no_risk_cpm_array)
            low = []
            high = []
            value = []
            name=[]
            def sort_second(val):
                return val[1]
            
            for i in range(1, (t+1)):
                x = no_risk_ciripath[i]
                s=no_risk_task[i]
                avs=np.mean(s)
                avx=np.mean(x)
                
                if round(np.average(s))==0:
                   i+1
                   
                   
                   
                   
                else:
                    a=round(0.05*len(s))
                    l=[]
                    h=[]
                    for j in range(0,a):
                        
                        minpos = s.index(min(s)) 
                        s[minpos]= avs
                        u=x[minpos]
                        x[minpos]= avx
                        maxpos = s.index(max(s))
                        s[maxpos]=avs
                        v=x[maxpos]
                        x[maxpos]=avx
                        l.append(u)
                        h.append(v)
        
                    v=np.average(l)
                    u=np.average(h)
                    if u>v:
                        low.append(int(v))
                        high.append(int(u))
                    elif u<=v:
                        low.append(int(u))
                        high.append(int(v))
                        
                    name.append(n[i-1])
                 
               
            
            mains = []
        
            for i in range(0, len(low)):
                v = float(high[i]) - float(low[i])
                value.append(v)
                x = (name[i], value[i], low[i], high[i])
                mains.append(x)
        
            mains.sort(key=sort_second, reverse=True)
        
        
        
            values = []
            variables = []
            lows = []
            highs=[]
            
            for i in range(0, len(mains)):
                x = (mains[i][0])
                variables.append(x)
            
            for i in range(0, len(mains)):
                y = mains[i][1]
                values.append(y)
            
            for i in range(0, len(mains)):
                z = mains[i][2]
                lows.append(z)
            for i in range(0, len(mains)):
                h = mains[i][3]
                highs.append(h)
        
        
        
        
        
            base = avg_npv
            return variables,base, values, lows, highs
       
        
        ############################################################################################################
        
        
        def tornado_withoutrisk(no_risk_ciripath,no_risk_task,name,no_risk_cpm_array):
            
            
        
            variables,base, values, lows, highs=tornado(no_risk_ciripath,no_risk_task,name,no_risk_cpm_array)
            
            ys = range(len(values))[::-1]
            
            
            for y, low, high,value in zip(ys, lows,highs, values):
                
               
                pyplot.broken_barh(
                   [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=['lightgreen'],
                    edgecolors=['white'],
                    linewidth=1,
                )
                x = base + value / 4
                if x <= base + 5:
                   x = base + value / 4+ 5
                #pyplot.text(x, y, str(round(value)), va='center', ha='center')
                pyplot.text(high+3, y, str(round(value)), va='center', ha='center')
            
            
            
            pyplot.axvline(base, color='black')
           
            
            axes = pyplot.gca()
           # axes.title("Tornado Graph Without Risk")
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '{x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            axes.xaxis.set_major_formatter(tick)
            
            pyplot.yticks(ys, variables)
            
            pyplot.xlim(base - 25, base + 25)
            pyplot.ylim(-1, len(variables))
            pyplot.title('Project Duration without Risks Events, days ')        
            base_money = '{:,.0f}'.format(round(base))
            pyplot.text(round(base)-5,y-0.15*len(variables),'Average = '+str(base_money)+' days', fontsize=12)
            pyplot.show()
        
        def tornado_risk(no_risk_ciripath,no_risk_task,name,no_risk_cpm_array):
            
            
        
            variables,base, values, lows, highs=tornado(no_risk_ciripath,no_risk_task,name,no_risk_cpm_array)
            
            ys = range(len(values))[::-1]
            
            
            for y, low, high,value in zip(ys, lows,highs, values):
                
               
                pyplot.broken_barh(
                   [(low,value)],
                    (y - 0.4, 0.8),
                    facecolors=['lightgreen'],
                    edgecolors=['white'],
                    linewidth=1,
                )
                x = base + value/4
                if x <= base + 50:
                   x = base + value/4+ 50
            
            
                #pyplot.text(high+5,y-0.2,str(round(high)),  color='lightgreen', fontweight='bold')
                #pyplot.text(x,y-0.2, str(round(value)),  va='center', ha='center')
                pyplot.text(high+6,y-0.2, str(round(value)),  va='center', ha='center')
            
            pyplot.axvline(base, color='black')
           
            
            axes = pyplot.gca()
           # axes.title("Tornado Graph Without Risk")
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            fmt = '{x:,.0f}'
            tick = mtick.StrMethodFormatter(fmt)
            axes.xaxis.set_major_formatter(tick)
            
            pyplot.yticks(ys, variables)
            
            pyplot.xlim(base - 100, base + 200)
            pyplot.ylim(-1, len(variables))
            pyplot.title('Project Duration with Risks Events, days ')
            base_money = '{:,.0f}'.format(round(base))
            pyplot.text(round(base)-30,y-0.15*len(variables),'Average = '+str(base_money)+' days', fontsize=12)
            #pyplot.xlabel('Average ='+str(base_money), fontsize=12)
            pyplot.show()

        
        
        ###########################################################################################
            
        def index_graph(no_risk_duration,no_risk_cpm_array,crit_index,itera,name):
            def sort_second(val):
                return val[1]
            
            def sort(name, critindex):
                 mains = []
                 for i in range(0, len(name)):
                    
                     x = (name[i],critindex[i])
                     mains.append(x)
        
                 mains.sort(key=sort_second, reverse=True)
                 values = []
                 variables = []
                
                 for i in range(0, len(mains)):
                     x = (mains[i][0])
                     variables.append(x)
                 for i in range(0, len(mains)):
                     y = mains[i][1]
                     values.append(y)
                 return variables,values
                 
            critindex=[]
            for i in range(0,t):
                
                critindex.append(crit_index[i+1]/itera)
                
            name1,critindex=sort(name, critindex)    
            ys = range(len(critindex))[::-1]
            
            
            for y, value in zip (ys,critindex):
                
               
                pyplot.broken_barh(
                   [(0,value)],
                    (y-0.2, 0.8),
                    facecolors=['blue'],
                    edgecolors=['white'],
                    linewidth=1,
                )
            
            
                pyplot.text(max(0,value)+0.1,y-0.2,str(round(value*100,2))+'%',  color='blue', fontweight='bold')
                
            
            pyplot.axvline(0, color='black')
            
            
           # axes.title("Critical Index for no risk task")
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            axes.set(ylim=(15, len(name1)), xlim=([0.0, 1.0]))
             
            pyplot.yticks(ys, name1)
            
            
            pyplot.title('Critical Index')
            pyplot.show()
            
        
        
        
            '''
            no_risk_corelation=[]
            no_risk_correl_coeff=[]
            #corelattion coefficient
            def column(matrix, i):    #extract i th column
                return [row[i] for row in matrix]
            
            for i in range(0,t):
                no_risk_corelation.append(spearmanr(column(no_risk_duration,i), no_risk_cpm_array))   #i+1 ,since 1st column of distret is 0,0,0...
                #no_risk_corelation.append(pearsonr(column(no_risk_duration,i+1), no_risk_cpm_array))
            
            
            for i in range(0,t):
                no_risk_correl_coeff.append(column(no_risk_corelation,0)[i])
                no_risk_correl_coeff=pd.Series(no_risk_correl_coeff).fillna(0).tolist()
                print("The Corelation coefficient for  no risk task %d are:- "%(i))
                print(column(no_risk_corelation,0)[i])
            name2,norisk_correl=sort(name, no_risk_correl_coeff)      
            ys = range(len(no_risk_correl_coeff))[::-1]
            
            for y, value in zip (ys,norisk_correl):
                
               
                pyplot.broken_barh(
                   [(0,value)],
                    (y-0.2, 0.8),
                    facecolors=['blue'],
                    edgecolors=['white'],
                    linewidth=1,
                )
            
            
                pyplot.text(max(0,value)+0.2,y-0.2,str(round(value,2)),  color='blue', fontweight='bold')
                
            
            pyplot.axvline(0, color='black')
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            axes.set(ylim=(15, len(name2)), xlim=([-1.0, 1.0]))
            
            pyplot.yticks(ys, name2)
            
            
            pyplot.title('Correlation Coefficient')
            pyplot.show()
            
            
            
            no_risk_cruciality_index=[]
            
            
            #print("The Crutiality index for the tasks are:- ")
            for i in range(0,t):
                
                no_risk_cruciality_index.append(no_risk_correl_coeff[i] * crit_index[i+1]/itera)
                
            name3,no_risk_cruciality_index=sort(name, no_risk_cruciality_index)      
            ys = range(len(no_risk_correl_coeff))[::-1]
            
            for y, value in zip (ys,no_risk_cruciality_index):
                
               
                pyplot.broken_barh(
                   [(0,value)],
                    (y-0.2, 0.8),
                    facecolors=['blue'],
                    edgecolors=['white'],
                    linewidth=1,
                )
            
            
                pyplot.text(max(0,value)+0.2,y-0.2,str(round(value*100,2))+'%',  color='blue', fontweight='bold')
                
            
            pyplot.axvline(0, color='black')
            axes = pyplot.gca()
            axes.spines['left'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_visible(False)
            axes.xaxis.set_ticks_position('top')
            axes.set(ylim=(15, len(name3)), xlim=([-1.0, 1.0]))
             
            pyplot.yticks(ys, name3)
            
            
            pyplot.title('Schedule Impact Indicator')
            pyplot.show()
            '''
        def savetoflie():
            write_to_excel2(name, no_risk_cpm_array, no_risk_duration)
            risk_write_to_excel(name, risk_cpm_array, d2)
            risk_result(dura_rt,dura_ro)
    
            
        def check_distro1(task_no,minim,mode,maxim, distro, task):
            
            i = task_no
            if i > len(distro)-1 or i < 0:
    
                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")
            elif distro[i]=='':
                messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + " has no distribution.")
            elif distro[i] == "Normal":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    mu = minim[i]
                    sigma = mode[i]
                    x = np.linspace(mu - 5 * sigma, mu + 5 * sigma, 100)
                    pyplot.plot(x, ss.norm.pdf(x, mu, sigma))
                    pyplot.fill(x, ss.norm.pdf(x, mu, sigma))
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("Task Duration, days")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Normal Distribution")
                    pyplot.show()
            
            elif distro[i] == "Pert":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                  
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    a = minim[i]
                    b = mode[i]
                    c = maxim[i]
                    
                    x = pert_distro(a,b, c, 10000)
                    sns.kdeplot(x, data2=None, shade=True, vertical=False, kernel='gau', bw='scott',
                                    gridsize=100, cut=3, clip=None, legend=True, cumulative=False, shade_lowest=True, cbar=False, cbar_ax=None, cbar_kws=None, ax=None,color="b")
                    
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("Task Duration, days")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Pert Distribution")
                    
                    pyplot.show()
                    

    
            elif distro[i] == "Triangular":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("Task Duration, days")
                    c = (mode[i] - minim[i]) / (maxim[i]-minim[i])
                    mean, var, skew, kurt = triang.stats(c, moments='mvsk')
                    x = np.linspace(minim[i], maxim[i], 1000)
                    host.plot(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]), 'b', lw=2)
                    pyplot.fill(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]))
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title( 'Task '+str(i+1)+': '+task[i]+"\n"+"Triangular Distribution")
                    pyplot.show()
                
            elif distro[i]=="Uniform":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("Task Duration, days")
                    a=minim[i]
                    b=mode[i]
                    mean, var, skew, kurt = ss.uniform.stats(moments='mvsk')
                    x = np.linspace(a, b, 100)
                    host.plot(x, ss.uniform.pdf(x, loc=a, scale=b-1), 'b', lw=2, alpha=0.6, label='uniform pdf')
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('Task ' + str(i+1) + ': ' + task[i]+"\n"+"Uniform Distribution")
                    pyplot.show()
    
    
       
        def task_stats1(task_no,minim,mode,maxim, distro, task):

            i = task_no
            
            if i > len(distro)-1or i < 0:

                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")
            else:
                
               
                mcerp.npts = int(entry_simul.get())
    
                title = 'Task ' + str(i+1) + ': ' + task[i]
    
                if distro[i]=='':
                    messagebox.showerror("Hey!", 'Task ' + str(i + 1) + ': ' + task[i] + " has no distribution.")
    
                elif distro[i] == "Normal":
    
                    data = np.random.normal(minim[i], mode[i], int(entry_simul.get()))
    
                    make_table(data, title,' days','r')
    
                elif distro[i] == "Pert":
    
                    a = minim[i]
    
                    b = mode[i]
    
                    c = maxim[i]
    
                    alpha = ((4 * b) + c - (5 * a)) / (c - a)
    
                    beta = ((5 * c) - a - (4 * b)) / (c - a)
    
                    x = np.random.beta(alpha, beta, int(entry_simul.get()))
    
                    data = a + (x * (c - a))
    
                    make_table(data, title,' days','r')
    
                elif distro[i] == "Triangular":
    
                    data = np.random.triangular(minim[i], mode[i], maxim[i], int(entry_simul.get()))
    
                    make_table(data, title,' days','r')
    
                elif distro[i] == "Uniform":
    
                    data = np.random.uniform(minim[i],mode[i], int(entry_simul.get()))
    
                    make_table(data, title,' days','r')
                    
        
        def check_distro_risk1 (risk_no,even,minim,mode,maxim,event,risk_dis,risk_tsk,impact_task):
            b1,b2,c1,c2= upload_riskfile(directory,10000)

            """

            Displays the distribution graph for the given input / task.

            :param task_no: The index number of the task in risk register to display.

            :return: None.

            """

            i = risk_no

            if i > len(risk_dis)-1 or i < 0:

                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")

            elif risk_dis[i]=='':
                messagebox.showerror("Hey!", 'Risk '+str(i+1)+': '+risk_tsk[i]+" has no distribution.")

            elif risk_dis[i] == "Pert":

                if minim[i]>0:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle("Risk Impact Duration & Pert Distribution"+ "\n" + 'Risk ' + str(i+1) + ': ' + risk_tsk[i] ,

                                 fontsize=12)



                    host = pyplot.subplot(122)

                    a = minim[i]

                    b = mode[i]

                    c = maxim[i]



                    x = PERT(a, b, c)

                    x.plot()

                    x.plot(hist=True)



                    host.set_ylabel("Probability Density")

                    host.set_xlabel("Pert Distribution")



                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host.xaxis.set_major_formatter(tick)



                    #pyplot.title('Risk ' + str(i) + ': ' + risk_tsk[i]+"\n"+"Pert Distribution")

                    host1 = pyplot.subplot(121)

                    data_pert = pert_distro(a, b, c, 10000)
                    if even=='rt':
                        s = b1[i]
                    elif even=='ro':
                        s = b2[i]
                  
                    data = s *  data_pert

                    host1.set_ylabel("Relative Frequency, %")

                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')

                    host1.hist(data, bins=50, alpha=0.6, color='blue')

                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host1.xaxis.set_major_formatter(tick)

                    pyplot.show()

                else:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle("Risk Impact Duration & Pert Distribution"+ "\n" + 'Risk ' + str(i+1) + ': ' + risk_tsk[i],

                                 fontsize=12)

                    host = pyplot.subplot(122)

                    a = maxim[i]

                    b = mode[i]

                    c = minim[i]



                    x = PERT(a, b, c)

                    x.plot()

                    x.plot(hist=True)



                    host.set_ylabel("Probability Density")

                    host.set_xlabel("Pert Distribution")



                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host.xaxis.set_major_formatter(tick)

                    host1 = pyplot.subplot(121)

                    data_pert = pert_distro(c,b, a, 10000)

                    if even=='rt':
                        s = b1[i]
                    elif even=='ro':
                        s = b2[i]

                    data = s * data_pert

                    host1.set_ylabel("Relative Frequency, %")

                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')

                    host1.hist(data, bins=50, alpha=0.6, color='blue')

                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host1.xaxis.set_major_formatter(tick)

                    pyplot.show()

            elif risk_dis[i] == "Triangular":

                if minim[i]>0:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle( "Risk Impact Duration & Triangular Distribution"+ "\n"+ 'Risk ' + str(i+1) + ': ' + risk_tsk[i] ,

                                 fontsize=12)

                    host = pyplot.subplot(122)

                    host.set_ylabel("Probability Density")

                    host.set_xlabel("Triangular Distribution")

                    c = (mode[i] - minim[i]) / (maxim[i] - minim[i])

                    mean, var, skew, kurt = triang.stats(c, moments='mvsk')

                    x = np.linspace(minim[i], maxim[i], 1000)

                    host.plot(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i]- minim[i]), 'b', lw=2)

                    pyplot.fill(x, triang.pdf(x, c, loc=minim[i][0], scale=maxim[i] - minim[i]))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host.xaxis.set_major_formatter(tick)

                    host1 = pyplot.subplot(121)

                    data_trian = np.random.triangular(minim[i], mode[i], maxim[i], 10000)

                    #s = np.random.binomial(1,risk_bin[i], 10000)
                    if even=='rt':
                        s = b1[i]
                    elif even=='ro':
                        s = b2[i]

                    data = s * data_trian

                    host1.set_ylabel("Relative Frequency, %")

                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')

                    host1.hist(data, bins=50, alpha=0.6, color='blue')

                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host1.xaxis.set_major_formatter(tick)

                    pyplot.show()

                else:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle(

                        "Risk Impact Duration & Triangular Distribution"+ "\n"+ 'Risk ' + str(i+1) + ': ' + risk_tsk[i],

                        fontsize=12)

                    host = pyplot.subplot(122)

                    host.set_ylabel("Probability Density")

                    host.set_xlabel("Triangular Distribution")

                    c = (mode[i] - maxim[i]) / (minim[i] - maxim[i])

                    mean, var, skew, kurt = triang.stats(c, moments='mvsk')

                    x = np.linspace(risk_sim[i][2], risk_sim[i][0], 1000)

                    host.plot(x, triang.pdf(x, c, loc=maxim[i], scale=minim[i] - maxim[i]), 'b', lw=2)

                    pyplot.fill(x, triang.pdf(x, c, loc=maxim[i], scale=minim[i] - maxim[i]))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host.xaxis.set_major_formatter(tick)

                    host1 = pyplot.subplot(121)

                    data_trian = np.random.triangular(maxim[i], mode[i], minim[i], 10000)

                    if even=='rt':
                        s = b1[i]
                    elif even=='ro':
                        s = b2[i]

                    data = s * data_trian

                    host1.set_ylabel("Relative Frequency, %")

                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')

                    host1.hist(data, bins=50, alpha=0.6, color='blue')

                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host1.xaxis.set_major_formatter(tick)

                    pyplot.show()

            elif risk_dis[i] == "Poisson":

                if event[i]>0:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle("Risk Impact Duration & Poisson Distribution"+ "\n"+ ' Risk ' + str(i+1) + ': ' + risk_tsk[i]  , fontsize=12)

                    host = pyplot.subplot(122)

                    host.set_ylabel("Probability Density, %")

                    host.set_xlabel("The number of events")

                    m=event[i]

                    data_poisson = poisson.rvs(mu=m, size=10000)
                    sns.distplot(data_poisson,
                                 bins=30,
                                 kde=False,
                                 color='skyblue',
                                 hist_kws={"linewidth": 15, 'alpha': 1})
                    host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data_poisson), symbol=''))

                    host1 = pyplot.subplot(121)

                    data_trian = np.random.triangular(minim[i],mode[i], maxim[i], 10000)

                    s = np.random.poisson(m, 10000)

                    data = s * data_trian

                    host1.set_ylabel("Relative Frequency, %")

                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')

                    host1.hist(data, bins=50, alpha=0.6, color='blue')

                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))

                    fmt = '{x:,.0f}'

                    tick = mtick.StrMethodFormatter(fmt)

                    host1.xaxis.set_major_formatter(tick)

                    pyplot.show()

                else:

                    fig = pyplot.figure()

                    pyplot.clf()

                    fig.suptitle(

                        "Risk Impact Duration & Poisson Distribution"+ "\n"+ ' Risk ' + str(i+1) + ': ' + risk_tsk[i],

                        fontsize=12)

                    host = pyplot.subplot(122)
                    host.set_ylabel("Probability Density, %")
                    host.set_xlabel("The number of events")
                    m = event[i]
                    data_poisson = poisson.rvs(mu=m, size=10000)
                    sns.distplot(data_poisson,
                                 bins=30,
                                 kde=False,
                                 color='skyblue',
                                 hist_kws={"linewidth": 15, 'alpha': 1})
                    host.yaxis.set_major_formatter(mtick.PercentFormatter(len(data_poisson), symbol=''))
                    host1 = pyplot.subplot(121)
                    data_trian = np.random.triangular(maxim[i], mode[i], minim[i], 10000)
                    s = np.random.poisson(m, 10000)
                    data = s * data_trian
                    host1.set_ylabel("Relative Frequency, %")
                    host1.set_xlabel("Risk Impact on" + "\n"+'Task '+str(round(impact_task[i]))+ ' Duration')
                    host1.hist(data, bins=50, alpha=0.6, color='blue')
                    host1.yaxis.set_major_formatter(mtick.PercentFormatter(len(data), symbol=''))
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host1.xaxis.set_major_formatter(tick)
                    pyplot.show()
        
        def task_stats_risk1(task_no,even,minim,mode,maxim,event,risk_dis,risk_tsk):
            b1,b2,c1,c2= upload_riskfile(directory,int(entry_simul.get()))

            i = task_no
            if i > len(risk_dis)-1 or i < 0:
                messagebox.showwarning("Hey!", "This risk lies outside of the range of specified.")
            else:    
                mcerp.npts = int(entry_simul.get())
                title = 'Risk ' + str(i+1) + ': ' + risk_tsk[i]
                if risk_dis[i] == '':
                    messagebox.showerror("Hey!",'Risk '+str(i)+': '+risk_tsk[i]+ " has no distribution.")
                elif risk_dis[i] == "Pert":
                    a = minim[i]
                    b = mode[i]
                    c = maxim[i]
                    data_pert=pert_distro(a,b,c,int(entry_simul.get()))
                    if even=='rt':
                        s = b1[i]
                    elif even=='ro':
                        s = b2[i]
                    data=data_pert*s
                    make_table(data, title,' days','r')
                elif risk_dis[i] == "Triangular":   
                    data_tri=tri_distro(minim[i], mode[i][1], maxim[i][2], int(entry_simul.get()))
                    data=data_tri*risk_bin[i]
                    make_table(data, title,' days','r')
                elif risk_dis[i]=="Poisson":
                    data_poi=pois_distro(event[i], int(entry_simul.get()))
                    data_tri=tri_distro(minim[i],mode[i],maxim[i],int(entry_simul.get()))
                    data=data_tri*data_poi
                    make_table(data, title,' days','r')
                    
    
        def clear_val():
            global no_risk_duration,dura_rt,dura_ro,b1,b2,d1,d2,affect_rt,affect_ro
            global crit_index, no_risk_cpm_array, no_risk_ciripath,risk_crit_index, risk_cpm_array, risk_ciripath,no_risk_task,risk_task
            no_risk_duration=[]
            dura_rt=[]
            dura_ro=[]
            b1=[]
            b2=[]
            d1=[]
            d2=[]
            affect_rt=[]
            affect_ro=[]
            crit_index=[]
            no_risk_cpm_array=[]
            no_risk_ciripath=[]
            risk_crit_index=[]
            risk_cpm_array=[]
            risk_ciripath=[]
            no_risk_task=[]
            risk_task=[]
                   
            del no_risk_duration,dura_rt,dura_ro,b1,b2,d1,d2,affect_rt,affect_ro
            del crit_index, no_risk_cpm_array, no_risk_ciripath,risk_crit_index, risk_cpm_array, risk_ciripath,no_risk_task,risk_task
            
            controller.show_frame(ProSchedule)
            entry_simul.delete(0, 'end')
            entry_risk1.delete(0, 'end')
            entry_task.delete(0, 'end')
            entry_risk2.delete(0, 'end')
            
        
        top_frame = tk.Frame(self)
        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")
        center_frame = tk.Frame(self)
        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(side="top", fill='y', expand=True, padx=10, pady=5)
        bottom_frame = tk.Frame(self)
        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)
        label_1 = tk.Label(center_frame, text="Project Schedule with and without Risk Events| Simulation Settings ", font=HEADR_FONT)
        label_1.pack()
        
        message_1 = tk.Message(center_frame, text="Check the distributions of the individual tasks to confirm that "
    
                               "they are correct.\nThen input 'Number of Iterations' and click on 'Run Simulation'.",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        message_2 = tk.Message(center_frame, text="",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_2.pack()
        
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())
        button_1.pack(side="right", padx=5)
        entry_frame = tk.Frame(center_frame)
    
        entry_frame.pack()
        button_2 = ttk.Button(entry_frame, text="Task Distribution",
    
                              command=lambda: check_distro1(int(entry_task.get())-1,mim,task,maxm, distr, name))
    
        button_2.grid(row=0, column=1,columnspan=2)
    
        button_4 = ttk.Button(entry_frame, text="Task Statistics",command=lambda:task_stats1(int(entry_task.get())-1,mim,task,maxm, distr, name))
    
        button_4.grid(row=0, column=3,columnspan=2)
    
        button_5 = ttk.Button(entry_frame, text="Threat Distribution",command=lambda:check_distro_risk1(int(entry_risk1.get())-1,'rt',mim_rt,mod_rt,maxm_rt,event_rt,distr_rt,name_rt,task_rt))
    
        button_5.grid(row=1, column=1,columnspan=2)
    
        button_6 = ttk.Button(entry_frame, text="Threat Statistics",command=lambda:task_stats_risk1(int(entry_risk1.get())-1,'rt',mim_rt,mod_rt,maxm_rt,event_rt,distr_rt,name_rt))
    
        button_6.grid(row=1, column=3,columnspan=2)
        
        button_7 = ttk.Button(entry_frame, text="Opportunity Distribution",command=lambda:check_distro_risk1(int(entry_risk2.get())-1,'ro',mim_ro,mod_ro,maxm_ro,event_ro,distr_ro,name_ro,task_ro))
    
        button_7.grid(row=2, column=1,columnspan=2)
        button_8 = ttk.Button(entry_frame, text="Opportunity Statistics",command=lambda:task_stats_risk1(int(entry_risk2.get())-1,'ro',mim_ro,mod_ro,maxm_ro,event_ro,distr_ro,name_ro))
    
        button_8.grid(row=2, column=3,columnspan=2)
    
    
    
        entry_simul = ttk.Entry(entry_frame, width=8)
        entry_simul.grid(row=3, column=0,columnspan=1)
        entry_simul.insert(0, "100")
        entry_task = ttk.Entry(entry_frame, width=8)
    
        entry_task.grid(row=0, column=0,columnspan=1)
    
        entry_risk1 = ttk.Entry(entry_frame, width=8)
    
        entry_risk1.grid(row=1, column=0,columnspan=1)
        entry_risk2 = ttk.Entry(entry_frame, width=8)
    
        entry_risk2.grid(row=2, column=0,columnspan=1)
    
        entry_task.insert(0,"1")
        entry_risk1.insert(0, "1")
        entry_risk2.insert(0, "1")
        
        
    
        label_sims = tk.Label(entry_frame, text="Number of Iterations")
    
        label_sims.grid(row=3, column=1, columnspan=2)
    
        separate = ttk.Separator(entry_frame, orient="horizontal")
    
        separate.grid(row=5, column=0, columnspan=5, sticky="ew", pady=5)
    
        label_2 = tk.Label(entry_frame, text="Output Display without Risks Events:")
    
        label_2.grid(row=6, column=1, columnspan=2)
    
        button_3 = ttk.Button(entry_frame, text="Run Simulation", command=lambda:get_value(directory))
    
        button_3.grid(row=3, column=3, columnspan=2)
    
        b_pc = ttk.Button(entry_frame, text="PDF & CDF Plots",command=lambda: make_graph2(no_risk_cpm_array, "Project Duration (Without Risks Events), days ", 1, 1))
    
        b_pc.grid(row=7, column=0, sticky="ew")
    
        b_st = ttk.Button(entry_frame, text="Statistics",command=lambda: run_stats1() )
    
        b_st.grid(row=7, column=1, sticky="ew")
        
        b_tornado1 = ttk.Button(entry_frame, text="Tornado Diagram",command=lambda: tornado_withoutrisk(no_risk_ciripath,no_risk_task,name,no_risk_cpm_array))

        b_tornado1.grid(row=7, column=2, sticky="ew")
        
        b_schedule_performace=ttk.Button(entry_frame, text="Schedule Performance", command=lambda: index_graph(no_risk_duration,no_risk_cpm_array,crit_index,int(entry_simul.get()),name))
        b_schedule_performace.grid(row=7, column=3, sticky="ew")
        
        separate2 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate2.grid(row=8, column=0, columnspan=5, sticky="ew", pady=5)
    
        label_3 = tk.Label(entry_frame, text="Output Display with Risks Events:")
    
        label_3.grid(row=9, column=1, columnspan=2,sticky="ew")
        
        b_pc2 = ttk.Button(entry_frame, text="PDF & CDF Plots",command=lambda: make_graph2(risk_cpm_array, " Project Duration (With Risks Events), days ", 1, 1))
    
        b_pc2.grid(row=10, column=0,sticky="ew")
    
        b_st2 = ttk.Button(entry_frame, text="Statistics",command=lambda: run_stats2())
    
        b_st2.grid(row=10, column=1,sticky="ew")
        
        b_tornado = ttk.Button(entry_frame, text="Tornado Diagram",command=lambda: tornado_risk(risk_ciripath,risk_task,name,risk_cpm_array))

        b_tornado.grid(row=10, column=2,sticky="ew")
        b_schedule_performace2=ttk.Button(entry_frame, text="Schedule Performance", command=lambda: index_graph(d2,risk_cpm_array,risk_crit_index,int(entry_simul.get()),name))
        b_schedule_performace2.grid(row=10, column=3,sticky="ew")
       
        separate4 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate4.grid(row=11, column=0, columnspan=5, sticky="ew", pady=5)
    
        label_4 = tk.Label(entry_frame, text="Output Display Comparing with VS without Risks Events:")
    
        label_4.grid(row=12, column=1, columnspan=2,sticky="ew")
        
        b_pcw = ttk.Button(entry_frame, text="PDF Plots",command=lambda: make_risk_graph2(no_risk_cpm_array,risk_cpm_array, "Project Duration, days ", 0, 1))
    
        b_pcw.grid(row=13, column=1, sticky="ew")
    
        b_pcn = ttk.Button(entry_frame, text="CDF Plots",command=lambda: make_risk_graph2(no_risk_cpm_array,risk_cpm_array, "Project Duration, days ", 1, 0))
    
        b_pcn.grid(row=13, column=2, sticky="ew")
        
        
        separate3 = ttk.Separator(entry_frame, orient="horizontal")
    
        separate3.grid(row=14, column=0, columnspan=5, sticky="ew", pady=5)
        
        label_5 = tk.Label(entry_frame, text="                                        ")
    
        label_5.grid(row=15, column=1, columnspan=2)
    
    
    
    
        b_write_excel_risk = ttk.Button(entry_frame, text="Output Data", command=lambda: savetoflie())
    
        b_write_excel_risk.grid(row=16, column=1, columnspan=2, sticky="ew")

       


#################################################################################################################################
class Optimization(tk.Frame):
    """
    This page allows the user to select the excel file that contains the Optimization they would like to import.
    """
    def __init__(self, parent, controller):
        """
        Class constructor.
        :param parent: The parent frame.
        :param controller: The controller runs through every page and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)
        global directory
        directory =""
    

        def get_string():

            global directory

            directory = entry_1.get()

            if directory == "":

                messagebox.askretrycancel("Error", "An error has occurred in your File.")

            else:

                controller.show_frame(PortOOptions)
        
        


        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Portfolio Optimization | File Upload", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the MS Excel file with your portfolio data: (.xlsx)\n"

                               "Ensure your data has the following fields: project ID, investment, distribution, "

                               "minimum, most likely, maximum, standard deviation, and your constraint values",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()

        entry_1 = ttk.Entry(center_frame, width=50)

        entry_1.pack(side="left")

        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))

        button_2.pack(side="left", pady=5)

        button_3 = ttk.Button(center_frame, text="Open", command=lambda: get_string())

        button_3.pack(side="left", padx=5, pady=5)

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: controller.show_frame(ChoosePage))

        button_1.pack(side="right", padx=5)

####################################################################################################################################
        
class Optimization_Simulation(tk.Frame):
    """
    This page allows the user to select the excel file that contains the Optimization they would like to import.
    """
    def __init__(self, parent, controller):
        
        """
        Class constructor.
        :param parent: The parent frame.
        :param controller: The controller runs through every page and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)
        def write_to_excel(tsk, data):
            messagebox.showwarning("Hey","1. The iteration data are stored in new Excel files titled 'Optimization Output Data' "
                                   " inside the source folder. \n \n2. "
                                   "Before you run the next simulation, please make sure Excel files are closed,"
                                   " so that the new iteration data can override the old iteration data.")

   
            wb = xlsxwriter.Workbook('Optimization Output Data.xlsx')
            ws = wb.add_worksheet()
            A1 = 'Output Data'
            title = np.append(A1, tsk)
        
            item = title
            column = 0
            sim_task=[]
            
            for i in range(0,len(tsk)):
                 if distr[i] == 'Pert':
                        
                    sim_task.append(pert_distro(mim[i], task[i], maxm[i], 10000))
                 elif distr[i] == 'Triangular':
                    sim_task.append(tri_distro(mim[i], task[i], maxm[i], 10000))
                 else:
                     
                    sim_task.append(no_distro(task[i],10000))
    
            for i in item:
                ws.write(0, column, i)
                column += 1
            row = 1
        
            for i in data:
                ws.write(row, 0, i)
                row += 1
        
            col = 1
            
            for i in sim_task:
                
                r = 1
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    r += 1
                col += 1
        
            wb.close()

        def simul_data(distr,mim,task,maxm,itera):
            simul=[]
        
            if distr == 'Pert':
                
                if mim=='' or task=='' or maxm=='':
                    messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                    return
                else:
                    simul.append(pert_distro(mim, task, maxm, itera))
    
            elif distr == 'Triangular':
                if mim=='' or task=='' or maxm=='':
                    messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                    return
                else:
                    simul.append(tri_distro(mim, task, maxm, itera))
    
            else:
                if task=='':
                    messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                    return
                else:
                    simul.append(no_distro(task,itera))
            return simul
        
        def get_string():
            from collections import Counter
            if cons2==0 or cons2=='':
                y=plotdata
                x=np.linspace(0,len(y),len(y),False).tolist()
                
                z=Counter(y)
                
                m=list(z.keys())# value
                n=list(z.values())#number
                y_new=[]
                for i in range(0,len(m)-1):
                    y_new+=np.linspace(m[i],m[i+1],n[i],False).tolist()
                
                y_new=y_new+y[-n[-1]:]
                
                pyplot.plot(x,y_new)
                pyplot.xlabel('Investment, $')
                pyplot.ylabel('NPV, $')
                
                pyplot.title('Simulated Efficient Frontier')
                pyplot.show()


                show_table_one(result_cand, project_cand,inv_m,npv_m,inv,'Results')
            elif cons2!=0:
                y=plotdata
                x=np.linspace(0,len(y),len(y),False).tolist()
                
                z=Counter(y)
                
                m=list(z.keys())# value
                n=list(z.values())#number
                y_new=[]
                for i in range(0,len(m)-1):
                    y_new+=np.linspace(m[i],m[i+1],n[i],False).tolist()
                
                y_new=y_new+y[-n[-1]:]
                
                pyplot.plot(x,y_new)
                pyplot.xlabel('Investment, $')
                pyplot.ylabel('NPV, $')
                
                pyplot.title('Simulated Efficient Frontier')
                pyplot.show()

                show_table_two(result_cand, project_cand,inv_m,npv_m,res_m,inv,res,'Results')
        
        def get_value():
            global simu
            simu=[]
            
            for j in range(0,len(project_cand)):
                z=name.index(project_cand[j])
                simu+=simul_data(distr[z],mim[z],task[z],maxm[z],int(entry_simul.get()))
                
             
            simu=np.sum(simu, axis=0)
            messagebox.showwarning("Hey","Simulation is Done!")
        def clear_val():
            global simu
            del simu
            controller.show_frame(PortOOptions)
        
        def show_table_two(data1, data2,data3,data4,data5,inv,res,string_1):
            
            window = tk.Toplevel()
            window.resizable(width=True, height=True)
            window.title("Table")
            main = tk.Label(window, text=string_1, font=('Calibri', 24, 'bold') )
            main.pack()
           
            style = ttk.Style()
            style.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=40)  # Modify the font of the body
            style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
            
            
            tree = ttk.Treeview(window, selectmode="extended", show='tree',height="30", style="mystyle.Treeview")
            tree["columns"] = ("", "", "", "","")
            tree.column("#0", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#1",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#2", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#3", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#4", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            

            tree.tag_configure("evenrow",background='white',foreground='black')
            tree.tag_configure("oddrow",background='black',foreground='white')
            
            tree.insert('', 0, values=['Projects Selected', 'Investment', 'NPV','Number of Resources'],tags = ('oddrow',))
            m=len(data2)
            for i in range(len(data2)):
                
                tree.insert('', i+1, values=[str(data2[i]), '$'+str(data3[i]),'$'+str(data4[i]),str(data5[i])],tags = ('evenrow',))
                                           
            tree.insert('', m+1, values=['Total', '$'+str(int(inv)), '$'+str(int(data1)),str(int(res))],tags = ('oddrow',))
            
            tree.pack() 
            
            
            window1 = tk.Toplevel()
            window1.resizable(width=True, height=True)
            window1.title("Table")
            main1 = tk.Label(window1, text='Constraints', font=('Calibri', 24, 'bold') )
            main1.pack()
            style1 = ttk.Style()
            style1.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=60)  # Modify the font of the body
            style1.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
           
            
            tree2 = ttk.Treeview(window1, selectmode="extended", show='tree',height="100", style="mystyle.Treeview")
            tree2["columns"] = ("stat", "sdata", "perc", "pdata","mdata")
            
            tree2.column("stat",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("sdata", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("perc", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            tree2.column("pdata", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("mdata", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            
            tree2.tag_configure("evenrow",background='white',foreground='black')
            tree2.tag_configure("oddrow",background='black',foreground='white')
            
            tree2.insert('', 0, values=['Investment', 'Number of',' Resources'],tags = ('oddrow',))
            tree2.insert('', 1, values=['$'+conss1, conss2],tags = ('evenrow',))
            tree2.insert('', 3, values=['Fixed', 'Either','Or','If','Then'],tags = ('oddrow',))
            tree2.insert('', 4, values=[fix_11, either_11,or_11,if_11,then_11],tags = ('evenrow',))
            tree2.insert('', 5, values=[fix_22, either_22,or_22,if_22,then_22],tags = ('evenrow',))
            tree2.insert('', 6, values=[fix_33, either_33,or_33,if_33,then_33],tags = ('evenrow',))
            tree2.insert('', 7, values=[fix_44, either_44,or_44,if_44,then_44],tags = ('evenrow',))
            tree2.insert('', 8, values=[fix_55, either_55,or_55,if_55,then_55],tags = ('evenrow',))
            
            tree2.pack() 
            
        def show_table_one(data1, data2,data3,data4,inv,string_1):
            """ Displays the statistics of the generated data in a table format.
            :param data: The data.
            :param string_1: The title of the table.
            """
            window = tk.Toplevel()
            window.resizable(width=True, height=True)
            window.title("Table")
            main = tk.Label(window, text=string_1, font=('Calibri', 24, 'bold') )
            main.pack()
           
            style = ttk.Style()
            style.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=40)  # Modify the font of the body
            style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
            
            
            tree = ttk.Treeview(window, selectmode="extended", show='tree',height="30", style="mystyle.Treeview")
            tree["columns"] = ("", "", "", "","")
            tree.column("#0", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#1",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#2", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#3", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#4", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            
           
           
            
            tree.tag_configure("evenrow",background='white',foreground='black')
            tree.tag_configure("oddrow",background='black',foreground='white')
            
            tree.insert('', 0, values=['Projects Selected', 'Investment', 'NPV','Number of Resources'],tags = ('oddrow',))
            m=len(data2)
            for i in range(len(data2)):
                
                tree.insert('', i+1, values=[str(data2[i]), '$'+str(data3[i]),'$'+str(data4[i]),''],tags = ('evenrow',))
                                           
            tree.insert('', m+1, values=['Total', '$'+str(int(inv)), '$'+str(int(data1)),''],tags = ('oddrow',))
            
            tree.pack() 
            
            
            window1 = tk.Toplevel()
            window1.resizable(width=True, height=True)
            window1.title("Table")
            main1 = tk.Label(window1, text='Constraints', font=('Calibri', 24, 'bold') )
            main1.pack()
            style1 = ttk.Style()
            style1.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=60)  # Modify the font of the body
            style1.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
           
            
            tree2 = ttk.Treeview(window1, selectmode="extended", show='tree',height="100", style="mystyle.Treeview")
            tree2["columns"] = ("stat", "sdata", "perc", "pdata","mdata","")
            
            tree2.column("stat",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("sdata", stretch=tk.YES, minwidth=100, width=150,anchor="center")
            tree2.column("perc", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            tree2.column("pdata", stretch=tk.YES, minwidth=50, width=100,anchor="center")
            tree2.column("mdata", stretch=tk.YES, minwidth=50, width=60,anchor="center")
            tree2.column("", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            
            
            tree2.tag_configure("evenrow",background='white',foreground='black')
            tree2.tag_configure("oddrow",background='black',foreground='white')
            
            tree2.insert('', 0, values=['Investment', 'Number of \n Resources'],tags = ('oddrow',))
            tree2.insert('', 1, values=['$'+conss1, ''],tags = ('evenrow',))
            tree2.insert('', 3, values=['Fixed', 'Either','Or','If','Then'],tags = ('oddrow',))
            tree2.insert('', 4, values=[fix_11, either_11,or_11,if_11,then_11],tags = ('evenrow',))
            tree2.insert('', 5, values=[fix_22, either_22,or_22,if_22,then_22],tags = ('evenrow',))
            tree2.insert('', 6, values=[fix_33, either_33,or_33,if_33,then_33],tags = ('evenrow',))
            tree2.insert('', 7, values=[fix_44, either_44,or_44,if_44,then_44],tags = ('evenrow',))
            tree2.insert('', 8, values=[fix_55, either_55,or_55,if_55,then_55],tags = ('evenrow',))
            
            tree2.pack() 
        
        def check_distro1(task_no,minim,mode,maxim, distro, task):
            
            i = task_no
            if i > len(distro)-1 or i < 0:
    
                messagebox.showwarning("Hey!", "This task lies outside of the range of specified.")
            elif distro[i]=='':
                messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + " has no distribution.")
            elif distro[i] == "Normal":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    mu = minim[i]
                    sigma = mode[i]
                    x = np.linspace(mu - 5 * sigma, mu + 5 * sigma, 100)
                    pyplot.plot(x, ss.norm.pdf(x, mu, sigma))
                    pyplot.fill(x, ss.norm.pdf(x, mu, sigma))
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("NPV, $")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('NPV' + str(i+1) + ': ' + task[i]+"\n"+"Normal Distribution")
                    pyplot.show()
            
            elif distro[i] == "Pert":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                  
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    a = minim[i]
                    b = mode[i]
                    c = maxim[i]
                    
                    x = pert_distro(a,b, c, 10000)
                    sns.kdeplot(x, data2=None, shade=True, vertical=False, kernel='gau', bw='scott',
                                    gridsize=100, cut=3, clip=None, legend=True, cumulative=False, shade_lowest=True, cbar=False, cbar_ax=None, cbar_kws=None, ax=None,color="b")
                    
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("NPV, $")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('NPV' + str(i+1) + ': ' + task[i]+"\n"+"Pert Distribution")
                    
                    pyplot.show()
                    

    
            elif distro[i] == "Triangular":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("NPV, $")
                    c = (mode[i] - minim[i]) / (maxim[i]-minim[i])
                    mean, var, skew, kurt = triang.stats(c, moments='mvsk')
                    x = np.linspace(minim[i], maxim[i], 1000)
                    host.plot(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]), 'b', lw=2)
                    pyplot.fill(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]))
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title( 'NPV'+str(i+1)+': '+task[i]+"\n"+"Triangular Distribution")
                    pyplot.show()
                
            elif distro[i]=="Uniform":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel("NPV, $")
                    a=minim[i]
                    b=mode[i]
                    mean, var, skew, kurt = ss.uniform.stats(moments='mvsk')
                    x = np.linspace(a, b, 100)
                    host.plot(x, ss.uniform.pdf(x, loc=a, scale=b-1), 'b', lw=2, alpha=0.6, label='uniform pdf')
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title('NPV' + str(i+1) + ': ' + task[i]+"\n"+"Uniform Distribution")
                    pyplot.show()
    
    
       
        def task_stats1(task_no,minim,mode,maxim, distro, task):

            i = task_no
            
            if i > len(distro)-1 or i < 0:

                messagebox.showwarning("Hey!", "This NPVlies outside of the range of specified.")
            else:
                
               
                mcerp.npts = int(entry_simul.get())
    
                title = 'NPV' + str(i+1) + ': ' + task[i]
    
                if distro[i]=='':
                    messagebox.showerror("Hey!", 'NPV' + str(i + 1) + ': ' + task[i] + " has no distribution.")
    
                elif distro[i] == "Normal":
    
                    data = np.random.normal(minim[i], mode[i], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Pert":
    
                    a = minim[i]
    
                    b = mode[i]
    
                    c = maxim[i]
    
                    alpha = ((4 * b) + c - (5 * a)) / (c - a)
    
                    beta = ((5 * c) - a - (4 * b)) / (c - a)
    
                    x = np.random.beta(alpha, beta, int(entry_simul.get()))
    
                    data = a + (x * (c - a))
    
                    make_table(data, title)
    
                elif distro[i] == "Triangular":
    
                    data = np.random.triangular(minim[i], mode[i], maxim[i], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Uniform":
    
                    data = np.random.uniform(minim[i],mode[i], int(entry_simul.get()))
    
                    make_table(data, title)
                    

            
            

           
        


        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Input data for Stochastic Optimization | File Upload", font=HEADR_FONT)

        label_1.pack()

        
        message_1 = tk.Message(center_frame, text="Check the distributions of the individual project to confirm that "
    
                               "they are correct.\nThen input 'Number of Iterations' and click on 'Run Simulation'.",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        message_2 = tk.Message(center_frame, text="",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_2.pack()
        
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: controller.show_frame(PortOOptions))
        button_1.pack(side="right", padx=5)
        entry_frame = tk.Frame(center_frame)
    
        entry_frame.pack()
        button_2 = ttk.Button(entry_frame, text="NPV Distribution",
    
                              command=lambda: check_distro1(int(entry_task.get())-1,mim,task,maxm, distr, name))
    
        button_2.grid(row=0, column=1,columnspan=2)
    
        button_4 = ttk.Button(entry_frame, text="NPV Statistics",command=lambda:task_stats1(int(entry_task.get())-1,mim,task,maxm, distr, name))
    
        button_4.grid(row=0, column=3,columnspan=2)
    
        entry_simul = ttk.Entry(entry_frame, width=8)
        entry_simul.grid(row=3, column=0,columnspan=1)
        entry_simul.insert(0, "100")
        entry_task = ttk.Entry(entry_frame, width=8)
    
        entry_task.grid(row=0, column=0,columnspan=1)
    
        
    
        entry_task.insert(0,"1")
       
    
        label_sims = tk.Label(entry_frame, text="Number of Iterations")
    
        label_sims.grid(row=3, column=1, columnspan=2)
    
        separate = ttk.Separator(entry_frame, orient="horizontal")
    
        separate.grid(row=5, column=0, columnspan=5, sticky="ew", pady=5)
    
        label_2 = tk.Label(entry_frame, text="Output Display:")
    
        label_2.grid(row=6, column=1, columnspan=2)
    
        button_3 = ttk.Button(entry_frame, text="Run Simulation", command=lambda:get_value())
    
        button_3.grid(row=3, column=3, columnspan=2)
    
        b_pc = ttk.Button(entry_frame, text="PDF & CDF Plots",command=lambda: make_graph(simu, "NPV, $ ", 1, 1))
    
        b_pc.grid(row=7, column=1, sticky="ew")
    
        b_st = ttk.Button(entry_frame, text="Statistics",command=lambda: make_table(simu, "NPV Statistics") )
    
        b_st.grid(row=7, column=2, sticky="ew")
        
        b_tornado1 = ttk.Button(entry_frame, text="Results",command=lambda: get_string() )

        b_tornado1.grid(row=7, column=0, sticky="ew")
        
        
    
    
        b_write_excel_risk = ttk.Button(entry_frame, text="Output Data", command=lambda: write_to_excel(name, simu))
    
        b_write_excel_risk.grid(row=8, column=1, columnspan=2, sticky="ew")

       


###################################################################################################################################
class PortOOptions(tk.Frame):

    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)
        
        numb_1=''
        numb_2=''
        numb_3=''

        var_cons2 = tk.IntVar()
        var_fix = tk.IntVar()
        var_either = tk.IntVar()
        var_if = tk.IntVar()
    
        def read_Port_excel():
        
            book=xlrd.open_workbook(directory)
            sheet=book.sheet_by_index(0)
            t=int(sheet.cell_value(0,0))
            name=[]
            for i in range(0,t):
                name.append(sheet.cell_value(i+1,1)) #task name
            task=[]
            for i in range(0,t):
                task.append(round(sheet.cell_value(i+1,5)))  #task duration
            distr=[]
            for i in range(0,t):
                distr.append(sheet.cell_value(i+1,7))  #task distribution
            mim=[]
            for i in range(0,t):
                mim.append(round(sheet.cell_value(i+1,4))) #min
            maxm=[]
            for i in range(0,t):
                maxm.append(round(sheet.cell_value(i+1,6)))
            invest=[]
            for i in range(0,t):
                invest.append(round(sheet.cell_value(i+1,2)))
            resource=[]
            for i in range(0,t):
                resource.append(round(sheet.cell_value(i+1,3)))
            
            
            sim_task=[]
            
            for i in range(0,t):
                 if distr[i] == 'Pert':
                        
                    if mim[i]=='' or task[i]=='' or maxm[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                    else:
                        sim_task.append(round(np.average(pert_distro(mim[i], task[i], maxm[i], 10000))))
                 elif distr[i] == 'Triangular':
                    if mim[i]=='' or task[i]=='' or maxm[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                    else:
                        sim_task.append(round(np.average(tri_distro(mim[i], task[i], maxm[i], 10000)),2))
                 else:
                     
                     if task[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                     else:
                        sim_task.append(round(np.average(no_distro(task[i],10000)),2))
            
        
            return name,task, t, mim, maxm,invest,resource,sim_task,distr
       
        
        def split(word):  return list(word)
   
        def combine_input():
            if field_11.get()=='':
                messagebox.showwarning("Alert!", "Investment should not be empty.") 
                return
            if var_cons2.get() == 1:
                cons2=int(field_21.get())
                if cons2=='' :
                    messagebox.showwarning("Alert!", "Resources should not be empty.") 
                    return
            elif var_cons2.get() == 0:
                cons2=0
            
            if var_fix.get()  == 1:
                if fix_1.get()!='':
                    fix1=str(fix_1.get())
                elif str(fix_1.get())=='':
                    messagebox.showwarning("Alert!", "Fixed should not be empty.") 
                    return
                if str(fix_2.get())!='':
                    fix2=str(fix_2.get())
                elif str(fix_2.get())=='':
                    fix2=''
                if str(fix_3.get())!='':
                    fix3=str(fix_3.get())
                elif str(fix_3.get())=='':
                    fix3=''
                if str(fix_4.get())!='':
                    fix4=str(fix_4.get())
                elif str(fix_4.get())=='':
                    fix4=''
                if str(fix_5.get())!='':
                    fix5=str(fix_5.get())
                elif str(fix_5.get())=='':
                    fix5=''
                fix_ini=fix1+fix2+fix3+fix4+fix5
                fix_ini=split(fix_ini)
            
            elif var_fix.get()  == 0:
                fix_ini=''
                
            if var_either.get() == 1:
                if str(either_1.get())!='':
                    either1=str(either_1.get())
                elif str(either_1.get())=='':
                    messagebox.showwarning("Alert!", "Either should not be empty.") 
                    return
                if str(either_2.get())!='':
                    either2=str(either_2.get())
                elif str(either_2.get())=='':
                    either2=''
                if str(either_3.get())!='':
                    either3=str(either_3.get())
                elif str(either_3.get())=='':
                    either3=''
                if str(either_4.get())!='':
                    either4=str(either_4.get())
                elif str(either_4.get())=='':
                    either4=''
                if str(either_5.get())!='':
                    either5=str(either_5.get())
                elif str(either_5.get())=='':
                    either5=''
                either_ini=either1+either2+either3+either4+either5
                either_ini=split(either_ini)
                
                
                if str(or_1.get())!='':
                    or1=str(or_1.get())
                elif str(or_1.get())=='':
                    messagebox.showwarning("Alert!", "Or should not be empty.") 
                    return
                if str(or_2.get())!='':
                    or2=str(or_2.get())
                elif str(or_2.get())=='':
                    or2=''
                if str(or_3.get())!='':
                    or3=str(or_3.get())
                elif str(or_3.get())=='':
                    or3=''
                if str(or_4.get())!='':
                    or4=str(or_4.get())
                elif str(or_4.get())=='':
                    or4=''
                if str(or_5.get())!='':
                    or5=str(or_5.get())
                elif str(or_5.get())=='':
                    or5=''
                or_ini=or1+or2+or3+or4+or5
                or_ini=split(or_ini)
                
                if len(or_ini)!=len(either_ini):
                     messagebox.showwarning("Alert!", "Input of Either Or should be same length.") 
                     return
            elif var_either.get()  == 0:
                either_ini=''
                or_ini=''
                
            if var_if.get() == 1:
                
                if str(if_1.get())!='':
                    if1=str(if_1.get())
                elif str(if_1.get())=='':
                    messagebox.showwarning("Alert", "If should not be empty.") 
                    return
                if str(if_2.get())!='':
                    if2=str(if_2.get())
                elif str(if_2.get())=='':
                    if2=''
                if str(if_3.get())!='':
                    if3=str(if_3.get())
                elif str(if_3.get())=='':
                    if3=''
                if str(if_4.get())!='':
                    if4=str(if_4.get())
                elif str(if_4.get())=='':
                    if4=''
                if str(if_5.get())!='':
                    if5=str(if_5.get())
                elif str(if_5.get())=='':
                    if5=''
                if_ini=if1+if2+if3+if4+if5
                if_ini=split(if_ini)
                
                if str(then_1.get())!='':
                    then1=str(then_1.get())
                elif str(then_1.get())=='':
                    messagebox.showwarning("Alert", "Then should not be empty.") 
                    return
                if str(then_2.get())!='':
                    then2=str(then_2.get())
                elif str(then_2.get())=='':
                    then2=''
                if str(then_3.get())!='':
                    then3=str(then_3.get())
                elif str(then_3.get())=='':
                    then3=''
                if str(then_4.get())!='':
                    then4=str(then_4.get())
                elif str(then_4.get())=='':
                    then4=''
                if str(then_5.get())!='':
                    then5=str(then_5.get())
                elif str(then_5.get())=='':
                    then5=''
                then_ini=then1+then2+then3+then4+then5
                then_ini=split(then_ini)
                
                if len(if_ini)!=len(then_ini):
                     messagebox.showwarning("Alert", "Input of If Then should be same length.") 
                     return
            
            elif var_if.get()  == 0:
                if_ini=''
                then_ini=''
            
            if fix_ini !='':
                if if_ini !='':
                    for i in range(len(if_ini)):
                        if then_ini[i] in fix_ini: 
                            if_ini[i]=''
                            then_ini[i]=''
                        else:
                            if if_ini[i] in fix_ini:
                                fix_ini.append(then_ini[i])
                                if_ini[i]=''
                                then_ini[i]=''
                if either_ini!='':
                    for i in range(len(either_ini)):
                        
                        if (either_ini[i] in fix_ini) and (or_ini[i] in fix_ini):
                            messagebox.showwarning("Alert!", "There is conflicting input information.") 
                            return 
                        elif (either_ini[i] in fix_ini) or (or_ini[i] in fix_ini):
                            either_ini[i]=''
                            or_ini[i]=''
            if fix_ini!='':
                
                while '' in fix_ini: fix_ini.remove('')
            if if_ini!='':
                
                while '' in if_ini:if_ini.remove('')
            if then_ini!='':
                
                while '' in then_ini:then_ini.remove('')
            if either_ini!='':
                
                while '' in either_ini: either_ini.remove('')
            if or_ini!='':
                
                while '' in or_ini: or_ini.remove('')
            if set(either_ini) == set(or_ini) and either_ini !='':
                 messagebox.showwarning("Alert!", "There is conflicting input information.") 
                 return 
            if numb_1.get()!='':
                if int(numb_1.get())!=len(fix_ini):
                     messagebox.showwarning("Alert!", "There is conflicting input information.") 
                     return 
            if numb_2.get()!='':
                
                if int(numb_2.get())!=len(either_ini):
                     messagebox.showwarning("Alert!", "There is conflicting input information.") 
                     return 
            if numb_3.get()!='':
                if int(numb_3.get())!=len(if_ini):
                     messagebox.showwarning("Alert!", "There is conflicting input information.") 
                     return    
            return cons2, fix_ini, either_ini, or_ini, if_ini, then_ini
        
        def fixed_two(number,project,profit,c1,c2,compulsory,invest,resource,task,name):
            name1=name.copy()
            invest1=invest.copy()
            resource1=resource.copy()
            task1=task.copy()
            for j in range(0,int(number)):
                i=name.index(compulsory[j]) #find position of  
                profit+=task[i]
                project.append(name[i])
                 
                c1-=int(invest[i])
                c2-=int(resource[i])
                
                name1.remove(name[i])
                invest1.remove(invest[i])
                resource1.remove(resource[i])
                task1.remove(task[i])
            if c1<0 or c2<0:
                messagebox.showwarning("Alert!", "There is no feasible solution") 
                return
            return project,profit,c1,c2,invest1,resource1,task1,name1 
        
        def fixed_one(number,project,profit,c1,compulsory,invest,task,name):
            name1=name.copy()
            invest1=invest.copy()
            task1=task.copy()
            for j in range(0,int(number)):
                i=name.index(compulsory[j]) #find position of  
                profit+=task[i]
                project.append(name[i])
                 
                c1-=int(invest[i])
                
                name1.remove(name[i])
                invest1.remove(invest[i])
                task1.remove(task[i])
            if c1<0:
                messagebox.showwarning("Alert!", "There is no feasible solution") 
                return
            return project,profit,c1,invest1,task1,name1 
        
        def constrain_select(name1,project,profit,c1,c2,invest1,resource1,N_1):
            if len(c1)>1:
                for i in c1:
                    if i<0:
                        z=c1.index(i)
                        c1[z]=-1
                        c2[z]=-1
                        name1[z]=''
                        project[z]=''
                        profit[z]=''
                        invest1[z]=''
                        resource1[z]=''
                        N_1[z]=''
                        
                for i in c2:
                    if i<0:
                        z=c2.index(i)
                        c1[z]=-1
                        c2[z]=-1
                        name1[z]=''
                        project[z]=''
                        profit[z]=''
                        invest1[z]=''
                        resource1[z]=''
                        N_1[z]=''
                        
                while -1 in c1: c1.remove(-1)
                while -1 in c2:c2.remove(-1)
                while '' in name1:name1.remove('')
                while '' in project: project.remove('')
                while '' in profit:profit.remove('')
                while '' in invest1: invest1.remove('')
                while '' in resource1:resource1.remove('')
                while '' in N_1: N_1.remove('')    
            return name1,project,profit,c1,c2,invest1,resource1,N_1
        def constrain_select_one(name1,project,profit,c1,invest1,N_1):
            if len(c1)>1:
                for i in c1:
                    if i<0:
                        z=c1.index(i)
                        c1[z]=-1
                        name1[z]=''
                        project[z]=''
                        profit[z]=''
                        invest1[z]=''
                        N_1[z]=''
                while -1 in c1: c1.remove(-1)
                while '' in name1:name1.remove('')
                while '' in project: project.remove('')
                while '' in profit:profit.remove('')
                while '' in invest1: invest1.remove('')
                while '' in N_1: N_1.remove('')    
                    
            return name1,project,profit,c1,invest1,N_1
        
        def if_then_two(number2,project,profit,c1,c2,if1,then,invest,resource,task,name,name2):
        
            if number2!=0:
                c_1=[]
                c_2=[]
                for j in range(0,int(number2)):
                    c_1.append(name.index(if1[j]))
                    c_2.append(name.index(then[j]))
                
                name1=[copy.deepcopy(name2)]
                project=[project]
                profit=[profit]
                c1=[c1]
                c2=[c2]
                invest1=[]
                resource1=[]
                N_1=[]
                
                for j in range(len(if1)):
                    cand_1=copy.deepcopy(name1)
                    cand_2=copy.deepcopy(name1)
                    pro_1=copy.deepcopy(project)
                    pro_2=copy.deepcopy(project)
                    fit_1=profit.copy()
                    fit_2=profit.copy()
                    cons1=c1.copy()
                    cons2=c1.copy()
                    cons3=c2.copy()
                    cons4=c2.copy()
                    
                    for l in range(0,2**j):
                        cand_1[l].remove(if1[j])
                        cand_2[l].remove(if1[j])
                        cand_2[l].remove(then[j])
                        pro_2[l]+=[if1[j]]
                        pro_2[l]+=[then[j]]
                        fit_2[l]+=task[c_2[j]]
                        fit_2[l]+=task[c_1[j]]
                        cons2[l]-=invest[c_2[j]]
                        cons2[l]-=invest[c_1[j]]
                        
                        cons4[l]-=resource[c_2[j]]
                        cons4[l]-=resource[c_1[j]]
                    name1=cand_1.copy()+cand_2.copy()
                    project=pro_1.copy()+pro_2.copy()
                    profit=fit_1.copy()+fit_2.copy()
                    c1=cons1.copy()+cons2.copy()
                    c2=cons3.copy()+cons4.copy()
                
                    
                for j in range(len(name1)):
                 
                    res=[]
                    inv=[]
                    npv1=[]
                    for k in name1[j]:
                        posi=name.index(k)
                        inv.append(invest[posi])
                        res.append(resource[posi])
                        npv1.append(task[posi])
                    invest1.append(inv)
                    resource1.append(res)
                    N_1.append(npv1)
            name1,project,profit,c1,c2,invest1,resource1,N_1= constrain_select(name1,project,profit,c1,c2,invest1,resource1,N_1)  
            
            return name1,project,profit, c1,c2,invest1,resource1 ,N_1
        def if_then_one(number2,project,profit,c1,if1,then,invest,task,name,name2):
        
            if number2!=0:
                c_1=[]
                c_2=[]
                for j in range(0,int(number2)):
                    c_1.append(name.index(if1[j]))
                    c_2.append(name.index(then[j]))
                
                name1=[copy.deepcopy(name2)]
                project=[project]
                profit=[profit]
                c1=[c1]
               
                invest1=[]
                
                N_1=[]
                
                for j in range(len(if1)):
                    cand_1=copy.deepcopy(name1)
                    cand_2=copy.deepcopy(name1)
                    pro_1=copy.deepcopy(project)
                    pro_2=copy.deepcopy(project)
                    fit_1=profit.copy()
                    fit_2=profit.copy()
                    cons1=c1.copy()
                    cons2=c1.copy()
                    
                    
                    for l in range(0,2**j):
                        cand_1[l].remove(if1[j])
                        cand_2[l].remove(if1[j])
                        cand_2[l].remove(then[j])
                        pro_2[l]+=[if1[j]]
                        pro_2[l]+=[then[j]]
                        fit_2[l]+=task[c_2[j]]
                        fit_2[l]+=task[c_1[j]]
                        cons2[l]-=invest[c_2[j]]
                        cons2[l]-=invest[c_1[j]]
                       
                    name1=cand_1.copy()+cand_2.copy()
                    project=pro_1.copy()+pro_2.copy()
                    profit=fit_1.copy()+fit_2.copy()
                    c1=cons1.copy()+cons2.copy()
                
                    
                for j in range(len(name1)):
                 
                    inv=[]
                    npv1=[]
                    for k in name1[j]:
                        posi=name.index(k)
                        inv.append(invest[posi])
                        npv1.append(task[posi])
                    invest1.append(inv)
                    N_1.append(npv1)
            name1,project,profit,c1,invest1,N_1= constrain_select_one(name1,project,profit,c1,invest1,N_1)
            
            return name1,project,profit, c1,invest1, N_1        
        def keep_remove_one(project,profit,u1,d1,name1,name,invest,cons1):
            
            
            pp=[]
            nn=[]
            
            for k in range(len(name1)):
                keep_p=[copy.deepcopy(project[k])]
                keep_n=[copy.deepcopy(name1[k])]
                
                for i in range(len(u1)):
                    uu_p=copy.deepcopy(keep_p)
                    dd_p=copy.deepcopy(keep_p)
                    uu_n=copy.deepcopy(keep_n)
                    dd_n=copy.deepcopy(keep_n)
                    for j in range(0,2**i):
                        
                        if (u1[i] in keep_p[j]) :
                            if (d1[i] in keep_n[j]):
                                uu_n[j].remove(d1[i])
                                dd_p[j].remove(u1[i])
                                dd_n[j].remove(d1[i])
                        if (d1[i] in keep_p[j]):
                            if (u1[i] in keep_n[j]) :
                                dd_n[j].remove(u1[i])
                                uu_p[j].remove(d1[i])
                                uu_n[j].remove(u1[i])
                        
                                
                        if (u1[i] not in keep_p[j]) and (d1[i] not in keep_p[j]):
                           
                                
                              
                            if(u1[i] in keep_n[j]) and (d1[i] in keep_n[j]):
                                uu_p[j].append(u1[i])
                                uu_n[j].remove(d1[i])
                                uu_n[j].remove(u1[i])
                                dd_p[j].append(d1[i])
                                dd_n[j].remove(d1[i])
                                dd_n[j].remove(u1[i])
                               
                            elif (u1[i] in keep_n[j]) and (d1[i] not in keep_n[j]):
                                uu_p[j].append(u1[i])
                                uu_n[j].remove(u1[i])
                                
                                dd_n[j].remove(u1[i])
                                
                            elif  (d1[i] in keep_n[j] )and (u1[i] not in keep_n[j]):
                                
                                dd_p[j].append(d1[i])
                                
                                dd_n[j].remove(d1[i])
                                
                                uu_n[j].remove(d1[i])
                            
                   
                    keep_p=uu_p+dd_p
                    keep_n=uu_n+dd_n
                
                
                pp.append(keep_p)
                nn.append(keep_n)
                
                p1=[]
                n1=[]
                
            
            invest1=[]
            N_1=[]
            c_11=[]
            prof=[]
            
            for i in range(len(pp)):
                for j in range(len(pp[i])):
                    if pp[i][j] !=[]:
                        p1.append(pp[i][j])
                        n1.append(nn[i][j])
                        
            for j in range(len(profit)):
                c_1=cons1
                pro=profit[j]
             
                inv=[]
                npv1=[]
                for k in p1[j]:
                    posi=name.index(k)
                    c_1-=invest[posi]
                    pro+=task[posi]
                prof.append(pro)
                c_11.append(c_1)
                for k in n1[j]:
                    posi=name.index(k)
                    inv.append(invest[posi])
                    npv1.append(task[posi])
                invest1.append(inv)
                N_1.append(npv1)
                            
            name1,project,profit,c1,invest1,N_1= constrain_select_one(n1,p1,prof,c_11,invest1,N_1) 
            if len(name1)<1 and len(project)<1:
                messagebox.showwarning("Alert!", "There is no feasible solution") 
                return
            #return project,name1,invest1,resource1 ,N_1,c1,profit
            return project,name1,invest1,N_1,c1,prof
        
        def keep_remove_two(project,profit,u1,d1,name1,name,invest,resource,cons1,cons2):
            
            
            pp=[]
            nn=[]
            
            for k in range(len(name1)):
                keep_p=[copy.deepcopy(project[k])]
                keep_n=[copy.deepcopy(name1[k])]
                
                for i in range(len(u1)):
                    
                    uu_p=copy.deepcopy(keep_p)
                    dd_p=copy.deepcopy(keep_p)
                    uu_n=copy.deepcopy(keep_n)
                    dd_n=copy.deepcopy(keep_n)
                    
                    for j in range(0,2**i):
                        
                        if (u1[i] in keep_p[j]) :
                            if (d1[i] in keep_n[j]):
                                uu_n[j].remove(d1[i])
                                dd_p[j].remove(u1[i])
                                dd_n[j].remove(d1[i])
                        elif (d1[i] in keep_p[j]):
                            if (u1[i] in keep_n[j]) :
                                dd_n[j].remove(u1[i])
                                uu_p[j].remove(d1[i])
                                uu_n[j].remove(u1[i])
                        
                                
                        elif (u1[i] not in keep_p[j]) and (d1[i] not in keep_p[j]):
                           
                                
                              
                            if(u1[i] in keep_n[j]) and (d1[i] in keep_n[j]):
                                uu_p[j].append(u1[i])
                                uu_n[j].remove(d1[i])
                                uu_n[j].remove(u1[i])
                                dd_p[j].append(d1[i])
                                dd_n[j].remove(d1[i])
                                dd_n[j].remove(u1[i])
                               
                            elif (u1[i] in keep_n[j]) and (d1[i] not in keep_n[j]):
                                uu_p[j].append(u1[i])
                                uu_n[j].remove(u1[i])
                                
                                dd_n[j].remove(u1[i])
                                
                            elif  (d1[i] in keep_n[j] )and (u1[i] not in keep_n[j]):
                                
                                dd_p[j].append(d1[i])
                                
                                dd_n[j].remove(d1[i])
                                
                                uu_n[j].remove(d1[i])
                            '''    
                            elif(d1[i] not in keep_n[j] )and (u1[i] not in keep_n[j]):
                                
                                uu_p[j]=[]
                                uu_n[j]=[]
                                dd_p[j]=[]
                                dd_n[j]=[]
                            '''
                       
                    keep_p=uu_p+dd_p
                    keep_n=uu_n+dd_n
                
                
                pp.append(keep_p)
                nn.append(keep_n)
                
                p1=[]
                n1=[]
                
            
            invest1=[]
            resource1=[]
            N_1=[]
            c_11=[]
            c_22=[]
            prof=[]
            
            for i in range(len(pp)):
                for j in range(len(pp[i])):
                    if pp[i][j] !=[]:
                        p1.append(pp[i][j])
                        n1.append(nn[i][j])
                        
            for j in range(len(profit)):
                c_1=cons1
                c_2=cons2
                pro=profit[j]
             
                res=[]
                inv=[]
                npv1=[]
                for k in p1[j]:
                    posi=name.index(k)
                    c_1-=invest[posi]
                    c_2-=resource[posi]
                    pro+=task[posi]
                prof.append(pro)
                c_11.append(c_1)
                c_22.append(c_2)
                for k in n1[j]:
                    posi=name.index(k)
                    inv.append(invest[posi])
                    res.append(resource[posi])
                    npv1.append(task[posi])
                invest1.append(inv)
                resource1.append(res)
                N_1.append(npv1)
                      
            name1,project,profit,c1,c2,invest1,resource1,N_1= constrain_select(n1,p1,prof,c_11,c_22,invest1,resource1,N_1)
            if len(name1)<1 and len(project)<1:
                messagebox.showwarning("Alert!", "There is no feasible solution") 
                return
            return project,name1,invest1,resource1 ,N_1,c_11,c_22,profit
   
        
        def result_one(n,c,w,v,name,project,profit):
            res=[[0 for j in range(c+1)] for i in range(n+1)]
            for i in range(1,n+1):
                for j in range(1,c+1):
                    if w[i-1]<=j:
                        if res[i-1][j]<res[i-1][j-w[i-1]]+v[i-1]: 
                            res[i][j] = res[i-1][j-w[i-1]]+v[i-1]
                        elif res[i-1][j]>=res[i-1][j-w[i-1]]+v[i-1]:
                            res[i][j] = res[i-1][j]
                    else:
                        res[i][j]=res[i-1][j]
            result=res[n][c]+profit
            for i in reversed(range(0,n)):
                if i!=0:
                    if res[i][c]>res[i-1][c]:
                        project.append(name[i-1])
                        c-=w[i-1]
                       
                
                else:
                    if res[i][c]>0:
                        project.append(name[i])
        
            resl = list(map(lambda x : x + profit, res[n]))
            return result, project,resl
        def loop_result_one(project,name1,invest1,task1,c1,profit,invest,name):
            
                
            if len(project)<1:
                proj_cand=project
                name_cand=name1
                invest_cand=invest1.copy()
                
                task_cand=task1.copy()
                c1_cand=c1
                
                prof_cand=profit
                result_cand, project_cand,plotdata=result_one(len(name_cand),c1_cand,invest_cand,task_cand,name_cand,proj_cand,prof_cand)
                if result_cand==0:
                    messagebox.showwarning("Alert!", "There is no feasible solution") 
                    return
            
                  
            else:
                
                proj_cand=project[0]
                name_cand=name1[0]
                invest_cand=invest1[0].copy()
                
                task_cand=task1[0].copy()
                c1_cand=c1[0]
                
                prof_cand=profit[0]
                result_cand, project_cand,plotdata=result_one(len(name_cand),c1_cand,invest_cand,task_cand,name_cand,proj_cand,prof_cand)
                if len(invest1)>1:
                    c=0
                    d=[]
                    for i in range(1,len(invest1)):
                        proj_cand=project[i]
                        name_cand=name1[i]
                        invest_cand=invest1[i].copy()
                        
                        task_cand=task1[i].copy()
                        c1_cand=c1[i]
                       
                        prof_cand=profit[i]
                        c, d,p=result_one(len(name_cand),c1_cand,invest_cand,task_cand,name_cand,proj_cand,prof_cand)
                        if c>result_cand:
                            result_cand=c
                            project_cand=d
                            plotdata=p
                            
            if result_cand==0:
                
                    messagebox.showwarning("Alert!", "There is no feasible solution") 
                    return
            inv=0
            inv_m=[]
            npv_m=[]
            for i in project_cand:
                posi=name.index(i)
                inv+=invest[posi]
                npv_m.append(task[posi])
                inv_m.append(invest[posi])
                
            return  result_cand, project_cand,inv_m,npv_m,inv,plotdata              

            
        def result_two(n,c,d,w,b,v,name,project,profit):
            res=[[[0 for k in range(d+1)] for j in range(c+1)] for i in range(n+1)]
            for i in range(1,n+1):
                for j in reversed(range(w[i-1],c+1)):
                    for k in reversed(range(b[i-1],d+1)):
                        res[i][j][k]=max(res[i-1][j][k],res[i-1][j-w[i-1]][k-b[i-1]]+v[i-1])
            result=res[n][c][d]+profit
            for i in reversed(range(0,n)):
                if i!=0:
                    if res[i][c][d]>res[i-1][c][d]:
                        project.append(name[i-1])
                        c-=w[i-1]
                        d-=b[i-1]
                
                else:
                    if res[i][c][d]>0:
                        project.append(name[i])
            ccc=np.transpose(res[-1]).tolist()[-1]
            resl = list(map(lambda x : x + profit, ccc))
            return result, project,resl
        
        def loop_result_two(project,name1,invest1,resource1,task1,c1,c2,profit,invest,resource,name,task):
            
            if len(project)<1:
                proj_cand=project
                name_cand=name1
                invest_cand=invest1
                resource_cand=resource1
                task_cand=task1
                c1_cand=c1
                c2_cand=c2
                prof_cand=profit
                result_cand, project_cand,plotdata=result_two(len(name_cand),c1_cand,c2_cand,invest_cand,resource_cand,task_cand,name_cand,proj_cand,prof_cand)
                if result_cand==0:
                    messagebox.showwarning("Alert!", "There is no feasible solution") 
                    return
          
            else:
                
                proj_cand=project[0]
                name_cand=name1[0]
                invest_cand=invest1[0].copy()
                resource_cand=resource1[0].copy()
                task_cand=task1[0].copy()
                c1_cand=c1[0]
                c2_cand=c2[0]
                prof_cand=profit[0]
                result_cand, project_cand,plotdata=result_two(len(name_cand),c1_cand,c2_cand,invest_cand,resource_cand,task_cand,name_cand,proj_cand,prof_cand)
                if len(invest1)>1:
                    c=0
                    d=[]
                    for i in range(1,len(invest1)):
                        proj_cand=project[i]
                        name_cand=name1[i]
                        invest_cand=invest1[i].copy()
                        resource_cand=resource1[i].copy()
                        task_cand=task1[i].copy()
                        c1_cand=c1[i]
                        c2_cand=c2[i]
                        prof_cand=profit[i]
                        c, d,p=result_two(len(name_cand),c1_cand,c2_cand,invest_cand,resource_cand,task_cand,name_cand,proj_cand,prof_cand)
                        if c>result_cand:
                            result_cand=c
                            project_cand=d
                            plotdata=p
            if result_cand==0:
                    messagebox.showwarning("Alert!", "There is no feasible solution") 
                    return
            inv=0
            res=0
            inv_m=[]
            res_m=[]
            npv_m=[]
            for i in project_cand:
                posi=name.index(i)
                inv+=invest[posi]
                npv_m.append(task[posi])
                inv_m.append(invest[posi])
                res+=resource[posi]
                res_m.append(resource[posi])
            return  result_cand, project_cand,inv_m,res_m,npv_m,inv,res, plotdata             

        def show_table_two(data1, data2,data3,data4,data5,inv,res,string_1):
            
            
            
            window = tk.Toplevel()
            window.resizable(width=True, height=True)
            window.title("Table")
            main = tk.Label(window, text=string_1, font=('Calibri', 24, 'bold') )
            main.pack()
           
            style = ttk.Style()
            style.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=40)  # Modify the font of the body
            style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
            
            
            tree = ttk.Treeview(window, selectmode="extended", show='tree',height="30", style="mystyle.Treeview")
            tree["columns"] = ("", "", "", "","")
            tree.column("#0", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#1",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#2", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#3", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#4", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            
           
           
            
            tree.tag_configure("evenrow",background='white',foreground='black')
            tree.tag_configure("oddrow",background='black',foreground='white')
            
            tree.insert('', 0, values=['Projects Selected', 'Investment', 'NPV','Number of Resources'],tags = ('oddrow',))
            m=len(data2)
            for i in range(len(data2)):
                
                tree.insert('', i+1, values=[str(data2[i]), '$'+str(data3[i]),'$'+str(data4[i]),str(data5[i])],tags = ('evenrow',))
                                           
            tree.insert('', m+1, values=['Total', '$'+str(int(inv)), '$'+str(int(data1)),str(int(res))],tags = ('oddrow',))
            
            tree.pack() 
            
            
            window1 = tk.Toplevel()
            window1.resizable(width=True, height=True)
            window1.title("Table")
            main1 = tk.Label(window1, text='Constraints', font=('Calibri', 24, 'bold') )
            main1.pack()
            style1 = ttk.Style()
            style1.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=60)  # Modify the font of the body
            style1.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
           
            
            tree2 = ttk.Treeview(window1, selectmode="extended", show='tree',height="100", style="mystyle.Treeview")
            tree2["columns"] = ("stat", "sdata", "perc", "pdata","mdata")
            
            tree2.column("stat",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("sdata", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("perc", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            tree2.column("pdata", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("mdata", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            
            tree2.tag_configure("evenrow",background='white',foreground='black')
            tree2.tag_configure("oddrow",background='black',foreground='white')
            
            tree2.insert('', 0, values=['Investment', 'Number of',' Resources'],tags = ('oddrow',))
            tree2.insert('', 1, values=['$'+str(int(field_11.get())), str(int(field_21.get()))],tags = ('evenrow',))
            tree2.insert('', 3, values=['Fixed', 'Either','Or','If','Then'],tags = ('oddrow',))
            tree2.insert('', 4, values=[str(fix_1.get()), str(either_1.get()),str(or_1.get()),str(if_1.get()),str(then_1.get())],tags = ('evenrow',))
            tree2.insert('', 5, values=[str(fix_2.get()), str(either_2.get()),str(or_2.get()),str(if_2.get()),str(then_2.get())],tags = ('evenrow',))
            tree2.insert('', 6, values=[str(fix_3.get()), str(either_3.get()),str(or_3.get()),str(if_3.get()),str(then_3.get())],tags = ('evenrow',))
            tree2.insert('', 7, values=[str(fix_4.get()), str(either_4.get()),str(or_4.get()),str(if_4.get()),str(then_4.get())],tags = ('evenrow',))
            tree2.insert('', 8, values=[str(fix_5.get()), str(either_5.get()),str(or_5.get()),str(if_5.get()),str(then_5.get())],tags = ('evenrow',))
              
            tree2.pack() 
            
        def show_table_one(data1, data2,data3,data4,inv,string_1):
            """ Displays the statistics of the generated data in a table format.
            :param data: The data.
            :param string_1: The title of the table.
            """
            window = tk.Toplevel()
            window.resizable(width=True, height=True)
            window.title("Table")
            main = tk.Label(window, text=string_1, font=('Calibri', 24, 'bold') )
            main.pack()
           
            style = ttk.Style()
            style.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=40)  # Modify the font of the body
            style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
            
            
            tree = ttk.Treeview(window, selectmode="extended", show='tree',height="30", style="mystyle.Treeview")
            tree["columns"] = ("", "", "", "","")
            tree.column("#0", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#1",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#2", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#3", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree.column("#4", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            
           
           
            
            tree.tag_configure("evenrow",background='white',foreground='black')
            tree.tag_configure("oddrow",background='black',foreground='white')
            
            tree.insert('', 0, values=['Projects Selected', 'Investment', 'NPV','Number of Resources'],tags = ('oddrow',))
            m=len(data2)
            for i in range(len(data2)):
                
                tree.insert('', i+1, values=[str(data2[i]), '$'+str(data3[i]),'$'+str(data4[i]),''],tags = ('evenrow',))
                                           
            tree.insert('', m+1, values=['Total', '$'+str(int(inv)), '$'+str(int(data1)),''],tags = ('oddrow',))
            
            tree.pack() 
            
            
            window1 = tk.Toplevel()
            window1.resizable(width=True, height=True)
            window1.title("Table")
            main1 = tk.Label(window1, text='Constraints', font=('Calibri', 24, 'bold') )
            main1.pack()
            style1 = ttk.Style()
            style1.configure("mystyle.Treeview", font=('Calibri', 20),rowheight=60)  # Modify the font of the body
            style1.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders
           
            
            tree2 = ttk.Treeview(window1, selectmode="extended", show='tree',height="100", style="mystyle.Treeview")
            tree2["columns"] = ("stat", "sdata", "perc", "pdata","mdata","")
            
            tree2.column("stat",stretch=tk.YES, minwidth=100, width=200,anchor="center")
            tree2.column("sdata", stretch=tk.YES, minwidth=100, width=150,anchor="center")
            tree2.column("perc", stretch=tk.YES, minwidth=50, width=100,anchor="w")
            tree2.column("pdata", stretch=tk.YES, minwidth=50, width=100,anchor="center")
            tree2.column("mdata", stretch=tk.YES, minwidth=50, width=60,anchor="center")
            tree2.column("", stretch=tk.YES, minwidth=100, width=200,anchor="center")
            
            
            tree2.tag_configure("evenrow",background='white',foreground='black')
            tree2.tag_configure("oddrow",background='black',foreground='white')
            
            tree2.insert('', 0, values=['Investment', 'Number of \n Resources'],tags = ('oddrow',))
            tree2.insert('', 1, values=['$'+str(int(field_11.get())), ''],tags = ('evenrow',))
            tree2.insert('', 3, values=['Fixed', 'Either','Or','If','Then'],tags = ('oddrow',))
            tree2.insert('', 4, values=[str(fix_1.get()), str(either_1.get()),str(or_1.get()),str(if_1.get()),str(then_1.get())],tags = ('evenrow',))
            tree2.insert('', 5, values=[str(fix_2.get()), str(either_2.get()),str(or_2.get()),str(if_2.get()),str(then_2.get())],tags = ('evenrow',))
            tree2.insert('', 6, values=[str(fix_3.get()), str(either_3.get()),str(or_3.get()),str(if_3.get()),str(then_3.get())],tags = ('evenrow',))
            tree2.insert('', 7, values=[str(fix_4.get()), str(either_4.get()),str(or_4.get()),str(if_4.get()),str(then_4.get())],tags = ('evenrow',))
            tree2.insert('', 8, values=[str(fix_5.get()), str(either_5.get()),str(or_5.get()),str(if_5.get()),str(then_5.get())],tags = ('evenrow',))
            
            tree2.pack() 
        '''
        def plot1(x, y):
            window1 = tk.Toplevel()
            window1.resizable(width=True, height=True)
            window1.title("Graph")
            main1 = tk.Label(window1, text='Constraints', font=('Calibri', 24, 'bold') )
            main1.pack()
            f = pyplot.Figure(figsize=(3,4), dpi=100)
            ax1 = f.add_subplot(111)
            canvas = FigureCanvasTkAgg(f, self)
            canvas.show()
            canvas.get_tk_widget().grid(row = 0, column = 0, rowspan = 3, columnspan = 1, sticky = W+E+N+S) 
            ax1.set_xticklabels('Investment, $')
            ax1.set_yticklabels('NPV, $')
            
            for tick in ax1.xaxis.get_major_ticks():
            	tick.label.set_fontsize(6) 
            			 
            	tick.label.set_rotation(30)
            
            		
            ax1.set_title('Efficient Frontier') 
            ax1.plot(x,y) 
        '''

        
            
            
        def get_value(directory):
            
            global name,task,  t, mim, maxm,invest,resource,distr
            global fix_ini, either_ini, or_ini, if_ini, then_ini
            global project,profit,c1,c2,invest1,resource1,task1,name1
            global result_cand, project_cand,inv,res,plotdata
            
    
            
            name,task,t, mim, maxm,invest,resource,sim_task,distr= read_Port_excel()
            cons2, fix_ini, either_ini, or_ini, if_ini, then_ini=combine_input()
            if cons2==0:
                if fix_ini =='' and if_ini =='' and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    invest1=invest.copy()
                    task1=task.copy()
                    name1=name.copy()
                elif fix_ini =='' and if_ini !=''and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    invest1=invest.copy()
                    task1=task.copy()
                    name1=name.copy()
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                elif fix_ini =='' and if_ini =='' and  either_ini !='':
                    project=[[]]
                    profit=[0]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    invest1=[invest.copy()]
                    task1=[task.copy()]
                    name1=[name.copy()]
                    
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
                
                elif fix_ini!=''and if_ini =='' and  either_ini =='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1=[name1]
                    project=[project]
                    profit=[profit]
                    c1=[c1]
                    invest1=[invest1]
                    task1=[task1]
                
                elif fix_ini!=''and if_ini!='' and  either_ini =='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                
                elif fix_ini!=''and if_ini=='' and  either_ini !='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    project=[project]
                    profit=[profit]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    invest1=[invest1.copy()]
                    task1=[task1.copy()]
                    name1=[name1.copy()]
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
              
                elif fix_ini!=''and if_ini!='' and  either_ini !='':    
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
                
                
                result_cand, project_cand,inv_m,npv_m,inv,plotdata=loop_result_one(project,name1,invest1,task1,c1,profit,invest,name) 
                
                #f=pyplot.Figure(figsize=(5,5),dpi=100)
                #a=f.add_subplot(111)
                #y=plotdata
                #x = np.linspace(0, int(field_11.get()), num=len(y),endpoint=True)
                #plot1(x, y)
                
                #a.plot(x,y)
                #a.xlabel('Investment, $')
                #a.ylabel('NPV, $')
                #a.title('Efficient Frontier')
                #canvas=FigureCanvasTkAgg(f,self)
                #canvas.show()
                #canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.Both, expand=True)


                
                y=plotdata
                x = np.linspace(0, int(field_11.get()), num=len(y),endpoint=True)
                
                pyplot.plot(x,y)
                pyplot.xlabel('Investment, $')
                pyplot.ylabel('NPV, $')
                
                pyplot.title(' Efficient Frontier')
                pyplot.show()
                
                
                show_table_one(result_cand, project_cand,inv_m,npv_m,inv,'Results')
                
                
            
            
            elif cons2!=0:
                
                if fix_ini =='' and if_ini =='' and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    c2=int(field_21.get())
                    invest1=invest.copy()
                    resource1=resource.copy()
                    task1=task.copy()
                    name1=name.copy()
                elif fix_ini =='' and if_ini !=''and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    c2=int(field_21.get())
                    invest1=invest.copy()
                    resource1=resource.copy()
                    task1=task.copy()
                    name1=name.copy()
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                elif fix_ini =='' and if_ini =='' and  either_ini !='':
                    project=[[]]
                    profit=[0]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    c2=[int(field_21.get())]
                    invest1=[invest.copy()]
                    resource1=[resource.copy()]
                    task1=[task.copy()]
                    name1=[name.copy()]
                    
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
                
                elif fix_ini!=''and if_ini =='' and  either_ini =='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    
                    name1=[name1]
                    project=[project]
                    profit=[profit]
                    c1=[c1]
                    c2=[c2]
                    invest1=[invest1]
                    resource1=[resource1]
                    task1=[task1]
                
                elif fix_ini!=''and if_ini!='' and  either_ini =='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                
                elif fix_ini!=''and if_ini=='' and  either_ini !='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    project=[project]
                    profit=[profit]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    c2=[int(field_21.get())]
                    invest1=[invest1.copy()]
                    resource1=[resource1.copy()]
                    task1=[task1.copy()]
                    name1=[name1.copy()]
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
              
                elif fix_ini!=''and if_ini!='' and  either_ini !='':    
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
                
                
                
                result_cand, project_cand,inv_m,res_m,npv_m,inv,res,plotdata =loop_result_two(project,name1,invest1,resource1 ,task1,c1,c2,profit,invest,resource,name,task) 
                y=plotdata
                x = np.linspace(0, int(field_11.get()), num=len(y),endpoint=True)
                
                pyplot.plot(x,y)
                pyplot.xlabel('Investment, $')
                pyplot.ylabel('NPV, $')
                
                pyplot.title(' Efficient Frontier')
                pyplot.show()
                
                  
                show_table_two(result_cand, project_cand,inv_m,npv_m,res_m,inv,res,'Results')
        
       
                
        def get_simulate(directory):
            
            global name,task,  t, mim, maxm,invest,resource,sim_task,distr
            global conss1,conss2, cons2, fix_ini, either_ini, or_ini, if_ini, then_ini
            global project,profit,c1,c2,invest1,resource1,task1,name1
            global result_cand, project_cand,inv_m,res_m,npv_m,inv,res,plotdata
            global fix_11, either_11,or_11,if_11,then_11
            global fix_22, either_22,or_22,if_22,then_22
            global fix_33, either_33,or_33,if_33,then_33
            global fix_44, either_44,or_44,if_44,then_44
            global fix_55, either_55,or_55,if_55,then_55

            
            fix_11=str(fix_1.get())
            either_11=str(either_1.get())
            or_11=str(or_1.get())
            if_11=str(if_1.get())
            then_11=str(then_1.get())
            fix_22=str(fix_2.get())
            either_22=str(either_2.get())
            or_22=str(or_2.get())
            if_22=str(if_2.get())
            then_22=str(then_2.get())
            fix_33=str(fix_3.get()) 
            either_33=str(either_3.get())
            or_33=str(or_3.get())
            if_33=str(if_3.get())
            then_33=str(then_3.get())
            fix_44=str(fix_4.get())
            either_44=str(either_4.get())
            or_44=str(or_4.get())
            if_44=str(if_4.get())
            then_44=str(then_4.get())
            fix_55=str(fix_5.get())
            either_55=str(either_5.get())
            or_55=str(or_5.get())
            if_55=str(if_5.get())
            then_55=str(then_5.get())
            
            conss1=str(int(field_11.get()))
            
            name,task,t, mim, maxm,invest,resource,sim_task,distr= read_Port_excel()
            cons2, fix_ini, either_ini, or_ini, if_ini, then_ini=combine_input()
            
            task=sim_task
            if cons2==0:
                conss2=str(field_21.get())
                if fix_ini =='' and if_ini =='' and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    invest1=invest.copy()
                    task1=task.copy()
                    name1=name.copy()
                elif fix_ini =='' and if_ini !=''and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    invest1=invest.copy()
                    task1=task.copy()
                    name1=name.copy()
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                elif fix_ini =='' and if_ini =='' and  either_ini !='':
                    project=[[]]
                    profit=[0]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    invest1=[invest.copy()]
                    task1=[task.copy()]
                    name1=[name.copy()]
                    
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
                
                elif fix_ini!=''and if_ini =='' and  either_ini =='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1=[name1]
                    project=[project]
                    profit=[profit]
                    c1=[c1]
                    invest1=[invest1]
                    task1=[task1]
                
                elif fix_ini!=''and if_ini!='' and  either_ini =='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                
                elif fix_ini!=''and if_ini=='' and  either_ini !='':
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    project=[project]
                    profit=[profit]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    invest1=[invest1.copy()]
                    task1=[task1.copy()]
                    name1=[name1.copy()]
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
              
                elif fix_ini!=''and if_ini!='' and  either_ini !='':    
                    project,profit,c1,invest1,task1,name1= fixed_one(len(fix_ini),[],0,int(field_11.get()),fix_ini,invest,task,name)
                    name1,project,profit, c1,invest1, task1=if_then_one(len(if_ini),project,profit,c1,if_ini,then_ini,invest,task,name,name1)
                    project,name1,invest1,task1,c1,profit=keep_remove_one(project,profit,either_ini,or_ini,name1,name,invest,int(field_11.get()))
                result_cand, project_cand,inv_m,npv_m,inv,plotdata=loop_result_one(project,name1,invest1,task1,c1,profit,invest,name) 
               # show_table_one(result_cand, project_cand,inv_m,npv_m,inv,'Results')
            elif cons2!=0:
                conss2=str(int(field_21.get()))
                if fix_ini =='' and if_ini =='' and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    c2=int(field_21.get())
                    invest1=invest.copy()
                    resource1=resource.copy()
                    task1=task.copy()
                    name1=name.copy()
                elif fix_ini =='' and if_ini !=''and  either_ini =='':
                    project=[]
                    profit=0
                    c1=int(field_11.get())
                    c2=int(field_21.get())
                    invest1=invest.copy()
                    resource1=resource.copy()
                    task1=task.copy()
                    name1=name.copy()
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                elif fix_ini =='' and if_ini =='' and  either_ini !='':
                    project=[[]]
                    profit=[0]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    c2=[int(field_21.get())]
                    invest1=[invest.copy()]
                    resource1=[resource.copy()]
                    task1=[task.copy()]
                    name1=[name.copy()]
                    
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
                
                elif fix_ini!=''and if_ini =='' and  either_ini =='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    
                    name1=[name1]
                    project=[project]
                    profit=[profit]
                    c1=[c1]
                    c2=[c2]
                    invest1=[invest1]
                    resource1=[resource1]
                    task1=[task1]
                
                elif fix_ini!=''and if_ini!='' and  either_ini =='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                
                elif fix_ini!=''and if_ini=='' and  either_ini !='':
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    project=[project]
                    profit=[profit]*(2**len(either_ini))
                    c1=[int(field_11.get())]
                    c2=[int(field_21.get())]
                    invest1=[invest1.copy()]
                    resource1=[resource1.copy()]
                    task1=[task1.copy()]
                    name1=[name1.copy()]
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
              
                elif fix_ini!=''and if_ini!='' and  either_ini !='':    
                    project,profit,c1,c2,invest1,resource1,task1,name1 = fixed_two(len(fix_ini),[],0,int(field_11.get()),int(field_21.get()),fix_ini,invest,resource,task,name)
                    name1,project,profit, c1,c2,invest1,resource1 ,task1=if_then_two(len(if_ini),project,profit,c1,c2,if_ini,then_ini,invest,resource,task,name,name1)
                    project,name1,invest1,resource1 ,task1,c1,c2,profit=keep_remove_two(project,profit,either_ini,or_ini,name1,name,invest,resource,int(field_11.get()),int(field_21.get()))
          
                result_cand, project_cand,inv_m,res_m,npv_m,inv,res,plotdata =loop_result_two(project,name1,invest1,resource1 ,task1,c1,c2,profit,invest,resource,name,task) 
                '''
                y=plotdata
                x = np.linspace(0, int(field_11.get()), num=len(y),endpoint=True)
                #f1 = np.polyfit(x, y, 3)
                #ploor = np.poly1d(f1)
                #yval=ploor(x)
                pyplot.plot(x,y)
                pyplot.xlabel('Investment, $')
                pyplot.ylabel('NPV, $')
                
                pyplot.title('Simulated Efficient Frontier')
                pyplot.show()
                #show_table_two(result_cand, project_cand,inv_m,npv_m,res_m,inv,res,'Results')
                '''
       
                

      
        
        def page2():
            get_simulate(directory)
            controller.show_frame(Optimization_Simulation)
            
        def clear_val():
            global name,task,  t, mim, maxm,invest,resource,sim_task,distr
            global conss1,conss2,cons2, fix_ini, either_ini, or_ini, if_ini, then_ini
            global project,profit,c1,c2,invest1,resource1,task1,name1
            global result_cand, project_cand,inv_m,res_m,npv_m,inv,res,plotdata
            global fix_11, either_11,or_11,if_11,then_11
            global fix_22, either_22,or_22,if_22,then_22
            global fix_33, either_33,or_33,if_33,then_33
            global fix_44, either_44,or_44,if_44,then_44
            global fix_55, either_55,or_55,if_55,then_55
            inv_m=[]
            res_m=[]
            npv_m=[]
            fix_11=[]
            either_11=[]
            or_11=[]
            if_11=[]
            then_11=[]
            fix_22=[]
            either_22=[]
            or_22=[]
            if_22=[]
            then_22=[]
            fix_33=[]
            either_33=[]
            or_33=[]
            if_33=[]
            then_33=[]
            fix_44=[]
            either_44=[]
            or_44=[]
            if_44=[]
            then_44=[]
            fix_55=[] 
            either_55=[]
            or_55=[]
            if_55=[]
            then_55=[]
            conss1=[] 
            conss2=[]
            name=[]
            task=[]
            t=[]
            mim=[]
            maxm=[]
            invest=[]
            resource=[]
            distr=[]
            plotdata=[]
            sim_task=[]
            cons2=[]
            fix_ini=[]
            either_ini=[]
            or_ini=[]
            if_ini=[]
            then_ini=[]
            project=[]
            profit=[]
            c1=[]
            c2=[]
            invest1=[]
            resource1=[]
            task1=[]
            name1=[]
            result_cand=[]
            project_cand=[]
            inv=[]
            res=[]
                   
            del name,task,  t, mim, maxm,invest,resource,sim_task,distr
            del conss1, conss2,cons2, fix_ini, either_ini, or_ini, if_ini, then_ini
            del project,profit,c1,c2,invest1,resource1,task1,name1
            del result_cand, project_cand,inv_m,res_m,npv_m,inv,res,plotdata
            del fix_11, either_11,or_11,if_11,then_11
            del fix_22, either_22,or_22,if_22,then_22
            del fix_33, either_33,or_33,if_33,then_33
            del fix_44, either_44,or_44,if_44,then_44
            del fix_55, either_55,or_55,if_55,then_55
            controller.show_frame(Optimization)
        
        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        center2_frame = tk.Frame(self)

        center2_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Portfolio Optimization | Settings", font=HEADR_FONT)

        label_1.pack()
        
        label_4 = tk.Label(center2_frame, text=" Constraints")

        label_4.grid(row=1, column=0, columnspan=5)
        
        separate = ttk.Separator(center2_frame, orient="horizontal")

        separate.grid(row=2, column=0, columnspan=5, sticky="ew", pady=5)

        check_1 = ttk.Label(center2_frame, text="Investment")

        check_1.grid(row=3, column=1, sticky='w')

        check_1.state(['!alternate'])

        field_11 = ttk.Entry(center2_frame, width=12)

        field_11.grid(row=3, column=2, columnspan=2)

        check_2 = ttk.Checkbutton(center2_frame, text="Resources", variable=var_cons2)

        check_2.grid(row=4, column=1, sticky='w')

        check_2.state(['!alternate'])

        field_21 = ttk.Entry(center2_frame, width=12)

        field_21.grid(row=4, column=2, columnspan=2)

        separate = ttk.Separator(center2_frame, orient="horizontal")

        separate.grid(row=5, column=0, columnspan=5, sticky="ew", pady=5)

        

        
        label_5 = tk.Label(center2_frame, text="Number of Other Constraints")

        label_5.grid(row=6, column=0, columnspan=5)
        
        numb_1=ttk.Entry(center2_frame, width=12)
        
        numb_1.grid(row=7, column=0, columnspan=1 ,sticky='w')
        
        numb_2=ttk.Entry(center2_frame, width=12)
        
        numb_2.grid(row=7, column=2, columnspan=1 ,sticky='w')
        
       
        numb_3=ttk.Entry(center2_frame, width=12)
        
        numb_3.grid(row=7, column=4, columnspan=1 ,sticky='w')
        
        label_6 = tk.Label(center2_frame, text="Insert Project ID")

        label_6.grid(row=8, column=0, columnspan=5)
    

        check_3 = ttk.Checkbutton(center2_frame, text="      Fixed       ", variable=var_fix)

        check_3.grid(row=9, column=0, sticky='w')

        check_3.state(['!alternate'])

        check_4 = ttk.Checkbutton(center2_frame, text="      Either                    Or            ", variable=var_either)

        check_4.grid(row=9, column=1,columnspan=2, sticky='w')

        check_4.state(['!alternate'])

        check_5 = ttk.Checkbutton(center2_frame, text="        If                        Then           ", variable=var_if)

        check_5.grid(row=9, column=3,columnspan=2, sticky='w')

        check_5.state(['!alternate'])
        '''
        label_7 = tk.Label(center2_frame, text="                           ")

        label_7.grid(row=10, column=0, columnspan=5)
        '''
        separate3 = ttk.Separator(center2_frame, orient="horizontal")

        separate3.grid(row=10, column=0, columnspan=5, sticky="w", pady=2)
       

        fix_1 = ttk.Entry(center2_frame, width=12)

        fix_1.grid(row=11, column=0, sticky='ew')

        either_1 = ttk.Entry(center2_frame, width=12)

        either_1.grid(row=11, column=1, sticky='ew')

        or_1 = ttk.Entry(center2_frame, width=12)

        or_1.grid(row=11, column=2, sticky='ew')

        if_1 = ttk.Entry(center2_frame, width=12)

        if_1.grid(row=11, column=3, sticky='ew')

        then_1 = ttk.Entry(center2_frame, width=12)

        then_1.grid(row=11, column=4, sticky='ew')
        
        fix_2 = ttk.Entry(center2_frame, width=12)

        fix_2.grid(row=12, column=0, sticky='ew')

        either_2 = ttk.Entry(center2_frame, width=12)

        either_2.grid(row=12, column=1, sticky='ew')

        or_2 = ttk.Entry(center2_frame, width=12)

        or_2.grid(row=12, column=2, sticky='ew')

        if_2 = ttk.Entry(center2_frame, width=12)

        if_2.grid(row=12, column=3, sticky='ew')

        then_2 = ttk.Entry(center2_frame, width=12)

        then_2.grid(row=12, column=4, sticky='ew')
        
        
        fix_3 = ttk.Entry(center2_frame, width=12)

        fix_3.grid(row=13, column=0, sticky='ew')

        either_3 = ttk.Entry(center2_frame, width=12)

        either_3.grid(row=13, column=1, sticky='ew')

        or_3 = ttk.Entry(center2_frame, width=12)

        or_3.grid(row=13, column=2, sticky='ew')

        if_3 = ttk.Entry(center2_frame, width=12)

        if_3.grid(row=13, column=3, sticky='ew')

        then_3 = ttk.Entry(center2_frame, width=12)

        then_3.grid(row=13, column=4, sticky='ew')
        
        
        fix_4 = ttk.Entry(center2_frame, width=12)

        fix_4.grid(row=14, column=0, sticky='ew')

        either_4 = ttk.Entry(center2_frame, width=12)

        either_4.grid(row=14, column=1, sticky='ew')

        or_4 = ttk.Entry(center2_frame, width=12)

        or_4.grid(row=14, column=2, sticky='ew')

        if_4 = ttk.Entry(center2_frame, width=12)

        if_4.grid(row=14, column=3, sticky='ew')

        then_4 = ttk.Entry(center2_frame, width=12)

        then_4.grid(row=14, column=4, sticky='ew')
        
        
        fix_5 = ttk.Entry(center2_frame, width=12)

        fix_5.grid(row=15, column=0, sticky='ew')

        either_5 = ttk.Entry(center2_frame, width=12)

        either_5.grid(row=15, column=1, sticky='ew')

        or_5 = ttk.Entry(center2_frame, width=12)

        or_5.grid(row=15, column=2, sticky='ew')

        if_5 = ttk.Entry(center2_frame, width=12)

        if_5.grid(row=15, column=3, sticky='ew')

        then_5 = ttk.Entry(center2_frame, width=12)

        then_5.grid(row=15, column=4, sticky='ew')
        
        separate4 = ttk.Separator(center2_frame, orient="horizontal")

        separate4.grid(row=16, column=0, columnspan=5, sticky="w", pady=5)
       



        button_run = ttk.Button(center2_frame, text='Run: Deterministic Optimization', command=lambda: get_value(directory))

        button_run.grid(row=17, column=1,columnspan=3, sticky='ew')
        
        #button_6=ttk.Button(center2_frame, text='Efficient Frontier')
        
       # button_6.grid(row=18, column=0, sticky="ew")
        
        separate5 = ttk.Separator(center2_frame, orient="horizontal")

        separate5.grid(row=19, column=0, columnspan=5, sticky="w", pady=5)
        
        button_simulate= ttk.Button(center2_frame, text='Input data for Stochastic Optimization', command=lambda: page2())#get_simulate(directory))
        button_simulate.grid(row=20, column=1, columnspan=3,sticky='ew')
        



        button_back = ttk.Button(bottom_frame, text="Back", command=lambda: clear_val())

        button_back.pack(side="right", padx=5)

              
    

####################################################################################################################################
class Decisiontree(tk.Frame):
    """
    This page allows the user to select the excel file that contains the Optimization they would like to import.
    """
    def __init__(self, parent, controller):
        """
        Class constructor.
        :param parent: The parent frame.
        :param controller: The controller runs through every page and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)
        global directory
        directory =""
        def read_decision(director):
            book=xlrd.open_workbook(director)
            sheet=book.sheet_by_name("Sheet1")
            df=int(sheet.cell_value(0,0))
            name = []
            cost = []
            distr = []
            minim = []
            maxim = []
            pu=[]
            pd=[]
            for i in range(0,df):
                name.append(sheet.cell_value(i+1,1))
                cost.append(sheet.cell_value(i+1,2))
                pu.append(sheet.cell_value(i+1,3))
                pd.append(sheet.cell_value(i+1,4))
                distr.append(sheet.cell_value(i+1,5))
                minim.append(sheet.cell_value(i+1,6))
                maxim.append(sheet.cell_value(i+1,7))
            return name, cost, distr, minim, maxim, pu, pd
            
        def get_string():

            global directory

            directory = entry_1.get()

            if directory == "":

                messagebox.askretrycancel("Error", "An error has occurred in your File.")

            else:
                global name, cost, distr, minim, maxim, pu, pd
                name, cost, distr, minim, maxim, pu, pd=read_decision(directory)

                controller.show_frame(DecisionSolution)
        
        

        def clear_value():
            global directory,name, cost, distr, minim, maxim, pu, pd
            directory=[]
            name=[]
            cost=[] 
            distr=[]
            minim=[]
            maxim=[]
            pu=[]
            pd=[]
            del directory,name, cost, distr, minim, maxim, pu, pd
            controller.show_frame(ChoosePage)
        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Decision Tree | File Upload", font=HEADR_FONT)

        label_1.pack()

        message_1 = tk.Message(center_frame, text="Pick the MS Excel file with your Decision_tree data: (.xlsx)\n"

                               "Ensure your data has the following fields: Name, Cost, Distribution "

                               "minimum,  maximum, probabilities",

                               justify="center", width=500, font=MESSG_FONT)

        message_1.pack()

        entry_1 = ttk.Entry(center_frame, width=50)

        entry_1.pack(side="left")

        button_2 = ttk.Button(center_frame, text="File...", command=lambda: open_onclick(entry_1))

        button_2.pack(side="left", pady=5)

        button_3 = ttk.Button(center_frame, text="Open", command=lambda: get_string())

        button_3.pack(side="left", padx=5, pady=5)

        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_value())

        button_1.pack(side="right", padx=5)


class DecisionSolution(tk.Frame):
    """
    This page allows the user to select the excel file that contains the Optimization they would like to import.
    """
    def __init__(self, parent, controller):
        
        """
        Class constructor.
        :param parent: The parent frame.
        :param controller: The controller runs through every page and allows them to call the show_page function.
        """
        tk.Frame.__init__(self, parent)
                   
        def binomial_grid(self,n,earn,pu,pd):
            
            G=nx.Graph() 
            for i in range(1,n+1): 
                j=2
                z=3 
                if i<n:             
                    G.add_edge((i,j),(i+1,j),color='b')
                    G.add_edge((i,j),(i+1,z),color='r')
                    
            edges = G.edges()
            colors = [G[u][v]['color'] for u,v in edges]        
            posG={}    #dictionary with nodes position
            
            x_position=[]
            y_position=[]
            for node in G.nodes():
                x_position.append(node[0]*0.9/(n-1)-0.9/(n-1))
                y_position.append((node[0]-2*node[1]+3)*0.9/n+0.9/n-0.09)
                posG[node]=(node[0],node[0]-2*node[1])
                #x=node[0]*0.9/(n-1)
                #y=(node[0]-2*node[1]+3)*0.9/n
                #plt.figtext(x-0.9/(n-1),y+0.9/n-0.09,'Stock ')
            
            pyplot.figtext(x_position[0],y_position[0],'Decesion'+ ' 1\n '+'$'+str(np.round(earn[0][0])),color='b')
            pyplot.figtext(1/2*(x_position[0]+x_position[1]),(y_position[0]+y_position[1]),str(pu[0]),color='b')
            for i in range(0,int((len(x_position)+1)*1/2)):
               z=2*i+1
               if z<(len(x_position)-2):
                   
                   pyplot.figtext(x_position[z],y_position[z],'Decesion '+ str(i+2)+'\n'+'$'+str(np.round(earn[i+1][0])),color='b')
                   pyplot.figtext(x_position[z+1],y_position[z+1],'Revenue '+ '\n'+'$'+str(np.round(earn[i+1][-1])),color='black')
                   pyplot.figtext(1/2*(x_position[z+1]+x_position[z+2]),2/3*(y_position[-1]+y_position[z+1]+y_position[z]),str(pu[i+1]),color='b')
                   pyplot.figtext(4/9*(x_position[z+1]+x_position[z]),5/14*(y_position[z+1]+y_position[z-1]+y_position[-1]),str(pd[i]),color='r')
            
           # plt.figtext(1/2*(x_position[-2]+x_position[-4]),2/3*(y_position[-2]+y_position[-4]),str(pu[-2]),color='b')
            pyplot.figtext(x_position[-2],y_position[-2],'Revenue '+ '\n'+'$'+str(np.round(earn[-1][0])),color='black')
            pyplot.figtext(x_position[-1],y_position[-1],'Revenue '+ '\n'+'$'+str(np.round(earn[-1][1])),color='black')
            pyplot.figtext(1/2*(x_position[-1]+x_position[-4]),1/2*(y_position[-1]+y_position[-4]),str(pd[-2]),color='r')
            node = G.node()
            fixed_nodes = posG.keys()
            nx.draw(G,pos=posG,edge_color=colors,fixed=fixed_nodes) 
                
        '''
        def binomial_grid(self,n,earn,pu,pd):
            
            #t = tk.Toplevel(self)
            #t.wm_title("Decision Tree" )
            #l = tk.Label(t, text="This is window Decision Tree")
            #l.pack(side="top", fill="both", expand=True, padx=100, pady=100)
            #f = pyplot.figure(figsize=(5,4))
            #a = f.add_subplot(111)
            
            
            G=nx.Graph() 
            for i in range(1,n+1): 
                j=2
                z=3 
                if i<n:             
                    G.add_edge((i,j),(i+1,j),color='b')
                    G.add_edge((i,j),(i+1,z),color='r')
                    
            edges = G.edges()
            colors = [G[u][v]['color'] for u,v in edges]        
            posG={}    #dictionary with nodes position
            
            x_position=[]
            y_position=[]
            for node in G.nodes():
                x_position.append(node[0]*0.9/(n-1)-0.9/(n-1))
                y_position.append((node[0]-2*node[1]+3)*0.9/n+0.9/n-0.09)
                posG[node]=(node[0],node[0]-2*node[1])
                #x=node[0]*0.9/(n-1)
                #y=(node[0]-2*node[1]+3)*0.9/n
                #plt.figtext(x-0.9/(n-1),y+0.9/n-0.09,'Stock ')
            
            pyplot.figtext(x_position[0],y_position[0],'Decesion'+ ' 1\n '+'$'+str(earn[0][0]),color='b')
            pyplot.figtext(1/2*(x_position[0]+x_position[1]),1/2*(y_position[0]+y_position[1]),str(pu[0]),color='b')
            for i in range(0,int((len(x_position)+1)*1/2)):
               z=2*i+1
               if z<(len(x_position)-2):
                   
                   pyplot.figtext(x_position[z],y_position[z],'Decesion '+ str(i+2)+'\n'+'$'+str(earn[i+1][0]),color='b')
                   pyplot.figtext(x_position[z+1],y_position[z+1],'Revenue '+ '\n'+'$'+str(earn[i+1][-1]),color='black')
                   pyplot.figtext(1/2*(x_position[z+1]+x_position[z-1]),1/2*(y_position[z+1]+y_position[z-1]),str(pd[i]),color='r')
            
            pyplot.figtext(1/2*(x_position[-2]+x_position[-4]),1/2*(y_position[-2]+y_position[-4]),str(pu[-2]),color='b')
            pyplot.figtext(x_position[-2],y_position[-2],'Revenue '+ '\n'+'$'+str(earn[-1][0]),color='black')
            pyplot.figtext(x_position[-1],y_position[-1],'Revenue '+ '\n'+'$'+str(earn[-1][1]),color='black')
            pyplot.figtext(1/2*(x_position[-1]+x_position[-4]),1/2*(y_position[-1]+y_position[-4]),str(pd[-2]),color='r')
    
                
            pyplot.title('Decision Tree')
            node = G.node()
            fixed_nodes = posG.keys()
            nx.draw(G,pos=posG,edge_color=colors,fixed=fixed_nodes) 
            #canvas = FigureCanvasTkAgg(f, master=t)
            #canvas.show()
            #canvas.get_tk_widget().pack(row=1, column=0, columnspan = 3, pady=(15,15), padx=(25,25), sticky=tk.N+tk.S+tk.E+tk.W)
        '''
            
        
        def value_calculate( cost, pu, pd):
            t=int(len(cost))
            expe = [[0] * 2 for i in range(t)]
            
            for i in range(0,t-1):
                expe[i+1][1]=expe[i][1]-cost[i]
            expe[t-1][0]=cost[t-1]+expe[t-1][1]
            
            earn=copy.deepcopy(expe)
            for i in reversed(range(1,t)):
                earn[i-1][0]=earn[i][0]*pu[i-1]+earn[i][1]*pd[i-1]
              
            return expe,earn

        
        def value_calculate2( cost, pu, pd):
            t=int(len(name))
            expe = [[0] * 2 for i in range(t)]
            
            for i in range(0,t-1):
                expe[i][1]=-cost[i]
            expe[t-1][0]=cost[t-1]+expe[t-1][1]
            
            earn=np.round(copy.deepcopy(expe))
            for i in reversed(range(1,t)):
                earn[i-1][0]=np.round(earn[i][0]*pu[i-1]+earn[i-1][1])
              
            return expe,earn
 


        def get_value():
            expe,earn=value_calculate( cost, pu, pd)
            
            binomial_grid(self,len(name),earn,pu,pd)
            
        def get_value2():
            
            expe2,earn2=value_calculate2( cost, pu, pd)
            
            binomial_grid(self,len(name),earn2,pu,pd)
        
        def simulation_111(director,itera):
            global simcost, result
            simcost=[]
            for i in range(0, len(cost)):
                if distr[i] == 'Pert':
                
                    if minim[i]=='' or cost[i]=='' or maxim[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                    else:
                        simcost.append(pert_distro(minim[i], cost[i], maxim[i], itera))
    
                elif distr[i] == 'Triangular':
                
                    if minim[i]=='' or cost[i]=='' or maxim[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                    else:
                        simcost.append(tri_distro(minim[i], cost[i], maxim[i], itera))
    
                else:
                    if cost[i]=='':
                        messagebox.showerror("Error", "Missing Spealling/ Formating/ Input.")
                        return
                    else:
                        simcost.append(no_distro(cost[i],itera))
                
            simcost=np.array(simcost).T.tolist()
            result=[]
            for j in range(0,itera):
                ex,ear=value_calculate(simcost[j], pu, pd)
                result.append(ear[0][0])
            messagebox.showwarning("Hey", "Simulation is done.")

   
    
            
        def write_to_excel(tsk, data):
            messagebox.showwarning("Hey","1. The iteration data are stored in new Excel files titled 'Decision Data' "
                                   " inside the source folder. \n \n2. "
                                   "Before you run the next simulation, please make sure Excel files are closed,"
                                   " so that the new iteration data can override the old iteration data.")

   
            wb = xlsxwriter.Workbook('Decision Tree Output Data.xlsx')
            ws = wb.add_worksheet()
            A1 = 'Output Data'
            title = np.append(A1, tsk)
        
            item = title
            column = 0
            
            sim_task=np.array(simcost).T.tolist()
            
            for i in item:
                ws.write(0, column, i)
                column += 1
            row = 1
        
            for i in data:
                ws.write(row, 0, i)
                row += 1
        
            col = 1
            
            for i in sim_task:
                
                r = 1
        
                for j in i:
                    
                    ws.write(r, col, j)
        
                    r += 1
                col += 1
        
            wb.close()
             
              
        def check_distro1(task_no,minim,mode,maxim, distro, task):
            
            i = task_no
            if i > len(distro)-1 or i < 0:
    
                messagebox.showwarning("Hey!", "This name lies outside of the range of specified.")
            elif distro[i]=='':
                messagebox.showerror("Hey!",  task[i] + " has no distribution.")
            elif distro[i] == "Normal":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!", task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    mu = minim[i]
                    sigma = mode[i]
                    x = np.linspace(mu - 5 * sigma, mu + 5 * sigma, 100)
                    pyplot.plot(x, ss.norm.pdf(x, mu, sigma))
                    pyplot.fill(x, ss.norm.pdf(x, mu, sigma))
                    host.set_ylabel("Probability Density")
                    host.set_xlabel(task[i]+", $")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title( task[i]+"\n"+"Normal Distribution")
                    pyplot.show()
            
            elif distro[i] == "Pert":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", task[i] + "\n has missing input data.")
                else:
                  
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    a = minim[i]
                    b = mode[i]
                    c = maxim[i]
                    
                    x = pert_distro(a,b, c, 10000)
                    sns.kdeplot(x, data2=None, shade=True, vertical=False, kernel='gau', bw='scott',
                                    gridsize=100, cut=3, clip=None, legend=True, cumulative=False, shade_lowest=True, cbar=False, cbar_ax=None, cbar_kws=None, ax=None,color="b")
                    
                    host.set_ylabel("Probability Density")
                    host.set_xlabel(task[i]+", $")
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title(task[i]+"\n"+"Pert Distribution")
                    
                    pyplot.show()
                    

    
            elif distro[i] == "Triangular":
                if minim[i]=='' or mode[i]==''or maxim[i]=='':
                    messagebox.showerror("Hey!", task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel(task[i]+", $")
                    c = (mode[i] - minim[i]) / (maxim[i]-minim[i])
                    mean, var, skew, kurt = triang.stats(c, moments='mvsk')
                    x = np.linspace(minim[i], maxim[i], 1000)
                    host.plot(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]), 'b', lw=2)
                    pyplot.fill(x, triang.pdf(x, c, loc=minim[i], scale=maxim[i] - minim[i]))
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title( task[i]+"\n"+"Triangular Distribution")
                    pyplot.show()
                
            elif distro[i]=="Uniform":
                if minim[i]=='' or mode[i]=='':
                    messagebox.showerror("Hey!",  task[i] + "\n has missing input data.")
                else:
                    
                    pyplot.clf()
                    host = pyplot.subplot(111)
                    host.set_ylabel("Probability Density")
                    host.set_xlabel(task[i]+", $")
                    a=minim[i]
                    b=mode[i]
                    mean, var, skew, kurt = ss.uniform.stats(moments='mvsk')
                    x = np.linspace(a, b, 100)
                    host.plot(x, ss.uniform.pdf(x, loc=a, scale=b-1), 'b', lw=2, alpha=0.6, label='uniform pdf')
                    fmt = '{x:,.0f}'
                    tick = mtick.StrMethodFormatter(fmt)
                    host.xaxis.set_major_formatter(tick)
                    pyplot.title( task[i]+"\n"+"Uniform Distribution")
                    pyplot.show()
    
    
       
        def task_stats1(task_no,minim,mode,maxim, distro, task):

            i = task_no
            
            if i > len(distro)-1 or i < 0:

                messagebox.showwarning("Hey!", "This name lies outside of the range of specified.")
            else:
                
               
                mcerp.npts = int(entry_simul.get())
    
                title =  task[i]
    
                if distro[i]=='':
                    messagebox.showerror("Hey!",  task[i] + " has no distribution.")
    
                elif distro[i] == "Normal":
    
                    data = np.random.normal(minim[i], mode[i], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Pert":
    
                    a = minim[i]
    
                    b = mode[i]
    
                    c = maxim[i]
    
                    alpha = ((4 * b) + c - (5 * a)) / (c - a)
    
                    beta = ((5 * c) - a - (4 * b)) / (c - a)
    
                    x = np.random.beta(alpha, beta, int(entry_simul.get()))
    
                    data = a + (x * (c - a))
    
                    make_table(data, title)
    
                elif distro[i] == "Triangular":
    
                    data = np.random.triangular(minim[i], mode[i], maxim[i], int(entry_simul.get()))
    
                    make_table(data, title)
    
                elif distro[i] == "Uniform":
    
                    data = np.random.uniform(minim[i],mode[i], int(entry_simul.get()))
    
                    make_table(data, title)
                    
        def clear_value():
            global simcost, result
            simcost=[]
            result=[]
            del simcost, result
            controller.show_frame(Decisiontree)
            
            

           
        


        top_frame = tk.Frame(self)

        top_frame.pack(side="top", fill="x", expand=True, padx=10, pady=5, anchor="n")

        center_frame = tk.Frame(self)

        center_frame.pack(side="top", fill="y", expand=True, padx=10, pady=5)

        bottom_frame = tk.Frame(self)

        bottom_frame.pack(side="bottom", fill="x", expand=True, padx=10, pady=5)



        label_1 = tk.Label(center_frame, text="Deterministic & Stochastic Decision Tree | File Upload", font=HEADR_FONT)

        label_1.pack()

        
        message_1 = tk.Message(center_frame, text="Check the distributions of the individual project to confirm that "
    
                               "they are correct.\nThen input 'Number of Iterations' and click on 'Run Simulation'.",
                               justify="center", width=500, font=MESSG_FONT)
        message_1.pack()
        message_2 = tk.Message(center_frame, text="",
    
                               justify="center", width=500, font=MESSG_FONT)
    
        message_2.pack()
        
        button_1 = ttk.Button(bottom_frame, text="Back", command=lambda: clear_value())
        button_1.pack(side="right", padx=5)
        entry_frame = tk.Frame(center_frame)
    
        entry_frame.pack()
        button_2 = ttk.Button(entry_frame, text="Distribution",
    
                              command=lambda: check_distro1(int(entry_task.get())-1,minim,cost,maxim, distr, name))
    
        button_2.grid(row=0, column=1,columnspan=2)
    
        button_4 = ttk.Button(entry_frame, text="Statistics",command=lambda:task_stats1(int(entry_task.get())-1,minim,cost,maxim, distr, name))
    
        button_4.grid(row=0, column=3,columnspan=2)
    
        entry_simul = ttk.Entry(entry_frame, width=8)
        entry_simul.grid(row=3, column=0,columnspan=1)
        entry_simul.insert(0, "100")
        entry_task = ttk.Entry(entry_frame, width=8)
    
        entry_task.grid(row=0, column=0,columnspan=1)
    
        
    
        entry_task.insert(0,"1")

    
        label_sims = tk.Label(entry_frame, text="Number of Iterations")
    
        label_sims.grid(row=3, column=1, columnspan=2)
    
        separate = ttk.Separator(entry_frame, orient="horizontal")
    
        separate.grid(row=5, column=0, columnspan=5, sticky="ew", pady=5)
    
        label_2 = tk.Label(entry_frame, text="Output Display:")
    
        label_2.grid(row=6, column=1, columnspan=2)
    
        button_3 = ttk.Button(entry_frame, text="Run Simulation", command=lambda:simulation_111(directory,int(entry_simul.get())))
    
        button_3.grid(row=3, column=3, columnspan=2)
    
        b_pc = ttk.Button(entry_frame, text="PDF & CDF Plots",command=lambda: make_graph(result, "Expected Value, $ ", 1, 1))
    
        b_pc.grid(row=8, column=0, sticky="ew")
    
        b_st = ttk.Button(entry_frame, text="Statistics",command=lambda: make_table(result, "Expected Value Statistics") )
    
        b_st.grid(row=8, column=2, sticky="ew")
        
        button_5 = ttk.Button(entry_frame, text="Deterministic Results", command=lambda:get_value())
    
        button_5.grid(row=7, column=0, sticky="ew",columnspan=2)
        button_6 = ttk.Button(entry_frame, text="Deterministic Results without Sank Cost", command=lambda:get_value2())
    
        button_6.grid(row=7, column=2, sticky="ew",columnspan=2)
    
    
       
    
    
        b_write_excel_risk = ttk.Button(entry_frame, text="Output Data", command=lambda: write_to_excel(name, result))
    
        b_write_excel_risk.grid(row=9, column=1, columnspan=2, sticky="ew")

       
###################################################################################################################################
if __name__ == "__main__":
    hawk = App()

    hawk.resizable(width=True, height=True)

    hawk.geometry('800x600')

    hawk.mainloop()

